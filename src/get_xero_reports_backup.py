"""
Nexa - Xero Financial Reports Backup (with Automatic Archiving)

This is the backup script for retrieving Xero financial reports. All reports are automatically 
archived with timestamps and financial year end codes.

Filename format: {report_type}_{YYYY_MM_DD}_{HHMMSS}_{FYEND}.csv
Example: balance_sheet_2025_06_30_143022_FEB26.csv

Usage:
    python get_xero_reports.py "June 2025"        # End of June 2025
    python get_xero_reports.py "2025-06-30"       # Specific date
    python get_xero_reports.py "31 December 2024" # Written format
    python get_xero_reports.py                    # Today's date
"""

import os
import sys
import pandas as pd
import re
from datetime import datetime, date
from dateutil.parser import parse
from dateutil.relativedelta import relativedelta
from config import XERO_CONFIG

# Company code mapping - add new companies as needed
COMPANY_CODES = {
    'Elenjical Solutions (Pty) Ltd': 'SA',
    'Elenjical Solutions': 'SA',
    'Elenjical Solutions MA (USD)': 'MA',
    'Elenjical Solutions MA': 'MA',
    'Elenjical Solutions Private Limited': 'IND',
    'Elenjical Solutions India': 'IND',
    'Elenjical Solutions IND': 'IND',
    # Add more companies here as you consolidate:
}

# Import Xero SDK
from xero_python.api_client import ApiClient, Configuration
from xero_python.api_client.oauth2 import OAuth2Token
from xero_python.accounting import AccountingApi
from xero_python.identity import IdentityApi

# Import FX rate reader for consolidation
from fx_reader import FXRateReader

def parse_date_input(date_input=None):
    """Parse various date input formats"""
    
    if not date_input:
        return date.today()
    
    try:
        # Handle month/year format like "June 2025"
        if len(date_input.split()) == 2 and date_input.split()[1].isdigit():
            month_name, year = date_input.split()
            # Parse to get first day of month, then get last day
            first_day = parse(f"1 {month_name} {year}").date()
            # Get last day of the month
            next_month = first_day + relativedelta(months=1)
            last_day = next_month - relativedelta(days=1)
            return last_day
        
        # Handle other formats
        parsed_date = parse(date_input).date()
        return parsed_date
        
    except Exception as e:
        print(f" Could not parse date '{date_input}': {e}")
        print(" Try formats like: 'June 2025', '2025-06-30', '30 June 2025'")
        return None

def get_financial_year_dates(target_date):
    """Get financial year start and end dates for a given date (South African FY: March to February)"""
    
    if target_date.month >= 3:  # March onwards = current financial year
        fy_start = date(target_date.year, 3, 1)
        fy_end = date(target_date.year + 1, 2, 28)  # Or 29 for leap year
        if target_date.year % 4 == 0 and (target_date.year % 100 != 0 or target_date.year % 400 == 0):
            fy_end = date(target_date.year + 1, 2, 29)
    else:  # January-February = previous financial year
        fy_start = date(target_date.year - 1, 3, 1)
        fy_end = date(target_date.year, 2, 28)
        if target_date.year % 4 == 0 and (target_date.year % 100 != 0 or target_date.year % 400 == 0):
            fy_end = date(target_date.year, 2, 29)
    
    return fy_start, fy_end

def get_fy_end_code(target_date):
    """Get financial year end code (e.g., FEB26 for FY ending Feb 2026)"""
    
    fy_start, fy_end = get_financial_year_dates(target_date)
    fy_end_code = fy_end.strftime("%b%y").upper()
    return fy_end_code

def get_company_code(org_name):
    """Get company short code from organization name"""
    
    # Direct lookup first
    if org_name in COMPANY_CODES:
        return COMPANY_CODES[org_name]
    
    # Fuzzy matching for similar names
    for company_name, code in COMPANY_CODES.items():
        if company_name.lower() in org_name.lower() or org_name.lower() in company_name.lower():
            return code
    
    # Default fallback - use first 4 chars of company name
    clean_name = ''.join(c.upper() for c in org_name if c.isalnum())
    return clean_name[:4] if clean_name else 'UNKN'

def create_archive_filename(report_type, target_date, timestamp=None, company_code=None):
    """Create archive-ready filename with company code, timestamp and FY end code"""
    
    if timestamp is None:
        timestamp = datetime.now()
    
    date_str = target_date.strftime("%Y_%m_%d")
    time_str = timestamp.strftime("%H%M%S")
    fy_code = get_fy_end_code(target_date)
    
    # Include company code if provided
    if company_code:
        filename = f"{company_code}_{report_type}_{date_str}_{time_str}_{fy_code}.xlsx"
    else:
        filename = f"{report_type}_{date_str}_{time_str}_{fy_code}.xlsx"
    
    return filename

def move_to_archive(filename):
    """Move file to archive directory (automatic archiving)"""
    
    source_path = f'xero_data/{filename}'
    archive_path = f'xero_data/archive/{filename}'
    
    try:
        # Create archive directory if it doesn't exist
        os.makedirs('xero_data/archive', exist_ok=True)
        
        if os.path.exists(source_path):
            os.rename(source_path, archive_path)
            print(f" Auto-archived: {filename}")
            return True
        else:
            print(f" File not found for archiving: {filename}")
            return False
    except Exception as e:
        print(f" Error archiving: {e}")
        return False

def setup_xero_client():
    """Set up Xero API client with proper token management and proactive refresh"""
    
    # Try to refresh token proactively to avoid expiry during execution
    try:
        import requests
        import base64
        
        print("Proactively refreshing token to ensure validity...")
        
        # Refresh the token proactively
        refresh_url = "https://identity.xero.com/connect/token"
        refresh_data = {
            'grant_type': 'refresh_token',
            'refresh_token': XERO_CONFIG['refresh_token']
        }
        
        credentials = base64.b64encode(f"{XERO_CONFIG['client_id']}:{XERO_CONFIG['client_secret']}".encode()).decode()
        headers = {
            'Authorization': f'Basic {credentials}',
            'Content-Type': 'application/x-www-form-urlencoded'
        }
        
        response = requests.post(refresh_url, data=refresh_data, headers=headers)
        
        if response.status_code == 200:
            token_data = response.json()
            
            # Update config.py with new tokens
            try:
                with open('../config/config.py', 'r', encoding='utf-8') as f:
                    config_content = f.read()
                
                # Update access_token and refresh_token in the file
                config_content = re.sub(
                    r"'access_token': '[^']*'",
                    f"'access_token': '{token_data['access_token']}'",
                    config_content
                )
                config_content = re.sub(
                    r"'refresh_token': '[^']*'",
                    f"'refresh_token': '{token_data['refresh_token']}'",
                    config_content
                )
                
                # Write updated config back to file
                with open('../config/config.py', 'w', encoding='utf-8') as f:
                    f.write(config_content)
                
                # Update the XERO_CONFIG dict in memory
                XERO_CONFIG['access_token'] = token_data['access_token']
                XERO_CONFIG['refresh_token'] = token_data['refresh_token']
                
                print("Token proactively refreshed and saved - expires in 1800 seconds")
                
            except Exception as e:
                print(f"Token refreshed but failed to save to ../config/config.py: {e}")
        else:
            print("Using existing token...")
            
    except Exception as e:
        print(f"Proactive token refresh failed, will rely on automatic refresh: {e}")
    
    # Create API client
    api_client = ApiClient(
        Configuration(
            debug=False,
            oauth2_token=OAuth2Token(
                client_id=XERO_CONFIG['client_id'],
                client_secret=XERO_CONFIG['client_secret']
            ),
        ),
        pool_threads=1,
    )
    
    # Set up token
    token_set = {
        "access_token": XERO_CONFIG['access_token'],
        "token_type": "Bearer",
        "refresh_token": XERO_CONFIG['refresh_token'],
        "expires_in": 1800,  # 30 minutes
        "scope": XERO_CONFIG['scopes']
    }
    
    # Set up token getter/setter functions
    @api_client.oauth2_token_getter
    def obtain_xero_oauth2_token():
        return token_set
    
    @api_client.oauth2_token_saver
    def store_xero_oauth2_token(token):
        token_set.update(token)
        
        # Automatically update config.py with new tokens
        try:
            # Read current config file
            with open('../config/config.py', 'r', encoding='utf-8') as f:
                config_content = f.read()
            
            # Update access_token and refresh_token in the file
            if 'access_token' in token:
                config_content = re.sub(
                    r"'access_token': '[^']*'",
                    f"'access_token': '{token['access_token']}'",
                    config_content
                )
            
            if 'refresh_token' in token:
                config_content = re.sub(
                    r"'refresh_token': '[^']*'",
                    f"'refresh_token': '{token['refresh_token']}'",
                    config_content
                )
            
            # Write updated config back
            with open('../config/config.py', 'w', encoding='utf-8') as f:
                f.write(config_content)
                
            print(f"🔄 Token automatically refreshed and saved - expires in {token.get('expires_in', 'unknown')} seconds")
            
        except Exception as e:
            print(f"⚠️ Token refreshed but failed to save to ../config/config.py: {e}")
    
    # Set the token
    api_client.set_oauth2_token(token_set)
    
    return api_client

def get_organization_info(api_client, tenant_id):
    """Get organization information including base currency"""
    
    accounting_api = AccountingApi(api_client)
    
    try:
        orgs = accounting_api.get_organisations(xero_tenant_id=tenant_id)
        if orgs and orgs.organisations:
            org = orgs.organisations[0]
            return {
                'name': org.name,
                'base_currency': org.base_currency,
                'country_code': org.country_code,
                'tax_number': org.tax_number
            }
    except Exception as e:
        print(f" Could not get organization info: {e}")
        return {'base_currency': 'ZAR'}  # Default to ZAR for South Africa

def format_accounting(amount):
    """Format amount in accounting format without currency symbol"""
    
    # Format with thousand separators and 2 decimal places
    if amount == 0:
        return "0.00"
    elif amount > 0:
        return f"{amount:,.2f}"
    else:
        return f"({abs(amount):,.2f})"  # Negative amounts in parentheses (accounting style)

def format_currency_display(amount, currency_code='ZAR'):
    """Format amount with currency symbol for display purposes only"""
    
    # Currency symbols mapping
    currency_symbols = {
        'ZAR': 'R',
        'USD': '$',
        'GBP': '',
        'EUR': '',
        'CurrencyCode.ZAR': 'R',
        'CurrencyCode.USD': '$',
        'CurrencyCode.GBP': '',
        'CurrencyCode.EUR': '',
    }
    
    # Clean currency code
    clean_code = currency_code.replace('CurrencyCode.', '') if currency_code else 'ZAR'
    symbol = currency_symbols.get(clean_code, f'{clean_code} ')
    
    # Format the amount with currency for display
    if amount == 0:
        return f"{symbol}0.00"
    elif amount > 0:
        return f"{symbol}{amount:,.2f}"
    else:
        return f"-{symbol}{abs(amount):,.2f}"

def format_data_for_export(data, data_type='general'):
    """Centrally format data with proper date formats (dd/mm/yyyy) and keep numbers as raw values for Excel"""
    
    if isinstance(data, list):
        # Handle list of dictionaries (like invoices)
        formatted_data = []
        
        for item in data:
            formatted_item = item.copy()
            
            # Format all date fields to dd/mm/yyyy
            date_fields = ['date', 'due_date', 'fully_paid_on_date', 'expected_payment_date', 'planned_payment_date', 'updated_date_utc']
            
            for field in date_fields:
                if field in formatted_item and formatted_item[field]:
                    try:
                        # Parse the date string and reformat to dd/mm/yyyy
                        if formatted_item[field] != '':
                            date_obj = pd.to_datetime(formatted_item[field])
                            # Special handling for datetime fields (include time)
                            if 'utc' in field.lower() or 'updated' in field.lower():
                                formatted_item[field] = date_obj.strftime('%d/%m/%Y %H:%M')
                            else:
                                formatted_item[field] = date_obj.strftime('%d/%m/%Y')
                    except:
                        # Keep original value if parsing fails
                        pass
            
            # Keep numeric fields as raw numbers - Excel formatting will handle the display
            formatted_data.append(formatted_item)
        
        return formatted_data
    
    else:
        # Handle DataFrame (like financial reports)
        return data

def export_to_excel_with_formatting(df, filename, sheet_name, data_type='general'):
    """Central Excel export function with consistent formatting for all data types"""
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Define formatting standards
        accounting_format = '#,##0.00_);(#,##0.00)'  # No currency, negative in parentheses
        
        # Define column types for different data formats
        financial_columns = {
            'balance', 'sub_total', 'total_tax', 'total', 'amount_due', 'amount_paid', 
            'amount_credited', 'currency_rate', 'first_line_quantity', 'first_line_unit_amount', 
            'first_line_line_amount'
        }
        
        # Apply formatting based on column names and data types
        from openpyxl.utils import get_column_letter
        
        for idx, col in enumerate(df.columns, 1):
            col_letter = get_column_letter(idx)
            
            # Handle YTD Total column specially for formulas
            if col == 'ytd_total' and data_type == 'financial':
                # Find the month columns to create proper Excel formulas
                month_cols = [c for c in df.columns if c.startswith('2025-')]
                
                for row in range(2, len(df) + 2):  # Start from row 2 (after header)
                    cell = worksheet[f'{col_letter}{row}']
                    
                    # Create Excel formula with proper range notation
                    if month_cols:
                        # Find first and last month column letters for range
                        first_month_idx = df.columns.get_loc(month_cols[0]) + 1
                        last_month_idx = df.columns.get_loc(month_cols[-1]) + 1
                        
                        first_col_letter = get_column_letter(first_month_idx)
                        last_col_letter = get_column_letter(last_month_idx)
                        
                        # Set Excel formula using range notation (e.g., =SUM(D5:G5))
                        formula = f"=SUM({first_col_letter}{row}:{last_col_letter}{row})"
                        cell.value = formula
                    
                    # Apply accounting format
                    cell.number_format = accounting_format
            
            # Handle monthly columns specially for totals row formulas
            elif col.startswith('2025-') and data_type == 'financial':
                # Check if this is the totals row (last row)
                last_row = len(df) + 1  # +1 because Excel is 1-indexed
                
                for row in range(2, len(df) + 2):  # Start from row 2 (after header)
                    cell = worksheet[f'{col_letter}{row}']
                    
                    # If this is the totals row, add column sum formula
                    if row == last_row:
                        # Create column sum formula (exclude totals row itself)
                        formula = f"=SUM({col_letter}2:{col_letter}{last_row-1})"
                        cell.value = formula
                    
                    # Apply accounting format
                    cell.number_format = accounting_format
            
            # Apply accounting format to other financial columns
            elif col.lower() in financial_columns or df[col].dtype in ['int64', 'float64']:
                for row in range(2, len(df) + 2):  # Start from row 2 (after header)
                    cell = worksheet[f'{col_letter}{row}']
                    cell.number_format = accounting_format

def create_sheet_name(report_type, target_date, company_code=None):
    """Create sheet name with company code, report type, month/year, and FYE"""
    
    # Format the month and year
    month_year = target_date.strftime("%B%Y")  # e.g., "June2025"
    
    # Get FYE code
    fy_code = get_fy_end_code(target_date)
    
    # Map report types to sheet prefixes (keeping under 31 char Excel limit)
    sheet_prefixes = {
        'balance_sheet': 'BS',
        'profit_and_loss': 'P&L', 
        'monthly_pnl_ytd': 'MonthlyPL',
        'trial_balance': 'TB',
        'chart_of_accounts': 'COA',
        'invoices': 'INV_YTD'
    }
    
    prefix = sheet_prefixes.get(report_type, report_type.upper())
    
    # Include company code if provided
    if company_code:
        return f"{company_code}_{prefix}_{month_year}_{fy_code}"
    else:
        return f"{prefix}_{month_year}_{fy_code}"

def parse_report_rows(rows, base_currency='ZAR', section_name="", level=0):
    """Recursively parse report rows to extract account data with currency"""
    accounts = []
    
    for row in rows:
        if hasattr(row, 'title') and row.title:
            # This is a section header
            current_section = row.title
            if hasattr(row, 'rows') and row.rows:
                accounts.extend(parse_report_rows(row.rows, base_currency, current_section, level + 1))
        
        elif hasattr(row, 'cells') and row.cells:
            # This is a data row
            cells = row.cells
            if len(cells) >= 2:
                account_name = cells[0].value if cells[0].value else ""
                balance_value = cells[1].value if len(cells) > 1 and cells[1].value else "0"
                
                # Skip empty rows, headers, and total/subtotal rows
                if account_name and account_name.strip() and balance_value and str(balance_value) != "0":
                    account_name_lower = account_name.strip().lower()
                    
                    # Skip total/subtotal rows and summary lines
                    skip_terms = [
                        'total ', 'subtotal', 'net profit', 'net loss', 'gross profit', 'gross loss',
                        'total assets', 'total liabilities', 'total equity', 'total income', 'total expenses',
                        'total current assets', 'total non-current assets', 'total current liabilities',
                        'total non-current liabilities', 'net assets', 'retained earnings', 'total revenue',
                        'total cost of sales', 'total operating expenses', 'total other income',
                        'total other expenses', 'earnings before', 'ebitda', 'ebit'
                    ]
                    
                    # Check if this is a total/summary row to skip
                    is_total_row = any(term in account_name_lower for term in skip_terms)
                    
                    if not is_total_row:
                        try:
                            # Clean up the balance value
                            balance_str = str(balance_value).replace(',', '').replace('(', '-').replace(')', '')
                            balance = float(balance_str) if balance_str and balance_str != "" else 0.0
                            
                            accounts.append({
                                'section': section_name or "Unknown",
                                'account_name': account_name.strip(),
                                'balance': balance,  # Keep raw number for Excel
                                'currency': base_currency,
                                'level': level
                            })
                        except (ValueError, AttributeError):
                            continue
        
        # Check for nested rows
        if hasattr(row, 'rows') and row.rows and not hasattr(row, 'title'):
            accounts.extend(parse_report_rows(row.rows, base_currency, section_name, level))
    
    return accounts

def get_all_invoices(accounting_api, tenant_id, base_currency, target_date=None, fy_start=None, page_size=100):
    """Fetch invoices for the financial year to date with pagination"""
    all_invoices = []
    page = 1
    
    # Determine date range for filtering
    if target_date and fy_start:
        print(f"   Fetching invoices for FY period: {fy_start.strftime('%d %b %Y')} to {target_date.strftime('%d %b %Y')} (page size: {page_size})...")
    else:
        print(f"   Fetching invoices (page size: {page_size})...")
    
    while True:
        try:
            # Build parameters for invoice query
            params = {
                'xero_tenant_id': tenant_id,
                'page': page,
                'page_size': page_size,
                'unitdp': 4,  # 4 decimal places
                'include_archived': True,  # Include archived invoices
                'summary_only': False  # Get full details
            }
            
            # If both target_date and fy_start are provided, filter for financial year to date
            if target_date and fy_start:
                # Filter for invoices within the financial year
                where_clause = f'Date >= DateTime({fy_start.year}, {fy_start.month}, {fy_start.day}) AND Date <= DateTime({target_date.year}, {target_date.month}, {target_date.day})'
                params['where'] = where_clause
            elif target_date:
                # Fallback: filter up to target date only
                where_clause = f'Date <= DateTime({target_date.year}, {target_date.month}, {target_date.day})'
                params['where'] = where_clause
            
            response = accounting_api.get_invoices(**params)
            
            if response.invoices:
                print(f"   Retrieved {len(response.invoices)} invoices from page {page}")
                
                # Process each invoice
                for invoice in response.invoices:
                    invoice_data = {
                        'invoice_id': invoice.invoice_id,
                        'invoice_number': invoice.invoice_number or '',
                        'reference': invoice.reference or '',
                        'type': str(invoice.type) if invoice.type else '',
                        'contact_name': invoice.contact.name if invoice.contact else '',
                        'contact_id': invoice.contact.contact_id if invoice.contact else '',
                        'date': invoice.date.strftime('%Y-%m-%d') if invoice.date else '',
                        'due_date': invoice.due_date.strftime('%Y-%m-%d') if invoice.due_date else '',
                        'status': str(invoice.status) if invoice.status else '',
                        'line_amount_types': str(invoice.line_amount_types) if invoice.line_amount_types else '',
                        'sub_total': float(invoice.sub_total) if invoice.sub_total else 0.0,
                        'total_tax': float(invoice.total_tax) if invoice.total_tax else 0.0,
                        'total': float(invoice.total) if invoice.total else 0.0,
                        'amount_due': float(invoice.amount_due) if invoice.amount_due else 0.0,
                        'amount_paid': float(invoice.amount_paid) if invoice.amount_paid else 0.0,
                        'amount_credited': float(invoice.amount_credited) if invoice.amount_credited else 0.0,
                        'currency_code': str(invoice.currency_code) if invoice.currency_code else base_currency,
                        'currency_rate': float(invoice.currency_rate) if invoice.currency_rate else 1.0,
                        'fully_paid_on_date': invoice.fully_paid_on_date.strftime('%Y-%m-%d') if invoice.fully_paid_on_date else '',
                        'expected_payment_date': invoice.expected_payment_date.strftime('%Y-%m-%d') if invoice.expected_payment_date else '',
                        'planned_payment_date': invoice.planned_payment_date.strftime('%Y-%m-%d') if invoice.planned_payment_date else '',
                        'updated_date_utc': invoice.updated_date_utc.strftime('%Y-%m-%d %H:%M:%S') if invoice.updated_date_utc else '',
                        'sent_to_contact': invoice.sent_to_contact if hasattr(invoice, 'sent_to_contact') else False,
                        'has_attachments': invoice.has_attachments if hasattr(invoice, 'has_attachments') else False
                    }
                    
                    # Add line items summary if available
                    if hasattr(invoice, 'line_items') and invoice.line_items:
                        line_count = len(invoice.line_items)
                        invoice_data['line_item_count'] = line_count
                        
                        # Get first line item description for reference
                        first_line = invoice.line_items[0]
                        invoice_data['first_line_description'] = first_line.description[:100] if first_line.description else ''
                        invoice_data['first_line_quantity'] = float(first_line.quantity) if first_line.quantity else 0.0
                        invoice_data['first_line_unit_amount'] = float(first_line.unit_amount) if first_line.unit_amount else 0.0
                        invoice_data['first_line_line_amount'] = float(first_line.line_amount) if first_line.line_amount else 0.0
                        invoice_data['first_line_account_code'] = first_line.account_code if first_line.account_code else ''
                    else:
                        invoice_data['line_item_count'] = 0
                        invoice_data['first_line_description'] = ''
                        invoice_data['first_line_quantity'] = 0.0
                        invoice_data['first_line_unit_amount'] = 0.0
                        invoice_data['first_line_line_amount'] = 0.0
                        invoice_data['first_line_account_code'] = ''
                    
                    all_invoices.append(invoice_data)
                
                # Check if we've reached the end
                if len(response.invoices) < page_size:
                    break
                
                page += 1
            else:
                break
                
        except Exception as e:
            print(f"   Error fetching invoices page {page}: {e}")
            break
    
    print(f"   Total invoices retrieved: {len(all_invoices)}")
    return all_invoices

def get_monthly_pnl_with_ytd(accounting_api, tenant_id, base_currency, target_date, fy_start):
    """Generate monthly P&L breakdown with YTD calculations"""
    
    print(f"   Building monthly P&L from {fy_start.strftime('%b %Y')} to {target_date.strftime('%b %Y')}...")
    
    # Calculate months to process
    months_data = []
    current_month = fy_start.replace(day=1)  # Start of financial year month
    
    while current_month <= target_date:
        # Calculate month end date
        if current_month.month == 12:
            month_end = current_month.replace(year=current_month.year + 1, month=1, day=1) - relativedelta(days=1)
        else:
            month_end = current_month.replace(month=current_month.month + 1, day=1) - relativedelta(days=1)
        
        # Don't go beyond target date
        if month_end > target_date:
            month_end = target_date
            
        print(f"     Fetching {current_month.strftime('%b %Y')} P&L...")
        
        try:
            # Get P&L for this month
            pnl_response = accounting_api.get_report_profit_and_loss(
                xero_tenant_id=tenant_id,
                from_date=current_month,
                to_date=month_end
            )
            
            if pnl_response and pnl_response.reports:
                report = pnl_response.reports[0]
                month_accounts = parse_report_rows(report.rows, base_currency)
                
                months_data.append({
                    'month': current_month.strftime('%Y-%m'),
                    'month_name': current_month.strftime('%b %Y'),
                    'start_date': current_month,
                    'end_date': month_end,
                    'accounts': month_accounts
                })
                
                print(f"       ✅ {len(month_accounts)} accounts processed")
            else:
                print(f"       ⚠️ No P&L data available")
                
        except Exception as e:
            print(f"       ❌ Error fetching {current_month.strftime('%b %Y')}: {e}")
        
        # Move to next month
        if current_month.month == 12:
            current_month = current_month.replace(year=current_month.year + 1, month=1)
        else:
            current_month = current_month.replace(month=current_month.month + 1)
    
    if not months_data:
        print(f"   No monthly P&L data available")
        return None
        
    # Build consolidated monthly P&L with YTD
    print(f"   Building consolidated monthly P&L with YTD calculations...")
    
    # Get all unique accounts across all months
    all_account_names = set()
    all_sections = set()
    
    for month_data in months_data:
        for account in month_data['accounts']:
            all_account_names.add(account['account_name'])
            all_sections.add(account['section'])
    
    # Create consolidated structure
    consolidated_pnl = []
    
    for section in sorted(all_sections):
        section_accounts = [acc for month_data in months_data 
                          for acc in month_data['accounts'] 
                          if acc['section'] == section]
        
        if section_accounts:
            # Get unique accounts in this section
            section_account_names = set(acc['account_name'] for acc in section_accounts)
            
            for account_name in sorted(section_account_names):
                monthly_row = {
                    'section': section,
                    'account_name': account_name,
                    'currency': base_currency
                }
                
                                # Add monthly columns and prepare for YTD formula
                month_columns_for_formula = []
                
                for month_data in months_data:
                    # Find this account in this month
                    month_balance = 0.0
                    for account in month_data['accounts']:
                        if (account['account_name'] == account_name and 
                            account['section'] == section):
                            month_balance = account['balance']
                            break
                    
                    monthly_row[month_data['month']] = month_balance
                    monthly_row[f"{month_data['month_name']}"] = month_balance
                    month_columns_for_formula.append(month_data['month'])

                # Store placeholder for YTD - will be converted to Excel formula during export
                monthly_row['ytd_total'] = 0  # Placeholder value, will be replaced with Excel formula
                consolidated_pnl.append(monthly_row)
    
    print(f"   ✅ Monthly P&L consolidated: {len(consolidated_pnl)} account rows across {len(months_data)} months")
    
    # Add totals row at the bottom
    if consolidated_pnl:
        # Create totals row
        totals_row = {
            'section': 'TOTALS',
            'account_name': 'TOTAL',
            'currency': base_currency
        }
        
        # Add placeholders for each month and YTD (will be converted to Excel formulas during export)
        for month_data in months_data:
            totals_row[month_data['month']] = 0  # Placeholder for Excel formula
        
        totals_row['ytd_total'] = 0  # Placeholder for Excel formula
        
        consolidated_pnl.append(totals_row)
        print(f"   ✅ Added totals row for column summation")
    
    return {
        'months': [m['month_name'] for m in months_data],
        'data': consolidated_pnl,
        'period_start': fy_start,
        'period_end': target_date
    }

def generate_reports_single_company(target_date, run_timestamp, skip_archiving=False):
    """Generate reports for a single company (existing logic)"""
    
    # Create archive-ready filenames
    fy_code = get_fy_end_code(target_date)
    month_year = target_date.strftime("%B %Y")
    
    # Check if a specific company was requested
    requested_company = sys.argv[2].upper() if len(sys.argv) > 2 else None
    if requested_company:
        print(f"Fetching Xero Financial Reports - {requested_company} (Auto-Archive)")
    else:
        print(f"Fetching Xero Financial Reports (Auto-Archive)")
    print(f"Report Date: {target_date.strftime('%d %B %Y')}")
    print(f"Financial Year End: {fy_code}")
    print(f"Timestamp: {run_timestamp.strftime('%H:%M:%S')}")
    print("=" * 70)
    
    # Create output directories
    os.makedirs('xero_data', exist_ok=True)
    os.makedirs('xero_data/archive', exist_ok=True)
    
    # STEP 1: Archive any existing old files first (only if not skipping)
    if not skip_archiving:
        print(f"\nCHECKING FOR EXISTING FILES TO ARCHIVE")
        print("-" * 50)
        
        import glob
        existing_files = glob.glob('xero_data/*.xlsx') + glob.glob('xero_data/*.csv')
        archived_old_files = 0
        
        for filepath in existing_files:
            filename = os.path.basename(filepath)
            # Skip README files
            if filename.lower().startswith('readme'):
                continue
                
            try:
                # Move existing file to archive with timestamp
                timestamp_str = datetime.now().strftime("%H%M%S")
                archive_name = f"archive_{timestamp_str}_{filename}"
                archive_path = f'xero_data/archive/{archive_name}'
                
                os.rename(filepath, archive_path)
                print(f"Archived existing file: {filename} -> {archive_name}")
                archived_old_files += 1
                
            except Exception as e:
                print(f"Could not archive {filename}: {e}")
        
        if archived_old_files == 0:
            print("No existing files found to archive")
        else:
            print(f"Archived {archived_old_files} existing files")
    else:
        print(f"\nSkipping archiving (multi-company mode)")
        print("-" * 40)
    
    # Set up client
    try:
        api_client = setup_xero_client()
        identity_api = IdentityApi(api_client)
        accounting_api = AccountingApi(api_client)
        
        # Get tenant ID
        connections = identity_api.get_connections()
        if not connections:
            print("No Xero connections found")
            return
        
        # Filter connections based on target company if specified
        target_company_code = sys.argv[2].upper() if len(sys.argv) > 2 else None
        target_connection = None
        
        if target_company_code and target_company_code in ['SA', 'MA', 'IND']:
            # Find connection matching the target company code
            for conn in connections:
                if get_company_code(conn.tenant_name) == target_company_code:
                    target_connection = conn
                    break
            
            if not target_connection:
                print(f"No connection found for company code: {target_company_code}")
                print("Available companies:")
                for conn in connections:
                    code = get_company_code(conn.tenant_name)
                    print(f"  {code}: {conn.tenant_name}")
                return
            
            tenant_id = target_connection.tenant_id
            tenant_name = target_connection.tenant_name
            print(f"Connected to: {tenant_name} (Target: {target_company_code})")
        else:
            # Use first available connection
            target_connection = connections[0]
            tenant_id = target_connection.tenant_id
            tenant_name = target_connection.tenant_name
            if target_company_code:
                print(f"Connected to: {tenant_name} (Invalid target: {target_company_code}, using auto-select)")
            else:
                print(f"Connected to: {tenant_name} (Auto-selected)")
        
        # Get organization info for base currency
        org_info = get_organization_info(api_client, tenant_id)
        base_currency_raw = org_info.get('base_currency', 'ZAR')
        # Clean up currency code - handle both string and CurrencyCode object
        if hasattr(base_currency_raw, 'value'):
            base_currency = base_currency_raw.value
        else:
            base_currency = str(base_currency_raw).replace('CurrencyCode.', '') if base_currency_raw else 'ZAR'
        
        org_name = org_info.get('name', tenant_name)
        company_code = get_company_code(org_name)
        
        print(f"Base Currency: {base_currency}")
        print(f"Organization: {org_name}")
        print(f"Company Code: {company_code}")
        
    except Exception as e:
        print(f"Connection failed: {e}")
        return
    
    # Financial year context
    fy_start, fy_end = get_financial_year_dates(target_date)
    days_into_fy = (target_date - fy_start).days
    
    print(f"\nFinancial Year: {fy_start.strftime('%d %b %Y')} to {fy_end.strftime('%d %b %Y')}")
    print(f"   Days into FY: {days_into_fy}")
    print(f"   Report as at: {target_date.strftime('%d %B %Y')}")
    
    # Track generated files for automatic archiving
    generated_files = []
    
    # 1. Get Balance Sheet
    print(f"\nGetting Balance Sheet as at {target_date.strftime('%d %B %Y')}...")
    try:
        balance_sheet = accounting_api.get_report_balance_sheet(
            xero_tenant_id=tenant_id,
            date=target_date
        )
        
        if balance_sheet and balance_sheet.reports:
            report = balance_sheet.reports[0]
            print(f"Retrieved Balance Sheet: {report.report_name}")
            
            # Parse balance sheet data
            bs_accounts = parse_report_rows(report.rows, base_currency)
            
            if bs_accounts:
                # Create filename with company code, timestamp and FY code
                filename = create_archive_filename('balance_sheet', target_date, run_timestamp, company_code)
                
                # Export balance sheet as current file (exclude balance_raw column)
                df_bs = pd.DataFrame(bs_accounts)
                export_columns = ['section', 'account_name', 'balance', 'currency', 'level']
                sheet_name = create_sheet_name('balance_sheet', target_date, company_code)
                export_to_excel_with_formatting(df_bs[export_columns], f'xero_data/{filename}', sheet_name, 'financial')
                generated_files.append(filename)
                
                # Calculate and show summary 
                total_assets = sum([acc['balance'] for acc in bs_accounts if any(word in acc['section'].lower() for word in ['asset', 'bank'])])
                total_liabilities = sum([acc['balance'] for acc in bs_accounts if 'liabilit' in acc['section'].lower()])
                total_equity = sum([acc['balance'] for acc in bs_accounts if 'equity' in acc['section'].lower()])
                
                print(f"   Assets: {format_currency_display(total_assets, base_currency)}")
                print(f"   Liabilities: {format_currency_display(total_liabilities, base_currency)}")
                print(f"   Equity: {format_currency_display(total_equity, base_currency)}")
                
    except Exception as e:
        print(f"Error getting balance sheet: {e}")
    
    # 2. Get Monthly P&L with YTD
    print(f"\n Getting Monthly P&L with YTD ({fy_start.strftime('%b %Y')} to {target_date.strftime('%b %Y')})...")
    try:
        monthly_pnl_data = get_monthly_pnl_with_ytd(accounting_api, tenant_id, base_currency, target_date, fy_start)
        
        if monthly_pnl_data and monthly_pnl_data['data']:
            print(f" Retrieved Monthly P&L: {len(monthly_pnl_data['data'])} account rows across {len(monthly_pnl_data['months'])} months")
            
            # Create filename with company code, timestamp and FY code
            filename = create_archive_filename('profit_and_loss', target_date, run_timestamp, company_code)
            
            # Create DataFrame with monthly P&L data
            df_monthly_pnl = pd.DataFrame(monthly_pnl_data['data'])
            
            # Define columns for export - dynamic based on available months
            base_columns = ['section', 'account_name', 'currency']
            month_columns = [month_data['month'] for month_data in monthly_pnl_data.get('months_detailed', [])]
            if not month_columns:
                # Fallback: try to find month columns in the data
                month_columns = [col for col in df_monthly_pnl.columns if col.startswith('2025-')]
            
            export_columns = base_columns + month_columns + ['ytd_total']
            
            # Filter to only include columns that exist in the DataFrame
            export_columns = [col for col in export_columns if col in df_monthly_pnl.columns]
            
            # Create sheet name
            sheet_name = create_sheet_name('profit_and_loss', target_date, company_code)
            
            # Export with centralized formatting for monthly P&L
            export_to_excel_with_formatting(df_monthly_pnl[export_columns], f'xero_data/{filename}', sheet_name, 'financial')
            generated_files.append(filename)
            
            # Show summary for each month and YTD
            months_list = monthly_pnl_data['months']
            print(f"   Monthly breakdown: {', '.join(months_list)}")
            
            # Calculate total revenue and expenses for each month and YTD (manually calculate since ytd_total will be Excel formulas)
            revenue_accounts = [acc for acc in monthly_pnl_data['data'] if 'income' in acc['section'].lower()]
            if revenue_accounts:
                # Calculate YTD manually by summing all monthly columns
                month_cols = [c for c in monthly_pnl_data['data'][0].keys() if c.startswith('2025-')]
                ytd_revenue = sum([sum([acc.get(month_col, 0) for month_col in month_cols]) for acc in revenue_accounts])
                print(f"   YTD Revenue: {format_currency_display(ytd_revenue, base_currency)}")
                
                # Calculate YTD net profit
                ytd_net_profit = sum([sum([acc.get(month_col, 0) for month_col in month_cols]) for acc in monthly_pnl_data['data']])
                print(f"   YTD Net Profit: {format_currency_display(ytd_net_profit, base_currency)}")
                
                if ytd_revenue > 0:
                    ytd_net_margin = (ytd_net_profit / ytd_revenue) * 100
                    print(f"   YTD Net Margin: {ytd_net_margin:.1f}%")
            
        else:
            print(f"   No monthly P&L data found for the specified period")
            
    except Exception as e:
        print(f" Error getting monthly P&L: {e}")
    
    # 3. Get Trial Balance
    print(f"\n Getting Trial Balance as at {target_date.strftime('%d %B %Y')}...")
    try:
        trial_balance = accounting_api.get_report_trial_balance(
            xero_tenant_id=tenant_id,
            date=target_date
        )
        
        if trial_balance and trial_balance.reports:
            report = trial_balance.reports[0]
            print(f" Retrieved Trial Balance: {report.report_name}")
            
            # Parse trial balance data
            tb_accounts = parse_report_rows(report.rows, base_currency)
            
            if tb_accounts:
                # Create filename with company code, timestamp and FY code
                filename = create_archive_filename('trial_balance', target_date, run_timestamp, company_code)
                
                # Export trial balance as current file (exclude balance_raw column)
                df_tb = pd.DataFrame(tb_accounts)
                export_columns = ['section', 'account_name', 'balance', 'currency', 'level']
                sheet_name = create_sheet_name('trial_balance', target_date, company_code)
                export_to_excel_with_formatting(df_tb[export_columns], f'xero_data/{filename}', sheet_name, 'financial')
                generated_files.append(filename)
                
                print(f"   Total accounts: {len(tb_accounts)}")
                
    except Exception as e:
        print(f" Error getting trial balance: {e}")
    
    # 4. Get Chart of Accounts
    print(f"\n Getting Chart of Accounts...")
    try:
        accounts_response = accounting_api.get_accounts(
            xero_tenant_id=tenant_id
        )
        
        if accounts_response.accounts:
            print(f" Retrieved Chart of Accounts")
            
            # Process accounts data
            accounts = []
            for account in accounts_response.accounts:
                account_data = {
                    'account_id': account.account_id,
                    'code': account.code,
                    'name': account.name,
                    'type': str(account.type),
                    'tax_type': account.tax_type,
                    'status': str(account.status),
                    'description': account.description,
                    'class': getattr(account, 'class', None),
                    'system_account': getattr(account, 'system_account', None),
                    'reporting_code': getattr(account, 'reporting_code', None),
                    'reporting_code_name': getattr(account, 'reporting_code_name', None),
                    'updated_date_utc': account.updated_date_utc.isoformat() if account.updated_date_utc else None
                }
                accounts.append(account_data)
            
            if accounts:
                # Create filename with company code, timestamp and FY code
                filename = create_archive_filename('chart_of_accounts', target_date, run_timestamp, company_code)
                
                # Export chart of accounts as current file
                df_accounts = pd.DataFrame(accounts)
                export_columns = ['code', 'name', 'type', 'tax_type', 'status', 'description', 'class', 'system_account', 'reporting_code', 'reporting_code_name']
                sheet_name = create_sheet_name('chart_of_accounts', target_date, company_code)
                export_to_excel_with_formatting(df_accounts[export_columns], f'xero_data/{filename}', sheet_name, 'general')
                generated_files.append(filename)
                
                # Analyze account types
                account_types = {}
                for account in accounts:
                    acc_type = account['type']
                    if acc_type not in account_types:
                        account_types[acc_type] = 0
                    account_types[acc_type] += 1
                
                print(f"   Total accounts: {len(accounts)}")
                print(f"   Account types: {len(account_types)}")
                
    except Exception as e:
        print(f" Error getting chart of accounts: {e}")
    
    # 5. Get All Invoices (Year to Date)
    print(f"\n Getting All Invoices for Financial Year to Date...")
    try:
        all_invoices = get_all_invoices(accounting_api, tenant_id, base_currency, target_date, fy_start)
        
        if all_invoices:
            print(f" Retrieved {len(all_invoices)} invoices")
            
            # Create filename with company code, timestamp and FY code
            filename = create_archive_filename('invoices_ytd', target_date, run_timestamp, company_code)
            
            # Format invoice data for proper date display (dd/mm/yyyy)
            formatted_invoices = format_data_for_export(all_invoices, 'invoices')
            
            # Create DataFrame with formatted invoice data
            df_invoices = pd.DataFrame(formatted_invoices)
            
            # Define columns to include in export (with accounting formatting for amounts)
            export_columns = [
                'invoice_number', 'reference', 'type', 'contact_name', 'date', 'due_date', 
                'status', 'sub_total', 'total_tax', 'total', 'amount_due', 'amount_paid', 
                'amount_credited', 'currency_code', 'currency_rate', 'fully_paid_on_date',
                'line_item_count', 'first_line_description', 'first_line_quantity', 
                'first_line_unit_amount', 'first_line_line_amount', 'first_line_account_code',
                'updated_date_utc'
            ]
            
            # Create sheet name
            sheet_name = create_sheet_name('invoices', target_date, company_code)
            
            # Export with centralized formatting for invoices
            export_to_excel_with_formatting(df_invoices[export_columns], f'xero_data/{filename}', sheet_name, 'invoices')
            generated_files.append(filename)
            
            # Calculate summary statistics
            total_invoiced = sum([inv['total'] for inv in all_invoices])
            total_outstanding = sum([inv['amount_due'] for inv in all_invoices])
            total_paid = sum([inv['amount_paid'] for inv in all_invoices])
            
            # Count by status
            status_counts = {}
            for invoice in all_invoices:
                status = invoice['status']
                if status not in status_counts:
                    status_counts[status] = 0
                status_counts[status] += 1
            
            # Count by type
            type_counts = {}
            for invoice in all_invoices:
                inv_type = invoice['type']
                if inv_type not in type_counts:
                    type_counts[inv_type] = 0
                type_counts[inv_type] += 1
            
            print(f"   Total Invoiced: {format_currency_display(total_invoiced, base_currency)}")
            print(f"   Total Outstanding: {format_currency_display(total_outstanding, base_currency)}")
            print(f"   Total Paid: {format_currency_display(total_paid, base_currency)}")
            print(f"   Status Breakdown: {status_counts}")
            print(f"   Type Breakdown: {type_counts}")
            
        else:
            print(f"   No invoices found for the specified date range")
            
    except Exception as e:
        print(f" Error getting invoices: {e}")
    
    # New files remain as current versions (no archiving needed)
    
    # Summary
    print(f"\n SUMMARY - {target_date.strftime('%d %B %Y')}")
    print("-" * 50)
    print(f"Organization: {org_info.get('name', tenant_name)}")
    print(f"Report Date: {target_date.strftime('%d %B %Y')}")
    print(f"Financial Year: {fy_code}")
    print(f"Timestamp: {run_timestamp.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Reports Generated: {len(generated_files)}")
    
    print(f"\n Current reports saved to: xero_data/")
    
    # List current files
    if generated_files:
        print(f"\n Current Files:")
        for filename in generated_files:
            print(f"    {filename}")
    
    print(f"\n Workflow: Old files archived -> New files become current")
    print(f" Archive contains historical versions with timestamps")

def convert_amount_to_zar(amount, from_currency, fx_reader, entity_name=""):
    """Convert amount to ZAR using FX rates"""
    if not amount or amount == 0:
        return 0.0, 1.0  # Return converted amount and rate used
    
    if from_currency == 'ZAR' or from_currency == 'CurrencyCode.ZAR':
        return float(amount), 1.0
    
    # Clean currency code
    clean_currency = from_currency.replace('CurrencyCode.', '') if from_currency else 'ZAR'
    
    # Get FX rate
    fx_rate = fx_reader.get_fx_rate(clean_currency, 'ZAR')
    
    if fx_rate:
        converted_amount = float(amount) * fx_rate
        return converted_amount, fx_rate
    else:
        print(f"⚠️ No FX rate found for {clean_currency} to ZAR for {entity_name}. Using original amount.")
        return float(amount), 1.0

def create_consolidated_dataframe(original_df, base_currency, fx_reader, entity_name, report_type):
    """Create consolidated DataFrame with entity tracking and ZAR conversion"""
    
    if original_df.empty:
        return original_df, {}
    
    # Create a copy of the original DataFrame
    consolidated_df = original_df.copy()
    
    # Add entity column at the beginning
    consolidated_df.insert(0, 'Entity', entity_name)
    
    # Clean base currency code
    clean_base_currency = base_currency.replace('CurrencyCode.', '') if base_currency else 'ZAR'
    
    # Find amount columns to convert
    amount_columns = []
    for col in consolidated_df.columns:
        if any(keyword in col.lower() for keyword in ['amount', 'balance', 'total', 'value', 'ytd', 'mar ', 'apr ', 'may ', 'jun ', 'jul ', 'aug ', 'sep ', 'oct ', 'nov ', 'dec ']):
            # Skip non-numeric columns
            if consolidated_df[col].dtype in ['float64', 'int64'] or pd.api.types.is_numeric_dtype(consolidated_df[col]):
                amount_columns.append(col)
    
    # Track FX rates used
    fx_rates_used = {}
    
    # Convert amount columns
    for col in amount_columns:
        # Create new columns for native and converted amounts
        native_col = f"{col}_Native_{clean_base_currency}"
        converted_col = f"{col}_ZAR"
        fx_rate_col = f"{col}_FX_Rate"
        
        # Store original amounts in native column
        consolidated_df[native_col] = consolidated_df[col]
        
        # Convert amounts
        converted_amounts = []
        fx_rates = []
        
        for amount in consolidated_df[col]:
            converted_amount, fx_rate = convert_amount_to_zar(amount, clean_base_currency, fx_reader, entity_name)
            converted_amounts.append(converted_amount)
            fx_rates.append(fx_rate)
        
        # Store converted amounts and FX rates
        consolidated_df[converted_col] = converted_amounts
        consolidated_df[fx_rate_col] = fx_rates
        
        # Track the FX rate used (use the first non-1.0 rate found)
        unique_rates = [r for r in fx_rates if r != 1.0]
        if unique_rates:
            fx_rates_used[f"{clean_base_currency}_ZAR"] = unique_rates[0]
        elif clean_base_currency != 'ZAR':
            fx_rates_used[f"{clean_base_currency}_ZAR"] = 1.0
        
        # Remove original column or rename it
        consolidated_df = consolidated_df.drop(columns=[col])
    
    return consolidated_df, fx_rates_used

def create_fx_rates_sheet(fx_rates_used, target_date):
    """Create a DataFrame with FX rates used for consolidation"""
    
    if not fx_rates_used:
        return pd.DataFrame()
    
    fx_data = []
    for currency_pair, rate in fx_rates_used.items():
        from_currency, to_currency = currency_pair.split('_')
        fx_data.append({
            'From_Currency': from_currency,
            'To_Currency': to_currency,
            'Exchange_Rate': rate,
            'Date_Applied': target_date.strftime('%Y-%m-%d'),
            'Source': 'fx/FX.xlsx'
        })
    
    return pd.DataFrame(fx_data)

def main():
    """Main function with multi-company support and consolidation options"""
    
    # Parse command line arguments
    if len(sys.argv) < 2:
        print("Usage: python get_xero_reports.py <date> [consolidation_mode]")
        print("Examples:")
        print("  python get_xero_reports.py \"June 2025\"            # All companies, native currencies (default)")
        print("  python get_xero_reports.py \"June 2025\" native     # All companies, native currencies")
        print("  python get_xero_reports.py \"June 2025\" consolidated # Single consolidated report (all entities in ZAR)")
        print("  python get_xero_reports.py \"June 2025\" both       # Both native + consolidated reports")
        print("")
        print("Consolidation modes:")
        print("  - native: Separate reports per entity in native currencies (default)")
        print("  - consolidated: Single consolidated report combining all entities, converted to ZAR")
        print("  - both: Generate both native reports AND consolidated reports")
        return
    
    # Parse parameters (simplified - no company filtering)
    date_input = sys.argv[1]
    consolidation_mode = sys.argv[2].lower() if len(sys.argv) > 2 else 'native'
    target_company = None  # Always process all companies
    
    target_date = parse_date_input(date_input)
    if not target_date:
        return
    
    # Validate consolidation mode
    if consolidation_mode not in ['native', 'consolidated', 'both']:
        print(f"❌ Invalid consolidation mode: {consolidation_mode}")
        print("Valid modes: native, consolidated, both")
        return
    
    # Initialize FX reader if consolidation is needed
    fx_reader = None
    if consolidation_mode in ['consolidated', 'both']:
        print("💱 Initializing FX rate reader...")
        fx_reader = FXRateReader()
        if not fx_reader.load_fx_data():
            print("❌ Failed to load FX data. Consolidation features will be disabled.")
            if consolidation_mode == 'consolidated':
                print("❌ Cannot proceed with consolidated mode without FX data.")
                return
            consolidation_mode = 'native'  # Fallback to native mode
        else:
            available_currencies = fx_reader.get_available_currencies()
            print(f"✅ FX data loaded. Available currencies: {', '.join(available_currencies)}")
    
    # Create timestamp for this run
    run_timestamp = datetime.now()
    
    print(f"📊 Consolidation mode: {consolidation_mode.upper()}")
    if fx_reader:
        fx_summary = fx_reader.get_fx_rates_summary()
        print(f"💱 FX rates loaded: {len(fx_summary)} currency pairs")
    
    # Handle different consolidation modes
    if consolidation_mode == 'native':
        # Generate native reports for all companies
        print(f"🌍 Generating NATIVE reports for ALL companies")
        print("=" * 50)
        generate_reports_all_companies_native(target_date, run_timestamp)
        
    elif consolidation_mode == 'consolidated':
        # Generate consolidated reports only
        print(f"🏦 Generating CONSOLIDATED reports (all entities in ZAR)")
        print("=" * 50)
        generate_consolidated_reports(target_date, run_timestamp, fx_reader)
        
    elif consolidation_mode == 'both':
        # Generate both native and consolidated
        print(f"📊 Generating BOTH native AND consolidated reports")
        print("=" * 50)
        
        # First generate native reports
        print(f"\n🌍 Step 1: Generating NATIVE reports for ALL companies")
        print("-" * 50)
        generate_reports_all_companies_native(target_date, run_timestamp)
        
        # Then generate consolidated reports
        print(f"\n🏦 Step 2: Generating CONSOLIDATED reports (all entities in ZAR)")
        print("-" * 50)
        generate_consolidated_reports(target_date, run_timestamp, fx_reader)

def generate_consolidated_reports(target_date, run_timestamp, fx_reader):
    """Generate consolidated reports combining all entities in ZAR"""
    
    if not fx_reader:
        print("❌ FX reader is required for consolidated reports")
        return
    
    print("🏦 Collecting data from all entities for consolidation...")
    
    # Get all available companies
    try:
        from xero_python.identity import IdentityApi
        api_client = setup_xero_client()
        identity_api = IdentityApi(api_client)
        connections = identity_api.get_connections()
        
        if not connections:
            print("❌ No Xero connections found")
            return
        
        # Process all companies
        companies_to_process = []
        for conn in connections:
            code = get_company_code(conn.tenant_name)
            if code not in [c['code'] for c in companies_to_process]:
                companies_to_process.append({
                    'code': code, 
                    'name': conn.tenant_name,
                    'tenant_id': conn.tenant_id
                })
        
        if not companies_to_process:
            print(f"❌ No companies found for consolidation")
            return
            
        print(f"📊 Consolidating data from {len(companies_to_process)} companies: {', '.join([c['code'] for c in companies_to_process])}")
        
        # Archive existing consolidated files
        import os
        import glob
        os.makedirs('xero_data', exist_ok=True)
        os.makedirs('xero_data/archive', exist_ok=True)
        
        # Archive only consolidated files (those with CONS in the name)
        existing_consolidated_files = glob.glob('xero_data/*CONS*.xlsx')
        for filepath in existing_consolidated_files:
            filename = os.path.basename(filepath)
            try:
                timestamp_str = run_timestamp.strftime("%H%M%S")
                archive_name = f"archive_{timestamp_str}_{filename}"
                archive_path = f'xero_data/archive/{archive_name}'
                os.rename(filepath, archive_path)
                print(f"Archived consolidated file: {filename} -> {archive_name}")
            except Exception as e:
                print(f"Could not archive {filename}: {e}")
        
        # Collect data from all companies
        consolidated_data = {
            'balance_sheets': [],
            'pnl_data': [],
            'trial_balances': [],
            'chart_of_accounts': [],
            'invoices': []
        }
        
        all_fx_rates_used = {}
        
        # Process each company
        for company in companies_to_process:
            print(f"\n📊 Collecting data from {company['code']} - {company['name']}")
            
            # Get company data (this would need to be extracted from the existing functions)
            company_data = collect_company_data_for_consolidation(
                company['tenant_id'], company['code'], company['name'], 
                target_date, fx_reader
            )
            
            if company_data:
                # Add to consolidated data
                for report_type, data in company_data.items():
                    if report_type == 'fx_rates_used':
                        all_fx_rates_used.update(data)
                    else:
                        consolidated_data[report_type].extend(data)
        
        # Generate consolidated reports
        generate_consolidated_excel_files(consolidated_data, all_fx_rates_used, target_date, run_timestamp)
        
        print("\n✅ Consolidated reports generation completed")
        
    except Exception as e:
        print(f"❌ Error generating consolidated reports: {e}")
        return

def generate_reports_all_companies_native(target_date, run_timestamp):
    """Generate native reports for all companies (original behavior)"""
    
    # Get all available companies
    try:
        from xero_python.identity import IdentityApi
        api_client = setup_xero_client()
        identity_api = IdentityApi(api_client)
        connections = identity_api.get_connections()
        
        if not connections:
            print("❌ No Xero connections found")
            return
        
        # Get unique company codes
        companies = []
        for conn in connections:
            code = get_company_code(conn.tenant_name)
            if code not in [c['code'] for c in companies]:
                companies.append({'code': code, 'name': conn.tenant_name})
        
        print(f"📊 Found {len(companies)} companies: {', '.join([c['code'] for c in companies])}")
        
        # Archive existing files ONCE at the beginning
        print(f"\n🗂️ ARCHIVING EXISTING FILES (ONCE FOR ALL COMPANIES)")
        print("-" * 60)
        
        import os
        import glob
        os.makedirs('xero_data', exist_ok=True)
        os.makedirs('xero_data/archive', exist_ok=True)
        
        existing_files = glob.glob('xero_data/*.xlsx') + glob.glob('xero_data/*.csv')
        archived_old_files = 0
        
        for filepath in existing_files:
            filename = os.path.basename(filepath)
            if filename.lower().startswith('readme'):
                continue
                
            try:
                timestamp_str = run_timestamp.strftime("%H%M%S")
                archive_name = f"archive_{timestamp_str}_{filename}"
                archive_path = f'xero_data/archive/{archive_name}'
                os.rename(filepath, archive_path)
                print(f"Archived: {filename} -> {archive_name}")
                archived_old_files += 1
            except Exception as e:
                print(f"Could not archive {filename}: {e}")
            
            if archived_old_files == 0:
                print("No existing files found to archive")
            else:
                print(f"✅ Archived {archived_old_files} existing files")
            
            print()
            
            # Generate reports for each company (skip archiving for each)
            for i, company in enumerate(companies, 1):
                print(f"\n{'='*60}")
                print(f"🏢 COMPANY {i}/{len(companies)}: {company['code']} - {company['name']}")
                print(f"{'='*60}")
                
                # Set target company in sys.argv for the single company function
                original_argv = sys.argv.copy()
                sys.argv = [sys.argv[0], date_input, company['code']]
                
                try:
                    # Skip archiving since we did it once at the beginning
                    generate_reports_single_company(target_date, run_timestamp, skip_archiving=True)
                    print(f"✅ {company['code']} reports completed")
                except Exception as e:
                    print(f"❌ Error generating {company['code']} reports: {e}")
                finally:
                    # Restore original argv
                    sys.argv = original_argv
            
            print(f"\n🎉 ALL COMPANY REPORTS COMPLETED")
            print(f"Generated reports for: {', '.join([c['code'] for c in companies])}")
            
            # Show final file list
            final_files = glob.glob('xero_data/*.xlsx')
            if final_files:
                print(f"\n📁 Final reports in xero_data/:")
                for filepath in sorted(final_files):
                    filename = os.path.basename(filepath)
                    if not filename.lower().startswith('readme'):
                        print(f"   📄 {filename}")
            
        except Exception as e:
            print(f"❌ Error in multi-company generation: {e}")

if __name__ == "__main__":
    main()
