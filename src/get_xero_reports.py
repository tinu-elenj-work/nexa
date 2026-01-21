"""
Nexa - Xero Financial Reports (with Automatic Archiving)

This is the main script for retrieving Xero financial reports. All reports are automatically 
archived with timestamps and financial year end codes.

Filename format: {report_type}_{YYYY_MM_DD}_{HHMMSS}_{FYEND}.csv
Example: balance_sheet_2025_06_30_143022_FEB26.csv

Usage:
    python get_xero_reports.py "June 2025"        # End of June 2025
    python get_xero_reports.py "2025-06-30"       # Specific date
    python get_xero_reports.py "31 December 2024" # Written format
    python get_xero_reports.py                    # Today's date
"""

import os
import sys
import pandas as pd
import re
import json
from datetime import datetime, date
from dateutil.parser import parse
from dateutil.relativedelta import relativedelta
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from config import XERO_CONFIG

# Company code mapping - add new companies as needed
COMPANY_CODES = {
    'Elenjical Solutions (Pty) Ltd': 'SA',
    'Elenjical Solutions': 'SA',
    'Elenjical Solutions MA (USD)': 'MA',
    'Elenjical Solutions MA': 'MA',
    'Elenjical Solutions Private Limited': 'IND',
    'Elenjical Solutions India': 'IND',
    'Elenjical Solutions IND': 'IND',
    # UK company - exact Xero organization name
    'Elenjical Solutions International Limited': 'UK',
    'Elenjical Solutions International': 'UK',
    'Elenjical Solutions UK': 'UK',
    # Add more companies here as you consolidate:
}

# Import Xero SDK
from xero_python.api_client import ApiClient, Configuration
from xero_python.api_client.oauth2 import OAuth2Token
from xero_python.accounting import AccountingApi
from xero_python.identity import IdentityApi

# Import FX rate reader for consolidation
from fx_reader import FXRateReader

# FX disaggregation removed - preserving ALL native Xero API data

def parse_date_input(date_input=None):
    """Parse various date input formats"""
    
    if not date_input:
        return date.today()
    
    try:
        # Handle month/year format like "June 2025"
        if len(date_input.split()) == 2 and date_input.split()[1].isdigit():
            month_name, year = date_input.split()
            # Parse to get first day of month, then get last day
            first_day = parse(f"1 {month_name} {year}").date()
            # Get last day of the month
            next_month = first_day + relativedelta(months=1)
            last_day = next_month - relativedelta(days=1)
            return last_day
        
        # Handle other formats
        parsed_date = parse(date_input).date()
        return parsed_date
        
    except Exception as e:
        print(f" Could not parse date '{date_input}': {e}")
        print(" Try formats like: 'June 2025', '2025-06-30', '30 June 2025'")
        return None

def get_financial_year_dates(target_date):
    """Get financial year start and end dates for a given date (South African FY: March to February)"""
    
    if target_date.month >= 3:  # March onwards = current financial year
        fy_start = date(target_date.year, 3, 1)
        fy_end = date(target_date.year + 1, 2, 28)  # Or 29 for leap year
        if target_date.year % 4 == 0 and (target_date.year % 100 != 0 or target_date.year % 400 == 0):
            fy_end = date(target_date.year + 1, 2, 29)
    else:  # January-February = previous financial year
        fy_start = date(target_date.year - 1, 3, 1)
        fy_end = date(target_date.year, 2, 28)
        if target_date.year % 4 == 0 and (target_date.year % 100 != 0 or target_date.year % 400 == 0):
            fy_end = date(target_date.year, 2, 29)
    
    return fy_start, fy_end

def get_fy_end_code(target_date):
    """Get financial year end code (e.g., FEB26 for FY ending Feb 2026)"""
    
    fy_start, fy_end = get_financial_year_dates(target_date)
    fy_end_code = fy_end.strftime("%b%y").upper()
    return fy_end_code

def load_pl_account_order():
    """Load P&L account type ordering configuration from JSON file"""
    try:
        config_file = 'pl_account_order.json'
        if os.path.exists(config_file):
            with open(config_file, 'r') as f:
                config = json.load(f)
            
            # Create account type order mapping
            type_order = {}
            for item in config.get('account_type_order', []):
                type_order[item['type']] = item['order']
            
            return type_order
        else:
            print(f"⚠️ P&L ordering config file not found: {config_file}")
            # Default ordering if file doesn't exist
            return {
                'REVENUE': 1,
                'DIRECTCOSTS': 2, 
                'OTHERINCOME': 3,
                'EXPENSE': 4
            }
    except Exception as e:
        print(f"⚠️ Error loading P&L ordering config: {e}")
        # Default ordering on error
        return {
            'REVENUE': 1,
            'DIRECTCOSTS': 2,
            'OTHERINCOME': 3, 
            'EXPENSE': 4
        }

def sort_pl_accounts(accounts_data, type_order_mapping):
    """Sort P&L accounts according to type order and then alphabetically by name"""
    
    def get_sort_key(account):
        # Get account type (convert to uppercase for matching)
        account_type_raw = str(account.get('type', '')).upper()
        
        # Remove ACCOUNTTYPE. prefix if present (Xero SDK format)
        if account_type_raw.startswith('ACCOUNTTYPE.'):
            account_type = account_type_raw.replace('ACCOUNTTYPE.', '')
        else:
            account_type = account_type_raw
        
        # Get order from mapping, default to 999 for unknown types
        type_order = type_order_mapping.get(account_type, 999)
        
        # Get account name for secondary sort
        account_name = account.get('account_name', '')
        
        return (type_order, account_name.upper())
    
    # Debug: Show account types before sorting
    account_types_found = {}
    for account in accounts_data:
        account_type_raw = str(account.get('type', '')).upper()
        # Remove ACCOUNTTYPE. prefix for display
        if account_type_raw.startswith('ACCOUNTTYPE.'):
            account_type = account_type_raw.replace('ACCOUNTTYPE.', '')
        else:
            account_type = account_type_raw
            
        if account_type not in account_types_found:
            account_types_found[account_type] = []
        account_types_found[account_type].append(account.get('account_name', ''))
    
    print(f"   📋 Account types found:")
    for acc_type, names in account_types_found.items():
        order = type_order_mapping.get(acc_type, 999)
        print(f"      {acc_type} (order {order}): {len(names)} accounts")
    
    # Sort the accounts
    sorted_accounts = sorted(accounts_data, key=get_sort_key)
    
    # Debug: Show first few accounts after sorting
    print(f"   📋 First 5 accounts after sorting:")
    for i, account in enumerate(sorted_accounts[:5]):
        acc_type_raw = str(account.get('type', '')).upper()
        # Remove ACCOUNTTYPE. prefix for display
        if acc_type_raw.startswith('ACCOUNTTYPE.'):
            acc_type = acc_type_raw.replace('ACCOUNTTYPE.', '')
        else:
            acc_type = acc_type_raw
        acc_name = account.get('account_name', '')
        order = type_order_mapping.get(acc_type, 999)
        print(f"      {i+1}. {acc_name} ({acc_type}, order {order})")
    
    return sorted_accounts

def get_company_code(org_name):
    """Get company short code from organization name"""
    
    # Direct lookup first
    if org_name in COMPANY_CODES:
        return COMPANY_CODES[org_name]
    
    # Fuzzy matching for similar names
    for company_name, code in COMPANY_CODES.items():
        if company_name.lower() in org_name.lower() or org_name.lower() in company_name.lower():
            return code
    
    # Default fallback - use first 4 chars of company name
    clean_name = ''.join(c.upper() for c in org_name if c.isalnum())
    return clean_name[:4] if clean_name else 'UNKN'

def create_archive_filename(report_type, target_date, timestamp=None, company_code=None):
    """Create archive-ready filename without company code for multi-sheet files"""
    
    if timestamp is None:
        timestamp = datetime.now()
    
    date_str = target_date.strftime("%Y_%m_%d")
    time_str = timestamp.strftime("%H%M%S")
    fy_code = get_fy_end_code(target_date)
    
    # No longer include company code in filename - using multi-sheet approach
    filename = f"{report_type}_{date_str}_{time_str}_{fy_code}.xlsx"
    
    return filename

def move_to_archive(filename):
    """Move file to archive directory (automatic archiving)"""
    
    source_path = f'../../output/xero_data/{filename}'
    archive_path = f'../../output/xero_data/archive/{filename}'
    
    try:
        # Create archive directory if it doesn't exist
        os.makedirs('../output/xero_data/archive', exist_ok=True)
        
        if os.path.exists(source_path):
            os.rename(source_path, archive_path)
            print(f" Auto-archived: {filename}")
            return True
        else:
            print(f" File not found for archiving: {filename}")
            return False
    except Exception as e:
        print(f" Error archiving: {e}")
        return False

def setup_xero_client():
    """Set up Xero API client with proper token management and proactive refresh"""
    
    # Try to refresh token proactively to avoid expiry during execution
    try:
        import requests
        import base64
        
        print("Proactively refreshing token to ensure validity...")
        
        # Refresh the token proactively
        refresh_url = "https://identity.xero.com/connect/token"
        refresh_data = {
            'grant_type': 'refresh_token',
            'refresh_token': XERO_CONFIG['refresh_token']
        }
        
        credentials = base64.b64encode(f"{XERO_CONFIG['client_id']}:{XERO_CONFIG['client_secret']}".encode()).decode()
        headers = {
            'Authorization': f'Basic {credentials}',
            'Content-Type': 'application/x-www-form-urlencoded'
        }
        
        response = requests.post(refresh_url, data=refresh_data, headers=headers)
        
        if response.status_code == 200:
            token_data = response.json()
            
            # Update config.py with new tokens
            try:
                with open('../config/config.py', 'r', encoding='utf-8') as f:
                    config_content = f.read()
                
                # Update access_token and refresh_token in the file
                config_content = re.sub(
                    r"'access_token': '[^']*'",
                    f"'access_token': '{token_data['access_token']}'",
                    config_content
                )
                config_content = re.sub(
                    r"'refresh_token': '[^']*'",
                    f"'refresh_token': '{token_data['refresh_token']}'",
                    config_content
                )
                
                # Write updated config back to file
                with open('../config/config.py', 'w', encoding='utf-8') as f:
                    f.write(config_content)
                
                # Update the XERO_CONFIG dict in memory
                XERO_CONFIG['access_token'] = token_data['access_token']
                XERO_CONFIG['refresh_token'] = token_data['refresh_token']
                
                print("Token proactively refreshed and saved - expires in 1800 seconds")
                
            except Exception as e:
                print(f"Token refreshed but failed to save to ../config/config.py: {e}")
        else:
            print("Using existing token...")
            
    except Exception as e:
        print(f"Proactive token refresh failed, will rely on automatic refresh: {e}")
    
    # Create API client
    api_client = ApiClient(
        Configuration(
            debug=False,
            oauth2_token=OAuth2Token(
                client_id=XERO_CONFIG['client_id'],
                client_secret=XERO_CONFIG['client_secret']
            ),
        ),
        pool_threads=1,
    )
    
    # Set up token
    token_set = {
        "access_token": XERO_CONFIG['access_token'],
        "token_type": "Bearer",
        "refresh_token": XERO_CONFIG['refresh_token'],
        "expires_in": 1800,  # 30 minutes
        "scope": XERO_CONFIG['scopes']
    }
    
    # Set up token getter/setter functions
    @api_client.oauth2_token_getter
    def obtain_xero_oauth2_token():
        return token_set
    
    @api_client.oauth2_token_saver
    def store_xero_oauth2_token(token):
        token_set.update(token)
        
        # Automatically update config.py with new tokens
        try:
            # Read current config file
            with open('../config/config.py', 'r', encoding='utf-8') as f:
                config_content = f.read()
            
            # Update access_token and refresh_token in the file
            if 'access_token' in token:
                config_content = re.sub(
                    r"'access_token': '[^']*'",
                    f"'access_token': '{token['access_token']}'",
                    config_content
                )
            
            if 'refresh_token' in token:
                config_content = re.sub(
                    r"'refresh_token': '[^']*'",
                    f"'refresh_token': '{token['refresh_token']}'",
                    config_content
                )
            
            # Write updated config back
            with open('../config/config.py', 'w', encoding='utf-8') as f:
                f.write(config_content)
                
            print(f"🔄 Token automatically refreshed and saved - expires in {token.get('expires_in', 'unknown')} seconds")
            
        except Exception as e:
            print(f"⚠️ Token refreshed but failed to save to ../config/config.py: {e}")
    
    # Set the token
    api_client.set_oauth2_token(token_set)
    
    return api_client

def get_organization_info(api_client, tenant_id):
    """Get organization information including base currency"""
    
    accounting_api = AccountingApi(api_client)
    
    try:
        orgs = accounting_api.get_organisations(xero_tenant_id=tenant_id)
        if orgs and orgs.organisations:
            org = orgs.organisations[0]
            return {
                'name': org.name,
                'base_currency': org.base_currency,
                'country_code': org.country_code,
                'tax_number': org.tax_number
            }
    except Exception as e:
        print(f" Could not get organization info: {e}")
        return {'base_currency': 'ZAR'}  # Default to ZAR for South Africa

def format_accounting(amount):
    """Format amount in accounting format without currency symbol"""
    
    # Format with thousand separators and 2 decimal places
    if amount == 0:
        return "0.00"
    elif amount > 0:
        return f"{amount:,.2f}"
    else:
        return f"({abs(amount):,.2f})"  # Negative amounts in parentheses (accounting style)

def format_currency_display(amount, currency_code='ZAR'):
    """Format amount with currency symbol for display purposes only"""
    
    # Currency symbols mapping
    currency_symbols = {
        'ZAR': 'R',
        'USD': '$',
        'GBP': '',
        'EUR': '',
        'CurrencyCode.ZAR': 'R',
        'CurrencyCode.USD': '$',
        'CurrencyCode.GBP': '',
        'CurrencyCode.EUR': '',
    }
    
    # Clean currency code
    clean_code = currency_code.replace('CurrencyCode.', '') if currency_code else 'ZAR'
    symbol = currency_symbols.get(clean_code, f'{clean_code} ')
    
    # Format the amount with currency for display
    if amount == 0:
        return f"{symbol}0.00"
    elif amount > 0:
        return f"{symbol}{amount:,.2f}"
    else:
        return f"-{symbol}{abs(amount):,.2f}"

def format_data_for_export(data, data_type='general'):
    """Centrally format data with proper date formats (dd/mm/yyyy) and keep numbers as raw values for Excel"""
    
    if isinstance(data, list):
        # Handle list of dictionaries (like invoices)
        formatted_data = []
        
        for item in data:
            formatted_item = item.copy()
            
            # Format all date fields to dd/mm/yyyy
            date_fields = ['date', 'due_date', 'fully_paid_on_date', 'expected_payment_date', 'planned_payment_date', 'updated_date_utc']
            
            for field in date_fields:
                if field in formatted_item and formatted_item[field]:
                    try:
                        # Parse the date string and reformat to dd/mm/yyyy
                        if formatted_item[field] != '':
                            date_obj = pd.to_datetime(formatted_item[field])
                            # Special handling for datetime fields (include time)
                            if 'utc' in field.lower() or 'updated' in field.lower():
                                formatted_item[field] = date_obj.strftime('%d/%m/%Y %H:%M')
                            else:
                                formatted_item[field] = date_obj.strftime('%d/%m/%Y')
                    except:
                        # Keep original value if parsing fails
                        pass
            
            # Keep numeric fields as raw numbers - Excel formatting will handle the display
            formatted_data.append(formatted_item)
        
        return formatted_data
    
    else:
        # Handle DataFrame (like financial reports)
        return data

def export_to_excel_with_formatting(df, filename, sheet_name, data_type='general'):
    """Central Excel export function with consistent formatting for all data types"""
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Define formatting standards
        accounting_format = '#,##0.00_);(#,##0.00)'  # No currency, negative in parentheses
        
        # Define column types for different data formats
        financial_columns = {
            'balance', 'sub_total', 'total_tax', 'total', 'amount_due', 'amount_paid', 
            'amount_credited', 'currency_rate', 'first_line_quantity', 'first_line_unit_amount', 
            'first_line_line_amount'
        }
        
        # Apply formatting based on column names and data types
        from openpyxl.utils import get_column_letter
        
        for idx, col in enumerate(df.columns, 1):
            col_letter = get_column_letter(idx)
            
            # Handle YTD Total column specially for formulas
            if col == 'ytd_total' and data_type == 'financial':
                # Find ALL month columns (both YYYY-MM format and any other date columns)
                month_cols = []
                for c in df.columns:
                    # Include date-like columns: YYYY-MM format or month names
                    if (c.startswith('2025-') or c.startswith('2024-') or c.startswith('2026-') or
                        c in ['Mar 2025', 'Apr 2025', 'May 2025', 'Jun 2025', 'Jul 2025', 'Aug 2025',
                              'Sep 2025', 'Oct 2025', 'Nov 2025', 'Dec 2025', 'Jan 2025', 'Feb 2025']):
                        month_cols.append(c)
                
                for row in range(2, len(df) + 2):  # Start from row 2 (after header)
                    cell = worksheet[f'{col_letter}{row}']
                    
                    # Create Excel formula with proper range notation
                    if month_cols:
                        # Find first and last month column letters for range
                        first_month_idx = df.columns.get_loc(month_cols[0]) + 1
                        last_month_idx = df.columns.get_loc(month_cols[-1]) + 1
                        
                        first_col_letter = get_column_letter(first_month_idx)
                        last_col_letter = get_column_letter(last_month_idx)
                        
                        # Set Excel formula using range notation (e.g., =SUM(D5:G5))
                        formula = f"=SUM({first_col_letter}{row}:{last_col_letter}{row})"
                        cell.value = formula
                    
                    # Apply accounting format
                    cell.number_format = accounting_format
            
            # Handle monthly columns specially for totals row formulas
            elif col.startswith('2025-') and data_type == 'financial':
                # Check if this is the totals row (last row)
                last_row = len(df) + 1  # +1 because Excel is 1-indexed
                
                for row in range(2, len(df) + 2):  # Start from row 2 (after header)
                    cell = worksheet[f'{col_letter}{row}']
                    
                    # If this is the totals row, add column sum formula
                    if row == last_row:
                        # Create column sum formula (exclude totals row itself)
                        formula = f"=SUM({col_letter}2:{col_letter}{last_row-1})"
                        cell.value = formula
                    
                    # Apply accounting format
                    cell.number_format = accounting_format
            
            # Apply accounting format to other financial columns
            elif col.lower() in financial_columns or df[col].dtype in ['int64', 'float64']:
                for row in range(2, len(df) + 2):  # Start from row 2 (after header)
                    cell = worksheet[f'{col_letter}{row}']
                    cell.number_format = accounting_format
        
        # COSMETIC ENHANCEMENTS
        # 1. Auto-expand all columns to fit content
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set column width with some padding
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 chars
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 2. Freeze the first row (headers)
        worksheet.freeze_panes = 'A2'  # Freeze everything above row 2

def export_multi_company_to_excel(company_data_dict, filename, data_type='general'):
    """Export multiple company data to a single Excel file with separate sheets per company"""
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        
        # Define formatting standards
        accounting_format = '#,##0.00_);(#,##0.00)'  # No currency, negative in parentheses
        financial_columns = {
            'balance', 'sub_total', 'total_tax', 'total', 'amount_due', 'amount_paid', 
            'amount_credited', 'currency_rate', 'first_line_quantity', 'first_line_unit_amount', 
            'first_line_line_amount'
        }
        
        from openpyxl.utils import get_column_letter
        
        # Process each company's data
        for company_code, df in company_data_dict.items():
            if df is not None and not df.empty:
                # Use company code as sheet name
                sheet_name = company_code
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Get the worksheet for formatting
                worksheet = writer.sheets[sheet_name]
                
                # Apply the same formatting as single sheet export
                for idx, col in enumerate(df.columns, 1):
                    col_letter = get_column_letter(idx)
                    
                    # Handle YTD Total column specially for formulas
                    if col == 'ytd_total' and data_type == 'financial':
                        # Find ALL month columns (both YYYY-MM format and any other date columns)
                        month_cols = []
                        for c in df.columns:
                            # Include date-like columns: YYYY-MM format or month names
                            if (c.startswith('2025-') or c.startswith('2024-') or c.startswith('2026-') or
                                c in ['Mar 2025', 'Apr 2025', 'May 2025', 'Jun 2025', 'Jul 2025', 'Aug 2025',
                                      'Sep 2025', 'Oct 2025', 'Nov 2025', 'Dec 2025', 'Jan 2025', 'Feb 2025']):
                                month_cols.append(c)
                        
                        for row in range(2, len(df) + 2):  # Start from row 2 (after header)
                            cell = worksheet[f'{col_letter}{row}']
                            
                            # Create Excel formula with proper range notation
                            if month_cols:
                                # Find first and last month column letters for range
                                first_month_idx = df.columns.get_loc(month_cols[0]) + 1
                                last_month_idx = df.columns.get_loc(month_cols[-1]) + 1
                                
                                first_col_letter = get_column_letter(first_month_idx)
                                last_col_letter = get_column_letter(last_month_idx)
                                
                                # Set Excel formula using range notation (e.g., =SUM(D5:G5))
                                formula = f"=SUM({first_col_letter}{row}:{last_col_letter}{row})"
                                cell.value = formula
                            
                            # Apply accounting format
                            cell.number_format = accounting_format
                    
                    # Handle monthly columns specially for totals row formulas
                    elif col.startswith('2025-') and data_type == 'financial':
                        for row in range(2, len(df) + 2):  # Start from row 2 (after header)
                            cell = worksheet[f'{col_letter}{row}']
                            cell.number_format = accounting_format
                    
                    # Apply accounting format to other financial columns
                    elif col.lower() in financial_columns or df[col].dtype in ['int64', 'float64']:
                        for row in range(2, len(df) + 2):  # Start from row 2 (after header)
                            cell = worksheet[f'{col_letter}{row}']
                            cell.number_format = accounting_format
                
                # COSMETIC ENHANCEMENTS
                # 1. Auto-expand all columns to fit content
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    # Set column width with some padding
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 chars
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                        # 2. Freeze the first row (headers)
        worksheet.freeze_panes = 'A2'  # Freeze everything above row 2

def add_consolidated_sheet_to_file(filepath, company_data_dict, fx_reader, target_date, report_type, data_format):
    """Add a consolidated sheet to an existing multi-sheet Excel file"""
    
    # Consolidate all company data into ZAR
    consolidated_data = []
    fx_rates_used = {}
    
    for company_code, df in company_data_dict.items():
        if df is not None and not df.empty:
            # Get the base currency for this company (from the entity column if available)
            base_currency = 'ZAR'  # Default
            if 'entity' in df.columns and not df.empty:
                # Try to infer currency from company code
                currency_mapping = {'SA': 'ZAR', 'MA': 'USD', 'IND': 'INR', 'UK': 'GBP'}
                base_currency = currency_mapping.get(company_code, 'ZAR')
            
            # Convert to ZAR if needed
            if base_currency != 'ZAR' and fx_reader:
                fx_rate = fx_reader.get_fx_rate(base_currency, 'ZAR')
                if fx_rate:
                    fx_rates_used[f"{base_currency}_ZAR"] = fx_rate
                    # Convert monetary columns
                    df_converted = df.copy()
                    for col in df.columns:
                        if col.startswith('2025-') or col == 'ytd_total':
                            if col in df_converted.columns:
                                df_converted[col] = df_converted[col] * fx_rate
                    consolidated_data.append(df_converted)
                else:
                    # Use original data if no FX rate available
                    consolidated_data.append(df)
            else:
                # ZAR data or no FX reader - use as is
                consolidated_data.append(df)
    
    # Combine all data
    if consolidated_data:
        consolidated_df = pd.concat(consolidated_data, ignore_index=True)
        
        # Add the consolidated sheet to the existing file
        with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
            consolidated_df.to_excel(writer, sheet_name='CONSOLIDATED', index=False)
            
            # Apply formatting to the consolidated sheet
            worksheet = writer.sheets['CONSOLIDATED']
            
            # Apply the same formatting as other sheets
            from openpyxl.utils import get_column_letter
            accounting_format = '#,##0.00_);(#,##0.00)'
            financial_columns = {
                'balance', 'sub_total', 'total_tax', 'total', 'amount_due', 'amount_paid', 
                'amount_credited', 'currency_rate'
            }
            
            for idx, col in enumerate(consolidated_df.columns, 1):
                col_letter = get_column_letter(idx)
                
                # Handle YTD Total column specially for formulas
                if col == 'ytd_total' and data_format == 'financial':
                    # Find ALL month columns
                    month_cols = []
                    for c in consolidated_df.columns:
                        if (c.startswith('2025-') or c.startswith('2024-') or c.startswith('2026-')):
                            month_cols.append(c)
                    
                    if month_cols:
                        first_month_idx = consolidated_df.columns.get_loc(month_cols[0]) + 1
                        last_month_idx = consolidated_df.columns.get_loc(month_cols[-1]) + 1
                        
                        first_col_letter = get_column_letter(first_month_idx)
                        last_col_letter = get_column_letter(last_month_idx)
                        
                        # Add SUM formula for each row
                        for row in range(2, len(consolidated_df) + 2):
                            cell = worksheet[f'{col_letter}{row}']
                            formula = f"=SUM({first_col_letter}{row}:{last_col_letter}{row})"
                            cell.value = formula
                            cell.number_format = accounting_format
                
                # Handle monthly columns
                elif col.startswith('2025-') and data_format == 'financial':
                    for row in range(2, len(consolidated_df) + 2):
                        cell = worksheet[f'{col_letter}{row}']
                        cell.number_format = accounting_format
                
                # Apply accounting format to other financial columns
                elif col.lower() in financial_columns or consolidated_df[col].dtype in ['int64', 'float64']:
                    for row in range(2, len(consolidated_df) + 2):
                        cell = worksheet[f'{col_letter}{row}']
                        cell.number_format = accounting_format
            
            # Auto-expand columns and freeze headers
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            worksheet.freeze_panes = 'A2'

def create_sheet_name(report_type, target_date, company_code=None, is_consolidated=False):
    """Create sheet name with company code, report type, month/year, and FYE"""
    
    # Format the month and year (shorter format for Excel sheet name limits)
    month_year = target_date.strftime("%b%y")  # e.g., "Jun25"
    
    # Get FYE code
    fy_code = get_fy_end_code(target_date)
    
    # Map report types to sheet prefixes (keeping under 31 char Excel limit)
    sheet_prefixes = {
        'balance_sheet': 'BS',
        'profit_and_loss': 'PL', 
        'monthly_pnl_ytd': 'MonthlyPL',
        'trial_balance': 'TB',
        'chart_of_accounts': 'COA',
        'invoices': 'INV',
        'fx_rates': 'FX'
    }
    
    prefix = sheet_prefixes.get(report_type, report_type[:3].upper())
    
    # Add consolidation suffix
    if is_consolidated:
        prefix += '_CONS'
    
    # Create sheet name (Excel limit is 31 characters)
    if company_code:
        sheet_name = f"{company_code}_{prefix}_{month_year}_{fy_code}"
    else:
        sheet_name = f"{prefix}_{month_year}_{fy_code}"
    
    # Ensure it's under 31 characters
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]
    
    return sheet_name

def get_chart_of_accounts_lookup(accounting_api, tenant_id):
    """Fetch chart of accounts and create lookup dictionary by account name"""
    try:
        accounts_response = accounting_api.get_accounts(xero_tenant_id=tenant_id)
        
        if accounts_response.accounts:
            # Create lookup dictionary by account name
            account_lookup = {}
            for account in accounts_response.accounts:
                account_name = getattr(account, 'name', '')
                account_lookup[account_name] = {
                    'id': getattr(account, 'account_id', ''),
                    'name': account_name,
                    'code': getattr(account, 'code', ''),
                    'type': str(getattr(account, 'type', '')) if getattr(account, 'type', None) else '',
                    'tax_type': str(getattr(account, 'tax_type', '')) if getattr(account, 'tax_type', None) else '',
                    'status': str(getattr(account, 'status', '')) if getattr(account, 'status', None) else ''
                }
            
            print(f"   📋 Chart of Accounts lookup created: {len(account_lookup)} accounts")
            return account_lookup
        else:
            print(f"   ⚠️ No chart of accounts found")
            return {}
            
    except Exception as e:
        print(f"   ❌ Error fetching chart of accounts: {e}")
        return {}

def get_chart_of_accounts(accounting_api, tenant_id, base_currency, target_date):
    """Fetch chart of accounts data for export"""
    try:
        accounts_response = accounting_api.get_accounts(xero_tenant_id=tenant_id)
        
        if accounts_response.accounts:
            accounts = []
            for account in accounts_response.accounts:
                account_data = {
                    'account_id': account.account_id,
                    'code': account.code,
                    'name': account.name,
                    'type': str(account.type),
                    'tax_type': account.tax_type,
                    'status': str(account.status),
                    'description': account.description,
                    'class': getattr(account, 'class', None),
                    'system_account': getattr(account, 'system_account', None),
                    'reporting_code': getattr(account, 'reporting_code', None),
                    'reporting_code_name': getattr(account, 'reporting_code_name', None),
                    'updated_date_utc': account.updated_date_utc.isoformat() if account.updated_date_utc else None
                }
                accounts.append(account_data)
            
            return accounts
        else:
            print(f"   ⚠️ No chart of accounts found")
            return []
            
    except Exception as e:
        print(f"   ❌ Error fetching chart of accounts: {e}")
        return []

def parse_report_rows(rows, account_lookup=None, entity_name=""):
    """Parse report rows to extract native account data with enhanced information"""
    accounts = []
    
    # Get total mode (default to 'none' if not set)
    total_mode = globals().get('TOTAL_MODE', 'none')
    
    for row in rows:
        # Process nested rows first
        if hasattr(row, 'rows') and row.rows:
            accounts.extend(parse_report_rows(row.rows, account_lookup, entity_name))
        
        # Process data rows with cells
        elif hasattr(row, 'cells') and row.cells:
            cells = row.cells
            if len(cells) >= 2:
                account_name = cells[0].value if cells[0].value else ""
                balance_value = cells[1].value if len(cells) > 1 and cells[1].value else "0"
                
                # Only include rows with meaningful data (allow zero balances for FX accounts)
                if account_name and account_name.strip() and balance_value:
                    # Check if this is a total/subtotal row
                    account_name_lower = account_name.strip().lower()
                    total_keywords = [
                        'total', 'subtotal', 'sub total', 'sub-total', 'net profit', 'net loss', 
                        'gross profit', 'gross loss', 'total assets', 'total liabilities', 
                        'total equity', 'total income', 'total expenses', 'total revenue',
                        'total current assets', 'total non-current assets', 'total current liabilities',
                        'total non-current liabilities', 'net assets', 'total operating expenses',
                        'total other income', 'total other expenses', 'total cost of sales',
                        'earnings before', 'ebitda', 'ebit'
                    ]
                    
                    is_total_row = any(keyword in account_name_lower for keyword in total_keywords)
                    
                    # Apply total mode filtering
                    should_include = True
                    if total_mode == 'none' and is_total_row:
                        should_include = False  # Skip all total rows when mode is 'none'
                    elif total_mode == 'native':
                        should_include = True  # Include all rows (totals and accounts) when mode is 'native'
                    
                    if should_include:
                        try:
                            # Clean up the balance value
                            balance_str = str(balance_value).replace(',', '').replace('(', '-').replace(')', '')
                            balance = float(balance_str) if balance_str and balance_str != "" else 0.0
                            
                            # Start with native Xero data
                            account_data = {
                                'account_name': account_name.strip(),
                                'balance': balance
                            }
                            
                            # Add entity information
                            if entity_name:
                                account_data['entity'] = entity_name
                            
                            # Enhance with chart of accounts information
                            if account_lookup and account_name.strip() in account_lookup:
                                account_info = account_lookup[account_name.strip()]
                                account_data['code'] = account_info.get('code', '')
                                account_data['type'] = account_info.get('type', '')
                                account_data['tax_type'] = account_info.get('tax_type', '')
                            else:
                                # Provide empty values if not found in chart of accounts
                                account_data['code'] = ''
                                account_data['type'] = ''
                                account_data['tax_type'] = ''
                            
                            accounts.append(account_data)
                        except (ValueError, AttributeError):
                            continue
                    else:
                        # Skip total/summary rows
                        continue
    
    # Return all native Xero API accounts without any manipulation
    return accounts

def get_all_invoices(accounting_api, tenant_id, base_currency, target_date=None, fy_start=None, page_size=100):
    """Fetch invoices for the financial year to date with pagination"""
    all_invoices = []
    page = 1
    
    # Determine date range for filtering
    if target_date and fy_start:
        print(f"   Fetching invoices for FY period: {fy_start.strftime('%d %b %Y')} to {target_date.strftime('%d %b %Y')}...")
    else:
        print(f"   Fetching invoices...")
    
    while True:
        try:
            # Build parameters for invoice query
            # Note: xero-python SDK v4.0.0 doesn't support page_size parameter
            # We'll use page parameter and rely on default page size (usually 100)
            params = {
                'xero_tenant_id': tenant_id,
                'page': page,
                'unitdp': 4,  # 4 decimal places
                'include_archived': True,  # Include archived invoices
                'summary_only': False  # Get full details
            }
            
            # If both target_date and fy_start are provided, filter for financial year to date
            if target_date and fy_start:
                # Filter for invoices within the financial year
                where_clause = f'Date >= DateTime({fy_start.year}, {fy_start.month}, {fy_start.day}) AND Date <= DateTime({target_date.year}, {target_date.month}, {target_date.day})'
                params['where'] = where_clause
            elif target_date:
                # Fallback: filter up to target date only
                where_clause = f'Date <= DateTime({target_date.year}, {target_date.month}, {target_date.day})'
                params['where'] = where_clause
            
            # Call API
            response = accounting_api.get_invoices(**params)
            
            if response.invoices:
                print(f"   Retrieved {len(response.invoices)} invoices from page {page}")
                
                # Process each invoice
                for invoice in response.invoices:
                    invoice_data = {
                        'invoice_id': invoice.invoice_id,
                        'invoice_number': invoice.invoice_number or '',
                        'reference': invoice.reference or '',
                        'type': str(invoice.type) if invoice.type else '',
                        'contact_name': invoice.contact.name if invoice.contact else '',
                        'contact_id': invoice.contact.contact_id if invoice.contact else '',
                        'date': invoice.date.strftime('%Y-%m-%d') if invoice.date else '',
                        'due_date': invoice.due_date.strftime('%Y-%m-%d') if invoice.due_date else '',
                        'status': str(invoice.status) if invoice.status else '',
                        'line_amount_types': str(invoice.line_amount_types) if invoice.line_amount_types else '',
                        'sub_total': float(invoice.sub_total) if invoice.sub_total else 0.0,
                        'total_tax': float(invoice.total_tax) if invoice.total_tax else 0.0,
                        'total': float(invoice.total) if invoice.total else 0.0,
                        'amount_due': float(invoice.amount_due) if invoice.amount_due else 0.0,
                        'amount_paid': float(invoice.amount_paid) if invoice.amount_paid else 0.0,
                        'amount_credited': float(invoice.amount_credited) if invoice.amount_credited else 0.0,
                        'currency_code': str(invoice.currency_code) if invoice.currency_code else base_currency,
                        'currency_rate': float(invoice.currency_rate) if invoice.currency_rate else 1.0,
                        'fully_paid_on_date': invoice.fully_paid_on_date.strftime('%Y-%m-%d') if invoice.fully_paid_on_date else '',
                        'expected_payment_date': invoice.expected_payment_date.strftime('%Y-%m-%d') if invoice.expected_payment_date else '',
                        'planned_payment_date': invoice.planned_payment_date.strftime('%Y-%m-%d') if invoice.planned_payment_date else '',
                        'updated_date_utc': invoice.updated_date_utc.strftime('%Y-%m-%d %H:%M:%S') if invoice.updated_date_utc else '',
                        'sent_to_contact': invoice.sent_to_contact if hasattr(invoice, 'sent_to_contact') else False,
                        'has_attachments': invoice.has_attachments if hasattr(invoice, 'has_attachments') else False
                    }
                    
                    # Add line items summary if available
                    if hasattr(invoice, 'line_items') and invoice.line_items:
                        line_count = len(invoice.line_items)
                        invoice_data['line_item_count'] = line_count
                        
                        # Get first line item description for reference
                        first_line = invoice.line_items[0]
                        invoice_data['first_line_description'] = first_line.description[:100] if first_line.description else ''
                        invoice_data['first_line_quantity'] = float(first_line.quantity) if first_line.quantity else 0.0
                        invoice_data['first_line_unit_amount'] = float(first_line.unit_amount) if first_line.unit_amount else 0.0
                        invoice_data['first_line_line_amount'] = float(first_line.line_amount) if first_line.line_amount else 0.0
                        invoice_data['first_line_account_code'] = first_line.account_code if first_line.account_code else ''
                    else:
                        invoice_data['line_item_count'] = 0
                        invoice_data['first_line_description'] = ''
                        invoice_data['first_line_quantity'] = 0.0
                        invoice_data['first_line_unit_amount'] = 0.0
                        invoice_data['first_line_line_amount'] = 0.0
                        invoice_data['first_line_account_code'] = ''
                    
                    all_invoices.append(invoice_data)
                
                # Check if we've reached the end
                # If we got fewer invoices than expected, we're done
                # Default page size in Xero API is usually 100
                if len(response.invoices) < 100:
                    break
                
                page += 1
            else:
                break
                
        except Exception as e:
            print(f"   Error fetching invoices page {page}: {e}")
            break
    
    print(f"   Total invoices retrieved: {len(all_invoices)}")
    return all_invoices

def get_monthly_pnl_with_ytd(accounting_api, tenant_id, base_currency, target_date, fy_start, account_lookup=None, entity_name=""):
    """Generate monthly P&L breakdown with YTD calculations"""
    
    print(f"   Building monthly P&L from {fy_start.strftime('%b %Y')} to {target_date.strftime('%b %Y')}...")
    
    # Calculate months to process
    months_data = []
    current_month = fy_start.replace(day=1)  # Start of financial year month
    
    while current_month <= target_date:
        # Calculate month end date
        if current_month.month == 12:
            month_end = current_month.replace(year=current_month.year + 1, month=1, day=1) - relativedelta(days=1)
        else:
            month_end = current_month.replace(month=current_month.month + 1, day=1) - relativedelta(days=1)
        
        # Don't go beyond target date
        if month_end > target_date:
            month_end = target_date
            
        print(f"     Fetching {current_month.strftime('%b %Y')} P&L...")
        
        try:
            # Get P&L for this month
            pnl_response = accounting_api.get_report_profit_and_loss(
                xero_tenant_id=tenant_id,
                from_date=current_month,
                to_date=month_end
            )
            
            if pnl_response and pnl_response.reports:
                report = pnl_response.reports[0]
                month_accounts = parse_report_rows(report.rows, account_lookup, entity_name)
                
                months_data.append({
                    'month': current_month.strftime('%Y-%m'),
                    'month_name': current_month.strftime('%b %Y'),
                    'start_date': current_month,
                    'end_date': month_end,
                    'accounts': month_accounts
                })
                
                print(f"       ✅ {len(month_accounts)} accounts processed")
            else:
                print(f"       ⚠️ No P&L data available")
                
        except Exception as e:
            print(f"       ❌ Error fetching {current_month.strftime('%b %Y')}: {e}")
        
        # Move to next month
        if current_month.month == 12:
            current_month = current_month.replace(year=current_month.year + 1, month=1)
        else:
            current_month = current_month.replace(month=current_month.month + 1)
    
    if not months_data:
        print(f"   No monthly P&L data available")
        return None
        
    # Build consolidated monthly P&L with YTD
    print(f"   Building consolidated monthly P&L with YTD calculations...")
    
    # Get all unique accounts across all months (preserve native API ordering)
    all_account_names = []  # Use list to preserve order
    seen_accounts = set()
    
    for month_data in months_data:
        for account in month_data['accounts']:
            account_name = account['account_name']
            # Keep ALL native Xero API accounts including "Foreign Currency Gains and Losses"
            if account_name not in seen_accounts:
                all_account_names.append(account_name)
                seen_accounts.add(account_name)
    
    print(f"   📋 Total unique accounts to consolidate: {len(all_account_names)}")
    print(f"   📋 Using only native accounts from Xero API (no artificial additions)")
    print(f"   📋 Preserving native API ordering (no alphabetical sorting)")
    
    # Create consolidated structure
    consolidated_pnl = []
    
    for account_name in all_account_names:
        # Find the first occurrence of this account to get the enhanced data
        account_template = None
        for month_data in months_data:
            for account in month_data['accounts']:
                if account['account_name'] == account_name:
                    account_template = account
                    break
            if account_template:
                break
        
        # If no template found, try to get from Chart of Accounts
        if not account_template and account_lookup:
            for account in account_lookup.values():
                if account.get('name', '') == account_name:
                    account_template = {
                        'account_name': account.get('name', ''),
                        'code': account.get('code', ''),
                        'type': account.get('type', ''),
                        'tax_type': account.get('tax_type', ''),
                        'entity': entity_name
                    }
                    break
        
        # If still no template found, create a basic one for FX accounts
        if not account_template:
            account_template = {
                'account_name': account_name,
                'code': '',
                'type': '',
                'tax_type': '',
                'entity': entity_name
            }
        
        # Start with enhanced account information
        monthly_row = {
            'account_name': account_name,
            'entity': account_template.get('entity', '') if account_template else '',
            'code': account_template.get('code', '') if account_template else '',
            'type': account_template.get('type', '') if account_template else '',
            'tax_type': account_template.get('tax_type', '') if account_template else ''
        }
        
        # Add monthly columns and prepare for YTD formula
        month_columns_for_formula = []
        
        for month_data in months_data:
            # Find this account in this month (sum all accounts with same name)
            month_balance = 0.0
            for account in month_data['accounts']:
                if account['account_name'] == account_name:
                    month_balance += account['balance']  # Sum all accounts with same name
            
            monthly_row[month_data['month']] = month_balance
            month_columns_for_formula.append(month_data['month'])

        # Store placeholder for YTD - will be converted to Excel formula during export
        monthly_row['ytd_total'] = 0  # Placeholder value, will be replaced with Excel formula
        consolidated_pnl.append(monthly_row)
    
    print(f"   ✅ Monthly P&L consolidated: {len(consolidated_pnl)} account rows across {len(months_data)} months")
    
    # Apply P&L specific sorting (account type order, then alphabetical)
    if consolidated_pnl:
        print(f"   📋 Applying P&L account ordering (type order + alphabetical)...")
        type_order_mapping = load_pl_account_order()
        consolidated_pnl = sort_pl_accounts(consolidated_pnl, type_order_mapping)
        print(f"   ✅ P&L accounts sorted by: {list(type_order_mapping.keys())}")
    
    # No totals row - keeping only individual account data
    
    return {
        'months': [m['month_name'] for m in months_data],
        'data': consolidated_pnl,
        'period_start': fy_start,
        'period_end': target_date
    }

def get_monthly_trial_balance_with_ytd(accounting_api, tenant_id, base_currency, target_date, fy_start, account_lookup=None, entity_name=""):
    """Generate monthly Trial Balance breakdown with YTD calculations"""
    
    print(f"   Building monthly Trial Balance from {fy_start.strftime('%b %Y')} to {target_date.strftime('%b %Y')}...")
    
    # Calculate months to process
    months_data = []
    current_month = fy_start.replace(day=1)  # Start of financial year month
    
    while current_month <= target_date:
        # Calculate month end date
        if current_month.month == 12:
            month_end = current_month.replace(year=current_month.year + 1, month=1, day=1) - relativedelta(days=1)
        else:
            month_end = current_month.replace(month=current_month.month + 1, day=1) - relativedelta(days=1)
        
        # Don't go beyond target date
        if month_end > target_date:
            month_end = target_date
            
        print(f"     Fetching {current_month.strftime('%b %Y')} Trial Balance...")
        
        try:
            # Get Trial Balance for this month end
            tb_response = accounting_api.get_report_trial_balance(
                xero_tenant_id=tenant_id,
                date=month_end
            )
            
            if tb_response and tb_response.reports:
                report = tb_response.reports[0]
                month_accounts = parse_report_rows(report.rows, account_lookup, entity_name)
                
                months_data.append({
                    'month': current_month.strftime('%Y-%m'),
                    'month_name': current_month.strftime('%b %Y'),
                    'start_date': current_month,
                    'end_date': month_end,
                    'accounts': month_accounts
                })
                
                print(f"       ✅ {len(month_accounts)} accounts processed")
            else:
                print(f"       ⚠️ No Trial Balance data available")
                
        except Exception as e:
            print(f"       ❌ Error fetching {current_month.strftime('%b %Y')}: {e}")
        
        # Move to next month
        if current_month.month == 12:
            current_month = current_month.replace(year=current_month.year + 1, month=1)
        else:
            current_month = current_month.replace(month=current_month.month + 1)
    
    if not months_data:
        print(f"   No monthly Trial Balance data available")
        return None
        
    # Build consolidated monthly Trial Balance with YTD
    print(f"   Building consolidated monthly Trial Balance with YTD calculations...")
    
    # Get all unique accounts across all months (preserve native API ordering)
    all_account_names = []  # Use list to preserve order
    seen_accounts = set()
    
    for month_data in months_data:
        for account in month_data['accounts']:
            account_name = account['account_name']
            # Keep ALL native Xero API accounts
            if account_name not in seen_accounts:
                all_account_names.append(account_name)
                seen_accounts.add(account_name)
    
    print(f"   📋 Total unique accounts to consolidate: {len(all_account_names)}")
    print(f"   📋 Using only native accounts from Xero API (no artificial additions)")
    print(f"   📋 Preserving native API ordering (no alphabetical sorting)")
    
    # Create consolidated structure
    consolidated_tb = []
    
    for account_name in all_account_names:
        # Find the first occurrence of this account to get the enhanced data
        account_template = None
        for month_data in months_data:
            for account in month_data['accounts']:
                if account['account_name'] == account_name:
                    account_template = account
                    break
            if account_template:
                break
        
        # If no template found, try to get from Chart of Accounts
        if not account_template and account_lookup:
            for account in account_lookup.values():
                if account.get('name', '') == account_name:
                    account_template = {
                        'account_name': account.get('name', ''),
                        'code': account.get('code', ''),
                        'type': account.get('type', ''),
                        'tax_type': account.get('tax_type', ''),
                        'entity': entity_name
                    }
                    break
        
        # If still no template found, create a basic one
        if not account_template:
            account_template = {
                'account_name': account_name,
                'code': '',
                'type': '',
                'tax_type': '',
                'entity': entity_name
            }
        
        # Start with enhanced account information
        monthly_row = {
            'account_name': account_name,
            'entity': account_template.get('entity', '') if account_template else '',
            'code': account_template.get('code', '') if account_template else '',
            'type': account_template.get('type', '') if account_template else '',
            'tax_type': account_template.get('tax_type', '') if account_template else ''
        }
        
        # Add monthly columns
        for month_data in months_data:
            # Find this account in this month (sum all accounts with same name)
            month_balance = 0.0
            for account in month_data['accounts']:
                if account['account_name'] == account_name:
                    month_balance += account['balance']  # Sum all accounts with same name

            monthly_row[month_data['month']] = month_balance

        # Store placeholder for YTD - will be converted to Excel formula during export
        monthly_row['ytd_total'] = 0  # Placeholder value, will be replaced with Excel formula
        consolidated_tb.append(monthly_row)
    
    print(f"   ✅ Monthly Trial Balance consolidated: {len(consolidated_tb)} account rows across {len(months_data)} months")
    
    return {
        'months': [m['month_name'] for m in months_data],
        'data': consolidated_tb,
        'period_start': fy_start,
        'period_end': target_date
    }

def get_monthly_balance_sheet_with_ytd(accounting_api, tenant_id, base_currency, target_date, fy_start, account_lookup=None, entity_name=""):
    """Generate monthly Balance Sheet breakdown with YTD calculations"""
    
    print(f"   Building monthly Balance Sheet from {fy_start.strftime('%b %Y')} to {target_date.strftime('%b %Y')}...")
    
    # Calculate months to process
    months_data = []
    current_month = fy_start.replace(day=1)  # Start of financial year month
    
    while current_month <= target_date:
        # Calculate month end date
        if current_month.month == 12:
            month_end = current_month.replace(year=current_month.year + 1, month=1, day=1) - relativedelta(days=1)
        else:
            month_end = current_month.replace(month=current_month.month + 1, day=1) - relativedelta(days=1)
        
        # Don't go beyond target date
        if month_end > target_date:
            month_end = target_date
            
        print(f"     Fetching {current_month.strftime('%b %Y')} Balance Sheet...")
        
        try:
            # Get Balance Sheet for this month end
            bs_response = accounting_api.get_report_balance_sheet(
                xero_tenant_id=tenant_id,
                date=month_end
            )
            
            if bs_response and bs_response.reports:
                report = bs_response.reports[0]
                month_accounts = parse_report_rows(report.rows, account_lookup, entity_name)
                
                months_data.append({
                    'month': current_month.strftime('%Y-%m'),
                    'month_name': current_month.strftime('%b %Y'),
                    'start_date': current_month,
                    'end_date': month_end,
                    'accounts': month_accounts
                })
                
                print(f"       ✅ {len(month_accounts)} accounts processed")
            else:
                print(f"       ⚠️ No Balance Sheet data available")
                
        except Exception as e:
            print(f"       ❌ Error fetching {current_month.strftime('%b %Y')}: {e}")
        
        # Move to next month
        if current_month.month == 12:
            current_month = current_month.replace(year=current_month.year + 1, month=1)
        else:
            current_month = current_month.replace(month=current_month.month + 1)
    
    if not months_data:
        print(f"   No monthly Balance Sheet data available")
        return None
        
    # Build consolidated monthly Balance Sheet with YTD
    print(f"   Building consolidated monthly Balance Sheet with YTD calculations...")
    
    # Get all unique accounts across all months (preserve native API ordering)
    all_account_names = []  # Use list to preserve order
    seen_accounts = set()
    
    for month_data in months_data:
        for account in month_data['accounts']:
            account_name = account['account_name']
            # Keep ALL native Xero API accounts
            if account_name not in seen_accounts:
                all_account_names.append(account_name)
                seen_accounts.add(account_name)
    
    print(f"   📋 Total unique accounts to consolidate: {len(all_account_names)}")
    print(f"   📋 Using only native accounts from Xero API (no artificial additions)")
    print(f"   📋 Preserving native API ordering (no alphabetical sorting)")
    
    # Create consolidated structure
    consolidated_bs = []
    
    for account_name in all_account_names:
        # Find the first occurrence of this account to get the enhanced data
        account_template = None
        for month_data in months_data:
            for account in month_data['accounts']:
                if account['account_name'] == account_name:
                    account_template = account
                    break
            if account_template:
                break
        
        # If no template found, try to get from Chart of Accounts
        if not account_template and account_lookup:
            for account in account_lookup.values():
                if account.get('name', '') == account_name:
                    account_template = {
                        'account_name': account.get('name', ''),
                        'code': account.get('code', ''),
                        'type': account.get('type', ''),
                        'tax_type': account.get('tax_type', ''),
                        'entity': entity_name
                    }
                    break
        
        # If still no template found, create a basic one
        if not account_template:
            account_template = {
                'account_name': account_name,
                'code': '',
                'type': '',
                'tax_type': '',
                'entity': entity_name
            }
        
        # Start with enhanced account information
        monthly_row = {
            'account_name': account_name,
            'entity': account_template.get('entity', '') if account_template else '',
            'code': account_template.get('code', '') if account_template else '',
            'type': account_template.get('type', '') if account_template else '',
            'tax_type': account_template.get('tax_type', '') if account_template else ''
        }
        
        # Add monthly columns
        for month_data in months_data:
            # Find this account in this month (sum all accounts with same name)
            month_balance = 0.0
            for account in month_data['accounts']:
                if account['account_name'] == account_name:
                    month_balance += account['balance']  # Sum all accounts with same name

            monthly_row[month_data['month']] = month_balance

        # Store placeholder for YTD - will be converted to Excel formula during export
        monthly_row['ytd_total'] = 0  # Placeholder value, will be replaced with Excel formula
        consolidated_bs.append(monthly_row)
    
    print(f"   ✅ Monthly Balance Sheet consolidated: {len(consolidated_bs)} account rows across {len(months_data)} months")
    
    return {
        'months': [m['month_name'] for m in months_data],
        'data': consolidated_bs,
        'period_start': fy_start,
        'period_end': target_date
    }

def generate_reports_single_company(target_date, run_timestamp, skip_archiving=False):
    """Generate reports for a single company (existing logic)"""
    
    # Create archive-ready filenames
    fy_code = get_fy_end_code(target_date)
    month_year = target_date.strftime("%B %Y")
    
    # Check if a specific company was requested
    requested_company = sys.argv[2].upper() if len(sys.argv) > 2 else None
    if requested_company:
        print(f"Fetching Xero Financial Reports - {requested_company} (Auto-Archive)")
    else:
        print(f"Fetching Xero Financial Reports (Auto-Archive)")
    print(f"Report Date: {target_date.strftime('%d %B %Y')}")
    print(f"Financial Year End: {fy_code}")
    print(f"Timestamp: {run_timestamp.strftime('%H:%M:%S')}")
    print("=" * 70)
    
    # Create output directories
    os.makedirs('../output/xero_data', exist_ok=True)
    os.makedirs('../output/xero_data/archive', exist_ok=True)
    
    # STEP 1: Archive any existing old files first (only if not skipping)
    if not skip_archiving:
        print(f"\nCHECKING FOR EXISTING FILES TO ARCHIVE")
        print("-" * 50)
        
        import glob
        existing_files = glob.glob('../output/xero_data/*.xlsx') + glob.glob('../output/xero_data/*.csv')
        archived_old_files = 0
        
        for filepath in existing_files:
            filename = os.path.basename(filepath)
            # Skip README files
            if filename.lower().startswith('readme'):
                continue
                
            try:
                # Move existing file to archive with timestamp
                timestamp_str = datetime.now().strftime("%H%M%S")
                archive_name = f"archive_{timestamp_str}_{filename}"
                archive_path = f'../output/xero_data/archive/{archive_name}'
                
                os.rename(filepath, archive_path)
                print(f"Archived existing file: {filename} -> {archive_name}")
                archived_old_files += 1
                
            except Exception as e:
                print(f"Could not archive {filename}: {e}")
        
        if archived_old_files == 0:
            print("No existing files found to archive")
        else:
            print(f"Archived {archived_old_files} existing files")
    else:
        print(f"\nSkipping archiving (multi-company mode)")
        print("-" * 40)
    
    # Set up client
    try:
        api_client = setup_xero_client()
        identity_api = IdentityApi(api_client)
        accounting_api = AccountingApi(api_client)
        
        # Get tenant ID
        connections = identity_api.get_connections()
        if not connections:
            print("No Xero connections found")
            return
        
        # Filter connections based on target company if specified
        target_company_code = sys.argv[2].upper() if len(sys.argv) > 2 else None
        target_connection = None
        
        if target_company_code and target_company_code in ['SA', 'MA', 'IND', 'UK']:
            # Find connection matching the target company code
            for conn in connections:
                if get_company_code(conn.tenant_name) == target_company_code:
                    target_connection = conn
                    break
            
            if not target_connection:
                print(f"No connection found for company code: {target_company_code}")
                print("Available companies:")
                for conn in connections:
                    code = get_company_code(conn.tenant_name)
                    print(f"  {code}: {conn.tenant_name}")
                return
            
            tenant_id = target_connection.tenant_id
            tenant_name = target_connection.tenant_name
            print(f"Connected to: {tenant_name} (Target: {target_company_code})")
        else:
            # Use first available connection
            target_connection = connections[0]
            tenant_id = target_connection.tenant_id
            tenant_name = target_connection.tenant_name
            if target_company_code:
                print(f"Connected to: {tenant_name} (Invalid target: {target_company_code}, using auto-select)")
            else:
                print(f"Connected to: {tenant_name} (Auto-selected)")
        
        # Get organization info for base currency
        org_info = get_organization_info(api_client, tenant_id)
        base_currency_raw = org_info.get('base_currency', 'ZAR')
        # Clean up currency code - handle both string and CurrencyCode object
        if hasattr(base_currency_raw, 'value'):
            base_currency = base_currency_raw.value
        else:
            base_currency = str(base_currency_raw).replace('CurrencyCode.', '') if base_currency_raw else 'ZAR'
        
        org_name = org_info.get('name', tenant_name)
        company_code = get_company_code(org_name)
        
        print(f"Base Currency: {base_currency}")
        print(f"Organization: {org_name}")
        print(f"Company Code: {company_code}")
        
    except Exception as e:
        print(f"Connection failed: {e}")
        return
    
    # Financial year context
    fy_start, fy_end = get_financial_year_dates(target_date)
    days_into_fy = (target_date - fy_start).days
    
    print(f"\nFinancial Year: {fy_start.strftime('%d %b %Y')} to {fy_end.strftime('%d %b %Y')}")
    print(f"   Days into FY: {days_into_fy}")
    print(f"   Report as at: {target_date.strftime('%d %B %Y')}")
    
    # Track generated files for automatic archiving
    generated_files = []
    
    # Get Chart of Accounts first for lookup
    print(f"\nGetting Chart of Accounts for lookup...")
    account_lookup = get_chart_of_accounts_lookup(accounting_api, tenant_id)
    
    # 1. Get Balance Sheet (enhanced with monthly data)
    print(f"\nGetting Balance Sheet with Monthly Data ({fy_start.strftime('%b %Y')} to {target_date.strftime('%b %Y')})...")
    try:
        monthly_bs_data = get_monthly_balance_sheet_with_ytd(accounting_api, tenant_id, base_currency, target_date, fy_start, account_lookup, org_name)
        
        if monthly_bs_data and monthly_bs_data['data']:
            print(f"Retrieved Balance Sheet: {len(monthly_bs_data['data'])} account rows across {len(monthly_bs_data['months'])} months")
            
            # Create filename with company code, timestamp and FY code (same filename as before)
            filename = create_archive_filename('balance_sheet', target_date, run_timestamp, company_code)
            
            # Create DataFrame with monthly Balance Sheet data
            df_bs = pd.DataFrame(monthly_bs_data['data'])
            
            # Define columns for export - dynamic based on available months
            base_columns = ['entity', 'account_name', 'code', 'type', 'tax_type']
            month_columns = [month_data['month'] for month_data in monthly_bs_data.get('months_detailed', [])]
            if not month_columns:
                # Fallback: try to find month columns in the data
                month_columns = [col for col in df_bs.columns if col.startswith('2025-')]
            
            export_columns = base_columns + month_columns + ['ytd_total']
            
            # Filter to only include columns that exist in the DataFrame
            export_columns = [col for col in export_columns if col in df_bs.columns]
            
            # Create sheet name
            sheet_name = create_sheet_name('balance_sheet', target_date, company_code)
            
            # Export with centralized formatting
            export_to_excel_with_formatting(df_bs[export_columns], f'../output/xero_data/{filename}', sheet_name, 'financial')
            generated_files.append(filename)
            
            # Show summary for each month and YTD
            months_list = monthly_bs_data['months']
            print(f"   Monthly breakdown: {', '.join(months_list)}")
            print(f"   Account Count: {len(monthly_bs_data['data'])}")
            
        else:
            print(f"   No Balance Sheet data found for the specified period")
            
    except Exception as e:
        print(f"Error getting balance sheet: {e}")
    
    # 2. Get Monthly P&L with YTD
    print(f"\n Getting Monthly P&L with YTD ({fy_start.strftime('%b %Y')} to {target_date.strftime('%b %Y')})...")
    try:
        monthly_pnl_data = get_monthly_pnl_with_ytd(accounting_api, tenant_id, base_currency, target_date, fy_start, account_lookup, org_name)
        
        if monthly_pnl_data and monthly_pnl_data['data']:
            print(f" Retrieved Monthly P&L: {len(monthly_pnl_data['data'])} account rows across {len(monthly_pnl_data['months'])} months")
            
            # Create filename with company code, timestamp and FY code
            filename = create_archive_filename('profit_and_loss', target_date, run_timestamp, company_code)
            
            # Create DataFrame with monthly P&L data
            df_monthly_pnl = pd.DataFrame(monthly_pnl_data['data'])
            
            # Define columns for export - dynamic based on available months
            base_columns = ['entity', 'account_name', 'code', 'type', 'tax_type']
            month_columns = [month_data['month'] for month_data in monthly_pnl_data.get('months_detailed', [])]
            if not month_columns:
                # Fallback: try to find month columns in the data
                month_columns = [col for col in df_monthly_pnl.columns if col.startswith('2025-')]
            
            export_columns = base_columns + month_columns + ['ytd_total']
            
            # Filter to only include columns that exist in the DataFrame
            export_columns = [col for col in export_columns if col in df_monthly_pnl.columns]
            
            # Create sheet name
            sheet_name = create_sheet_name('profit_and_loss', target_date, company_code)
            
            # Export with centralized formatting for monthly P&L
            export_to_excel_with_formatting(df_monthly_pnl[export_columns], f'../output/xero_data/{filename}', sheet_name, 'financial')
            generated_files.append(filename)
            
            # Show summary for each month and YTD
            months_list = monthly_pnl_data['months']
            print(f"   Monthly breakdown: {', '.join(months_list)}")
            
            # Calculate YTD totals (manually calculate since ytd_total will be Excel formulas)
            if monthly_pnl_data['data']:
                # Calculate YTD total across all accounts
                month_cols = [c for c in monthly_pnl_data['data'][0].keys() if c.startswith('2025-')]
                ytd_total = sum([sum([acc.get(month_col, 0) for month_col in month_cols]) for acc in monthly_pnl_data['data']])
                print(f"   YTD Total: {format_currency_display(ytd_total, base_currency)}")
                print(f"   Account Count: {len(monthly_pnl_data['data'])}")
            
        else:
            print(f"   No monthly P&L data found for the specified period")
            
    except Exception as e:
        print(f" Error getting monthly P&L: {e}")
    
    
    
    # 3. Get Trial Balance (enhanced with monthly data)
    print(f"\n Getting Trial Balance with Monthly Data ({fy_start.strftime('%b %Y')} to {target_date.strftime('%b %Y')})...")
    try:
        monthly_tb_data = get_monthly_trial_balance_with_ytd(accounting_api, tenant_id, base_currency, target_date, fy_start, account_lookup, org_name)
        
        if monthly_tb_data and monthly_tb_data['data']:
            print(f" Retrieved Trial Balance: {len(monthly_tb_data['data'])} account rows across {len(monthly_tb_data['months'])} months")
            
            # Create filename with company code, timestamp and FY code (same filename as before)
            filename = create_archive_filename('trial_balance', target_date, run_timestamp, company_code)
            
            # Create DataFrame with monthly Trial Balance data
            df_tb = pd.DataFrame(monthly_tb_data['data'])
            
            # Define columns for export - dynamic based on available months
            base_columns = ['entity', 'account_name', 'code', 'type', 'tax_type']
            month_columns = [month_data['month'] for month_data in monthly_tb_data.get('months_detailed', [])]
            if not month_columns:
                # Fallback: try to find month columns in the data
                month_columns = [col for col in df_tb.columns if col.startswith('2025-')]
            
            export_columns = base_columns + month_columns + ['ytd_total']
            
            # Filter to only include columns that exist in the DataFrame
            export_columns = [col for col in export_columns if col in df_tb.columns]
            
            # Create sheet name
            sheet_name = create_sheet_name('trial_balance', target_date, company_code)
            
            # Export with centralized formatting
            export_to_excel_with_formatting(df_tb[export_columns], f'../output/xero_data/{filename}', sheet_name, 'financial')
            generated_files.append(filename)
            
            # Show summary for each month and YTD
            months_list = monthly_tb_data['months']
            print(f"   Monthly breakdown: {', '.join(months_list)}")
            print(f"   Account Count: {len(monthly_tb_data['data'])}")
            
        else:
            print(f"   No Trial Balance data found for the specified period")
            
    except Exception as e:
        print(f" Error getting trial balance: {e}")
    
    # 4. Get Chart of Accounts
    print(f"\n Getting Chart of Accounts...")
    try:
        accounts_response = accounting_api.get_accounts(
            xero_tenant_id=tenant_id
        )
        
        if accounts_response.accounts:
            print(f" Retrieved Chart of Accounts")
            
            # Process accounts data
            accounts = []
            for account in accounts_response.accounts:
                account_data = {
                    'account_id': account.account_id,
                    'code': account.code,
                    'name': account.name,
                    'type': str(account.type),
                    'tax_type': account.tax_type,
                    'status': str(account.status),
                    'description': account.description,
                    'class': getattr(account, 'class', None),
                    'system_account': getattr(account, 'system_account', None),
                    'reporting_code': getattr(account, 'reporting_code', None),
                    'reporting_code_name': getattr(account, 'reporting_code_name', None),
                    'updated_date_utc': account.updated_date_utc.isoformat() if account.updated_date_utc else None
                }
                accounts.append(account_data)
            
            if accounts:
                # Create filename with company code, timestamp and FY code
                filename = create_archive_filename('chart_of_accounts', target_date, run_timestamp, company_code)
                
                # Export chart of accounts as current file
                df_accounts = pd.DataFrame(accounts)
                export_columns = ['code', 'name', 'type', 'tax_type', 'status', 'description', 'class', 'system_account', 'reporting_code', 'reporting_code_name']
                sheet_name = create_sheet_name('chart_of_accounts', target_date, company_code)
                export_to_excel_with_formatting(df_accounts[export_columns], f'../output/xero_data/{filename}', sheet_name, 'general')
                generated_files.append(filename)
                
                # Analyze account types
                account_types = {}
                for account in accounts:
                    acc_type = account['type']
                    if acc_type not in account_types:
                        account_types[acc_type] = 0
                    account_types[acc_type] += 1
                
                print(f"   Total accounts: {len(accounts)}")
                print(f"   Account types: {len(account_types)}")
                
    except Exception as e:
        print(f" Error getting chart of accounts: {e}")
    
    # 5. Get All Invoices (Year to Date)
    print(f"\n Getting All Invoices for Financial Year to Date...")
    try:
        all_invoices = get_all_invoices(accounting_api, tenant_id, base_currency, target_date, fy_start)
        
        if all_invoices:
            print(f" Retrieved {len(all_invoices)} invoices")
            
            # Create filename with company code, timestamp and FY code
            filename = create_archive_filename('invoices_ytd', target_date, run_timestamp, company_code)
            
            # Format invoice data for proper date display (dd/mm/yyyy)
            formatted_invoices = format_data_for_export(all_invoices, 'invoices')
            
            # Create DataFrame with formatted invoice data
            df_invoices = pd.DataFrame(formatted_invoices)
            
            # Define columns to include in export (with accounting formatting for amounts)
            export_columns = [
                'invoice_number', 'reference', 'type', 'contact_name', 'date', 'due_date', 
                'status', 'sub_total', 'total_tax', 'total', 'amount_due', 'amount_paid', 
                'amount_credited', 'currency_code', 'currency_rate', 'fully_paid_on_date',
                'line_item_count', 'first_line_description', 'first_line_quantity', 
                'first_line_unit_amount', 'first_line_line_amount', 'first_line_account_code',
                'updated_date_utc'
            ]
            
            # Create sheet name
            sheet_name = create_sheet_name('invoices', target_date, company_code)
            
            # Export with centralized formatting for invoices
            export_to_excel_with_formatting(df_invoices[export_columns], f'../output/xero_data/{filename}', sheet_name, 'invoices')
            generated_files.append(filename)
            
            # Calculate summary statistics
            total_invoiced = sum([inv['total'] for inv in all_invoices])
            total_outstanding = sum([inv['amount_due'] for inv in all_invoices])
            total_paid = sum([inv['amount_paid'] for inv in all_invoices])
            
            # Count by status
            status_counts = {}
            for invoice in all_invoices:
                status = invoice['status']
                if status not in status_counts:
                    status_counts[status] = 0
                status_counts[status] += 1
            
            # Count by type
            type_counts = {}
            for invoice in all_invoices:
                inv_type = invoice['type']
                if inv_type not in type_counts:
                    type_counts[inv_type] = 0
                type_counts[inv_type] += 1
            
            print(f"   Total Invoiced: {format_currency_display(total_invoiced, base_currency)}")
            print(f"   Total Outstanding: {format_currency_display(total_outstanding, base_currency)}")
            print(f"   Total Paid: {format_currency_display(total_paid, base_currency)}")
            print(f"   Status Breakdown: {status_counts}")
            print(f"   Type Breakdown: {type_counts}")
            
        else:
            print(f"   No invoices found for the specified date range")
            
    except Exception as e:
        print(f" Error getting invoices: {e}")
    
    # New files remain as current versions (no archiving needed)
    
    # Summary
    print(f"\n SUMMARY - {target_date.strftime('%d %B %Y')}")
    print("-" * 50)
    print(f"Organization: {org_info.get('name', tenant_name)}")
    print(f"Report Date: {target_date.strftime('%d %B %Y')}")
    print(f"Financial Year: {fy_code}")
    print(f"Timestamp: {run_timestamp.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Reports Generated: {len(generated_files)}")
    
    print(f"\n Current reports saved to: ../output/xero_data/")
    
    # List current files
    if generated_files:
        print(f"\n Current Files:")
        for filename in generated_files:
            print(f"    {filename}")
    
    print(f"\n Workflow: Old files archived -> New files become current")
    print(f" Archive contains historical versions with timestamps")

def convert_amount_to_zar(amount, from_currency, fx_reader, entity_name=""):
    """Convert amount to ZAR using FX rates"""
    if not amount or amount == 0:
        return 0.0, 1.0  # Return converted amount and rate used
    
    if from_currency == 'ZAR' or from_currency == 'CurrencyCode.ZAR':
        return float(amount), 1.0
    
    # Clean currency code
    clean_currency = from_currency.replace('CurrencyCode.', '') if from_currency else 'ZAR'
    
    # Get FX rate
    fx_rate = fx_reader.get_fx_rate(clean_currency, 'ZAR')
    
    if fx_rate:
        converted_amount = float(amount) * fx_rate
        return converted_amount, fx_rate
    else:
        print(f"⚠️ No FX rate found for {clean_currency} to ZAR for {entity_name}. Using original amount.")
        return float(amount), 1.0

def create_consolidated_dataframe(original_df, base_currency, fx_reader, entity_name, report_type):
    """Create consolidated DataFrame with entity tracking and ZAR conversion"""
    
    if original_df.empty:
        return original_df, {}
    
    # Create a copy of the original DataFrame
    consolidated_df = original_df.copy()
    
    # Add entity column at the beginning
    consolidated_df.insert(0, 'Entity', entity_name)
    
    # Clean base currency code
    clean_base_currency = base_currency.replace('CurrencyCode.', '') if base_currency else 'ZAR'
    
    # Find amount columns to convert
    amount_columns = []
    for col in consolidated_df.columns:
        if any(keyword in col.lower() for keyword in ['amount', 'balance', 'total', 'value', 'ytd', 'mar ', 'apr ', 'may ', 'jun ', 'jul ', 'aug ', 'sep ', 'oct ', 'nov ', 'dec ']):
            # Skip non-numeric columns
            if consolidated_df[col].dtype in ['float64', 'int64'] or pd.api.types.is_numeric_dtype(consolidated_df[col]):
                amount_columns.append(col)
    
    # Track FX rates used
    fx_rates_used = {}
    
    # Convert amount columns
    for col in amount_columns:
        # Convert amounts to ZAR (no native columns needed)
        converted_amounts = []
        fx_rates = []
        
        for amount in consolidated_df[col]:
            converted_amount, fx_rate = convert_amount_to_zar(amount, clean_base_currency, fx_reader, entity_name)
            converted_amounts.append(converted_amount)
            fx_rates.append(fx_rate)
        
        # Replace original column with ZAR amounts (keep same column name for simplicity)
        consolidated_df[col] = converted_amounts
        
        # Track the FX rate used for the FX rates sheet
        if clean_base_currency != 'ZAR':
            # Get the actual FX rate from the reader, regardless of whether amounts are 0
            actual_fx_rate = fx_reader.get_fx_rate(clean_base_currency, 'ZAR')
            if actual_fx_rate:
                fx_rates_used[f"{clean_base_currency}_ZAR"] = actual_fx_rate
            else:
                # Fallback to 1.0 only if no rate is available
                fx_rates_used[f"{clean_base_currency}_ZAR"] = 1.0
    
    return consolidated_df, fx_rates_used

def create_fx_rates_sheet(fx_rates_used, target_date):
    """Create a DataFrame with FX rates used for consolidation"""
    
    if not fx_rates_used:
        return pd.DataFrame()
    
    fx_data = []
    for currency_pair, rate in fx_rates_used.items():
        from_currency, to_currency = currency_pair.split('_')
        fx_data.append({
            'From_Currency': from_currency,
            'To_Currency': to_currency,
            'Exchange_Rate': rate,
            'Conversion_Example': f"1 {from_currency} = {rate:.4f} {to_currency}",
            'Date_Applied': target_date.strftime('%Y-%m-%d'),
            'Source': 'fx/FX.xlsx',
            'Notes': f'Used to convert all {from_currency} amounts to {to_currency}'
        })
    
    # Sort by currency for better readability
    fx_df = pd.DataFrame(fx_data)
    if not fx_df.empty:
        fx_df = fx_df.sort_values('From_Currency')
    
    return fx_df

def main():
    """Main function with multi-company support and consolidation options"""
    
    # Parse command line arguments
    if len(sys.argv) < 2:
        print("Usage: python get_xero_reports.py <date> [options]")
        print("Examples:")
        print("  python get_xero_reports.py \"June 2025\"                           # All companies, multi-sheet")
        print("  python get_xero_reports.py \"June 2025\" --mode native             # All companies, multi-sheet (default)")
        print("  python get_xero_reports.py \"June 2025\" --mode consolidated       # Multi-sheet + consolidated sheet")
        print("  python get_xero_reports.py \"June 2025\" --entity SA               # Only SA company")
        print("  python get_xero_reports.py \"June 2025\" --entity UK --mode consolidated # UK with consolidated sheet")
        print("  python get_xero_reports.py \"June 2025\" --total-mode native --entity SA # SA with totals")
        print("")
        print("Options:")
        print("  --mode <mode>       Output format (native, consolidated)")
        print("  --entity <code>     Filter to specific entity (SA, MA, IND, UK)")
        print("  --total-mode <mode> Control totals display (none, native)")
        print("")
        print("Mode options:")
        print("  - native: Multi-sheet files with one sheet per company (default)")
        print("  - consolidated: Multi-sheet files PLUS consolidated sheet (all entities in ZAR)")
        print("")
        print("Total modes:")
        print("  - none: No total/subtotal rows (default)")
        print("  - native: Include Xero's native calculated totals")
        return
    
    # Parse parameters
    date_input = sys.argv[1]
    
    # Parse mode parameter
    mode = 'native'  # default
    if '--mode' in sys.argv:
        try:
            mode_index = sys.argv.index('--mode')
            if mode_index + 1 < len(sys.argv):
                mode = sys.argv[mode_index + 1].lower()
        except (IndexError, ValueError):
            mode = 'native'
    
    # Parse entity parameter
    target_company = None
    if '--entity' in sys.argv:
        try:
            entity_index = sys.argv.index('--entity')
            if entity_index + 1 < len(sys.argv):
                target_company = sys.argv[entity_index + 1].upper()
        except (IndexError, ValueError):
            target_company = None
    
    # Parse total-mode parameter
    total_mode = 'none'  # default
    if '--total-mode' in sys.argv:
        try:
            total_mode_index = sys.argv.index('--total-mode')
            if total_mode_index + 1 < len(sys.argv):
                total_mode = sys.argv[total_mode_index + 1].lower()
        except (IndexError, ValueError):
            total_mode = 'none'
    
    target_date = parse_date_input(date_input)
    if not target_date:
        return
    
    # Validate mode
    if mode not in ['native', 'consolidated']:
        print(f"❌ Invalid mode: {mode}")
        print("Valid modes: native, consolidated")
        return
    
    # Validate total mode
    if total_mode not in ['none', 'native']:
        print(f"❌ Invalid total mode: {total_mode}")
        print("Valid modes: none, native")
        return
    
    # Validate entity code if specified
    if target_company and target_company not in ['SA', 'MA', 'IND', 'UK']:
        print(f"❌ Invalid entity code: {target_company}")
        print("Valid entity codes: SA, MA, IND, UK")
        return
    
    # Initialize FX reader if consolidation is needed
    fx_reader = None
    if mode == 'consolidated':
        print("💱 Initializing FX rate reader...")
        fx_reader = FXRateReader()
        if not fx_reader.load_fx_data():
            print("❌ Failed to load FX data. Consolidated sheets will not be generated.")
            fx_reader = None
        else:
            available_currencies = fx_reader.get_available_currencies()
            print(f"✅ FX data loaded. Available currencies: {', '.join(available_currencies)}")
    
    # Create timestamp for this run
    run_timestamp = datetime.now()
    
    print(f"📊 Output mode: {mode.upper()}")
    if target_company:
        print(f"🎯 Target entity: {target_company}")
    if fx_reader:
        fx_summary = fx_reader.get_fx_rates_summary()
        print(f"💱 FX rates loaded: {len(fx_summary)} currency pairs")
    
    # Generate reports based on mode
    if mode == 'native':
        # Generate multi-sheet reports only
        print(f"🌍 Generating MULTI-SHEET reports")
        print("=" * 50)
        generate_reports_all_companies_native(target_date, run_timestamp, total_mode, target_company)
        
    elif mode == 'consolidated':
        # Generate multi-sheet reports with consolidated sheets
        print(f"🏦 Generating MULTI-SHEET reports with CONSOLIDATED sheets")
        print("=" * 50)
        generate_reports_all_companies_native(target_date, run_timestamp, total_mode, target_company, fx_reader)

def generate_consolidated_reports(target_date, run_timestamp, fx_reader, skip_archiving=False, total_mode='none'):
    """Generate consolidated reports combining all entities in ZAR"""
    
    # Store total_mode globally for use in parse_report_rows
    global TOTAL_MODE
    TOTAL_MODE = total_mode
    
    if not fx_reader:
        print("❌ FX reader is required for consolidated reports")
        return
    
    print("🏦 Collecting data from all entities for consolidation...")
    
    # Get all available companies
    try:
        from xero_python.identity import IdentityApi
        api_client = setup_xero_client()
        identity_api = IdentityApi(api_client)
        connections = identity_api.get_connections()
        
        if not connections:
            print("❌ No Xero connections found")
            return
        
        # Process all companies
        companies_to_process = []
        for conn in connections:
            code = get_company_code(conn.tenant_name)
            if code not in [c['code'] for c in companies_to_process]:
                companies_to_process.append({
                    'code': code, 
                    'name': conn.tenant_name,
                    'tenant_id': conn.tenant_id
                })
        
        if not companies_to_process:
            print(f"❌ No companies found for consolidation")
            return
            
        print(f"📊 Consolidating data from {len(companies_to_process)} companies: {', '.join([c['code'] for c in companies_to_process])}")
        
        # Archive existing files (unless skipping)
        if not skip_archiving:
            print(f"\n🗂️ ARCHIVING EXISTING FILES")
            print("-" * 40)
            
            import os
            import glob
            os.makedirs('../output/xero_data', exist_ok=True)
            os.makedirs('../output/xero_data/archive', exist_ok=True)
            
            existing_files = glob.glob('../output/xero_data/*.xlsx') + glob.glob('../output/xero_data/*.csv')
            archived_files = 0
            
            for filepath in existing_files:
                filename = os.path.basename(filepath)
                if filename.lower().startswith('readme'):
                    continue
                    
                try:
                    timestamp_str = run_timestamp.strftime("%H%M%S")
                    archive_name = f"archive_{timestamp_str}_{filename}"
                    archive_path = f'../output/xero_data/archive/{archive_name}'
                    os.rename(filepath, archive_path)
                    print(f"Archived: {filename} -> {archive_name}")
                    archived_files += 1
                except Exception as e:
                    print(f"Could not archive {filename}: {e}")
            
            if archived_files == 0:
                print("No existing files found to archive")
            else:
                print(f"✅ Archived {archived_files} existing files")
            
            print()
        else:
            print(f"\n⏭️ Skipping archiving (already handled in previous step)")
            print()
            
            # Still need to create directories
            import os
            os.makedirs('../output/xero_data', exist_ok=True)
            os.makedirs('../output/xero_data/archive', exist_ok=True)
        
        # Collect data from all companies
        consolidated_data = {
            'balance_sheets': [],
            'pnl_data': [],
            'trial_balances': [],
            'chart_of_accounts': [],
            'invoices': []
        }
        
        all_fx_rates_used = {}
        
        # Process each company
        for company in companies_to_process:
            print(f"\n📊 Collecting data from {company['code']} - {company['name']}")
            
            # Get company data (this would need to be extracted from the existing functions)
            company_data = collect_company_data_for_consolidation(
                company['tenant_id'], company['code'], company['name'], 
                target_date, fx_reader
            )
            
            if company_data:
                # Add to consolidated data
                for report_type, data in company_data.items():
                    if report_type == 'fx_rates_used':
                        all_fx_rates_used.update(data)
                    else:
                        consolidated_data[report_type].extend(data)
        
        # Generate consolidated reports
        generate_consolidated_excel_files(consolidated_data, all_fx_rates_used, target_date, run_timestamp)
        
        print("\n✅ Consolidated reports generation completed")
        
    except Exception as e:
        print(f"❌ Error generating consolidated reports: {e}")
        return

def collect_company_data_for_consolidation(tenant_id, company_code, company_name, target_date, fx_reader):
    """Collect data from a single company for consolidation"""
    
    try:
        # Set up API client
        api_client = setup_xero_client()
        accounting_api = AccountingApi(api_client)
        
        # Get organization info for base currency
        org_response = accounting_api.get_organisations(tenant_id)
        if not org_response.organisations:
            print(f"   ❌ No organization info found for {company_code}")
            return None
            
        org = org_response.organisations[0]
        base_currency = str(org.base_currency)
        
        print(f"   📊 Base Currency: {base_currency}")
        
        # Initialize data collection
        company_data = {
            'balance_sheets': [],
            'pnl_data': [],
            'trial_balances': [],
            'chart_of_accounts': [],
            'invoices': [],
            'fx_rates_used': {}
        }
        
        # Get chart of accounts lookup for this company
        account_lookup = get_chart_of_accounts_lookup(accounting_api, tenant_id)
        
        # Collect Balance Sheet
        print(f"   📋 Collecting Balance Sheet...")
        try:
            bs_response = accounting_api.get_report_balance_sheet(tenant_id, date=target_date)
            if bs_response and hasattr(bs_response, 'reports') and bs_response.reports:
                bs_accounts = parse_report_rows(bs_response.reports[0].rows, account_lookup, company_name)
                if bs_accounts:
                    bs_df = pd.DataFrame(bs_accounts)
                    consolidated_bs, fx_rates = create_consolidated_dataframe(
                        bs_df, base_currency, fx_reader, company_name, 'balance_sheet'
                    )
                    company_data['balance_sheets'].append(consolidated_bs)
                    company_data['fx_rates_used'].update(fx_rates)
                    print(f"   ✅ Balance Sheet: {len(bs_accounts)} accounts")
        except Exception as e:
            print(f"   ⚠️ Balance Sheet error: {e}")
        
        # Collect Trial Balance
        print(f"   📋 Collecting Trial Balance...")
        try:
            tb_response = accounting_api.get_report_trial_balance(tenant_id, date=target_date)
            if tb_response and hasattr(tb_response, 'reports') and tb_response.reports:
                tb_accounts = parse_report_rows(tb_response.reports[0].rows, account_lookup, company_name)
                if tb_accounts:
                    tb_df = pd.DataFrame(tb_accounts)
                    consolidated_tb, fx_rates = create_consolidated_dataframe(
                        tb_df, base_currency, fx_reader, company_name, 'trial_balance'
                    )
                    company_data['trial_balances'].append(consolidated_tb)
                    company_data['fx_rates_used'].update(fx_rates)
                    print(f"   ✅ Trial Balance: {len(tb_accounts)} accounts")
        except Exception as e:
            print(f"   ⚠️ Trial Balance error: {e}")
        
        # Collect Profit & Loss
        print(f"   📋 Collecting Profit & Loss...")
        try:
            # Get financial year dates
            fy_start, fy_end = get_financial_year_dates(target_date)
            
            # Get monthly P&L with YTD (same as native mode)
            monthly_pnl_data = get_monthly_pnl_with_ytd(accounting_api, tenant_id, base_currency, target_date, fy_start, account_lookup, company_name)
            
            if monthly_pnl_data and 'data' in monthly_pnl_data and monthly_pnl_data['data']:
                pnl_df = pd.DataFrame(monthly_pnl_data['data'])
                consolidated_pnl, fx_rates = create_consolidated_dataframe(
                    pnl_df, base_currency, fx_reader, company_name, 'profit_and_loss'
                )
                company_data['pnl_data'].append(consolidated_pnl)
                company_data['fx_rates_used'].update(fx_rates)
                print(f"   ✅ Profit & Loss: {len(monthly_pnl_data['data'])} accounts across {len(monthly_pnl_data.get('months', []))} months")
            else:
                print(f"   ⚠️ No P&L data found for the specified period")
        except Exception as e:
            print(f"   ⚠️ Profit & Loss error: {e}")
        
        # Collect Chart of Accounts
        print(f"   📋 Collecting Chart of Accounts...")
        try:
            accounts_response = accounting_api.get_accounts(tenant_id)
            if accounts_response and accounts_response.accounts:
                accounts_data = []
                for account in accounts_response.accounts:
                    accounts_data.append({
                        'account_code': account.code,
                        'account_name': account.name,
                        'account_type': str(account.type) if account.type else '',
                        'tax_type': str(account.tax_type) if account.tax_type else '',
                        'status': str(account.status) if account.status else '',
                        'description': account.description or '',
                        'account_class': str(account.account_class) if hasattr(account, 'account_class') and account.account_class else '',
                        'system_account': str(account.system_account) if hasattr(account, 'system_account') and account.system_account else '',
                        'reporting_code': account.reporting_code if hasattr(account, 'reporting_code') else '',
                        'reporting_code_name': account.reporting_code_name if hasattr(account, 'reporting_code_name') else '',
                        'currency': base_currency
                    })
                
                if accounts_data:
                    coa_df = pd.DataFrame(accounts_data)
                    # Add entity column
                    coa_df.insert(0, 'Entity', company_name)
                    company_data['chart_of_accounts'].append(coa_df)
                    print(f"   ✅ Chart of Accounts: {len(accounts_data)} accounts")
        except Exception as e:
            print(f"   ⚠️ Chart of Accounts error: {e}")
        
        return company_data
        
    except Exception as e:
        print(f"   ❌ Error collecting data for {company_code}: {e}")
        return None

def generate_consolidated_excel_files(consolidated_data, fx_rates_used, target_date, run_timestamp):
    """Generate consolidated Excel files from collected data"""
    
    print("📊 Generating consolidated Excel files...")
    
    # Create consolidated reports
    reports_to_generate = []
    
    # Balance Sheet
    if consolidated_data['balance_sheets']:
        combined_bs = pd.concat(consolidated_data['balance_sheets'], ignore_index=True)
        if not combined_bs.empty:
            reports_to_generate.append(('balance_sheet', combined_bs, 'Consolidated Balance Sheet'))
    
    # Profit & Loss
    if consolidated_data['pnl_data']:
        combined_pnl = pd.concat(consolidated_data['pnl_data'], ignore_index=True)
        if not combined_pnl.empty:
            reports_to_generate.append(('profit_and_loss', combined_pnl, 'Consolidated Profit & Loss'))
    
    # Trial Balance  
    if consolidated_data['trial_balances']:
        combined_tb = pd.concat(consolidated_data['trial_balances'], ignore_index=True)
        if not combined_tb.empty:
            reports_to_generate.append(('trial_balance', combined_tb, 'Consolidated Trial Balance'))
    
    # Chart of Accounts
    if consolidated_data['chart_of_accounts']:
        combined_coa = pd.concat(consolidated_data['chart_of_accounts'], ignore_index=True)
        if not combined_coa.empty:
            reports_to_generate.append(('chart_of_accounts', combined_coa, 'Consolidated Chart of Accounts'))
    
    # FX Rates Sheet
    fx_rates_df = create_fx_rates_sheet(fx_rates_used, target_date)
    
    # Generate Excel files with proper formatting
    for report_type, data_df, description in reports_to_generate:
        try:
            # Create filename
            filename = create_archive_filename(f'{report_type}_CONSOLIDATED', target_date, run_timestamp, 'CONS')
            filepath = f'../output/xero_data/{filename}'
            
            # Use the same formatting function as native reports
            sheet_name = create_sheet_name(report_type, target_date, 'CONS', is_consolidated=True)
            
            # Determine data type for formatting
            data_type = 'financial' if report_type in ['balance_sheet', 'profit_and_loss', 'trial_balance'] else 'general'
            
            # Export main data with formatting
            export_to_excel_with_formatting(data_df, filepath, sheet_name, data_type)
            
            # Add FX rates sheet if we have FX data
            if not fx_rates_df.empty:
                # Append FX sheet to the existing file
                with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
                    fx_sheet_name = create_sheet_name('fx_rates', target_date, 'CONS', is_consolidated=True)
                    fx_rates_df.to_excel(writer, sheet_name=fx_sheet_name, index=False)
                
            print(f"   ✅ Generated: {filename}")
            
        except Exception as e:
            print(f"   ❌ Error generating {report_type}: {e}")
    
    print(f"📊 Consolidated reports saved to: ../output/xero_data/")

def generate_reports_all_companies_native(target_date, run_timestamp, total_mode='none', target_company=None, fx_reader=None):
    """Generate multi-sheet reports with one sheet per company (or single entity if specified)"""
    
    # Store total_mode globally for use in parse_report_rows
    global TOTAL_MODE
    TOTAL_MODE = total_mode
    
    # Get all available companies
    try:
        from xero_python.identity import IdentityApi
        from xero_python.accounting import AccountingApi
        api_client = setup_xero_client()
        identity_api = IdentityApi(api_client)
        accounting_api = AccountingApi(api_client)
        connections = identity_api.get_connections()
        
        if not connections:
            print("❌ No Xero connections found")
            return
        
        # Get unique company codes and connections
        companies = []
        for conn in connections:
            code = get_company_code(conn.tenant_name)
            if code not in [c['code'] for c in companies]:
                companies.append({
                    'code': code, 
                    'name': conn.tenant_name,
                    'tenant_id': conn.tenant_id
                })
        
        # Filter by target company if specified
        if target_company:
            available_entities = [c['code'] for c in companies]
            companies = [c for c in companies if c['code'] == target_company]
            if not companies:
                print(f"❌ Entity '{target_company}' not found in connected companies")
                print(f"Available entities: {', '.join(available_entities)}")
                return
            print(f"📊 Filtering to specific entity: {target_company}")
        
        print(f"📊 Processing {len(companies)} companies: {', '.join([c['code'] for c in companies])}")
        
        # Archive existing files ONCE at the beginning
        print(f"\n🗂️ ARCHIVING EXISTING FILES")
        print("-" * 40)
        
        import os
        import glob
        os.makedirs('../output/xero_data', exist_ok=True)
        os.makedirs('../output/xero_data/archive', exist_ok=True)
        
        existing_files = glob.glob('../output/xero_data/*.xlsx') + glob.glob('../output/xero_data/*.csv')
        archived_old_files = 0
        
        for filepath in existing_files:
            filename = os.path.basename(filepath)
            if filename.lower().startswith('readme'):
                continue
                
            try:
                timestamp_str = run_timestamp.strftime("%H%M%S")
                archive_name = f"archive_{timestamp_str}_{filename}"
                archive_path = f'../output/xero_data/archive/{archive_name}'
                os.rename(filepath, archive_path)
                print(f"Archived: {filename} -> {archive_name}")
                archived_old_files += 1
            except Exception as e:
                print(f"Could not archive {filename}: {e}")
            
        if archived_old_files == 0:
            print("No existing files found to archive")
        else:
            print(f"✅ Archived {archived_old_files} existing files")
        
        print()
        
        # Collect data from all companies
        print(f"📊 COLLECTING DATA FROM ALL COMPANIES")
        print("=" * 50)
        
        # Data storage for each report type
        balance_sheet_data = {}
        pnl_data = {}
        trial_balance_data = {}
        chart_of_accounts_data = {}
        invoices_data = {}
        
        # Process each company
        for i, company in enumerate(companies, 1):
            print(f"\n🏢 COMPANY {i}/{len(companies)}: {company['code']} - {company['name']}")
            print("-" * 60)
            
            try:
                tenant_id = company['tenant_id']
                company_code = company['code']
                org_name = company['name']
                
                # Get organization info for base currency
                org_info = get_organization_info(api_client, tenant_id)
                base_currency = org_info.get('base_currency', 'ZAR')
                
                # Get financial year info
                fy_start, fy_end = get_financial_year_dates(target_date)
                
                # Get Chart of Accounts for lookup
                account_lookup = get_chart_of_accounts_lookup(accounting_api, tenant_id)
                print(f"   📋 Chart of Accounts lookup created: {len(account_lookup)} accounts")
                
                # 1. Balance Sheet
                print(f"   📊 Getting Balance Sheet...")
                monthly_bs_data = get_monthly_balance_sheet_with_ytd(accounting_api, tenant_id, base_currency, target_date, fy_start, account_lookup, org_name)
                if monthly_bs_data and monthly_bs_data['data']:
                    df_bs = pd.DataFrame(monthly_bs_data['data'])
                    balance_sheet_data[company_code] = df_bs
                    print(f"      ✅ {len(df_bs)} accounts across {len(monthly_bs_data['months'])} months")
                
                # 2. P&L
                print(f"   📊 Getting P&L...")
                monthly_pnl_data = get_monthly_pnl_with_ytd(accounting_api, tenant_id, base_currency, target_date, fy_start, account_lookup, org_name)
                if monthly_pnl_data and monthly_pnl_data['data']:
                    df_pnl = pd.DataFrame(monthly_pnl_data['data'])
                    pnl_data[company_code] = df_pnl
                    print(f"      ✅ {len(df_pnl)} accounts across {len(monthly_pnl_data['months'])} months")
                
                # 3. Trial Balance
                print(f"   📊 Getting Trial Balance...")
                monthly_tb_data = get_monthly_trial_balance_with_ytd(accounting_api, tenant_id, base_currency, target_date, fy_start, account_lookup, org_name)
                if monthly_tb_data and monthly_tb_data['data']:
                    df_tb = pd.DataFrame(monthly_tb_data['data'])
                    trial_balance_data[company_code] = df_tb
                    print(f"      ✅ {len(df_tb)} accounts across {len(monthly_tb_data['months'])} months")
                
                # 4. Chart of Accounts
                print(f"   📊 Getting Chart of Accounts...")
                coa_data = get_chart_of_accounts(accounting_api, tenant_id, base_currency, target_date)
                if coa_data:
                    df_coa = pd.DataFrame(coa_data)
                    chart_of_accounts_data[company_code] = df_coa
                    print(f"      ✅ {len(df_coa)} accounts")
                
                # 5. Invoices (if any)
                print(f"   📊 Getting Invoices...")
                invoice_data = get_all_invoices(accounting_api, tenant_id, base_currency, target_date, fy_start)
                if invoice_data:
                    df_inv = pd.DataFrame(invoice_data)
                    invoices_data[company_code] = df_inv
                    print(f"      ✅ {len(df_inv)} invoices")
                
                print(f"   ✅ {company_code} data collection completed")
                
            except Exception as e:
                print(f"   ❌ Error collecting {company['code']} data: {e}")
        
        # Generate consolidated files
        print(f"\n📁 GENERATING MULTI-SHEET EXCEL FILES")
        print("=" * 50)
        
        report_types = [
            ('balance_sheet', balance_sheet_data, 'financial'),
            ('profit_and_loss', pnl_data, 'financial'),
            ('trial_balance', trial_balance_data, 'financial'),
            ('chart_of_accounts', chart_of_accounts_data, 'general'),
            ('invoices_ytd', invoices_data, 'general')
        ]
        
        generated_files = []
        
        for report_type, data_dict, data_format in report_types:
            if data_dict:  # Only create file if we have data
                filename = create_archive_filename(report_type, target_date, run_timestamp)
                filepath = f'../output/xero_data/{filename}'
                
                print(f"   📄 Creating {report_type}...")
                
                # Create multi-sheet file with company sheets
                export_multi_company_to_excel(data_dict, filepath, data_format)
                
                # Add consolidated sheet if fx_reader is provided
                if fx_reader and report_type in ['balance_sheet', 'profit_and_loss', 'trial_balance']:
                    print(f"      💱 Adding consolidated sheet to {report_type}...")
                    add_consolidated_sheet_to_file(filepath, data_dict, fx_reader, target_date, report_type, data_format)
                
                generated_files.append(filename)
                
                # Show summary
                company_counts = {code: len(df) if df is not None and not df.empty else 0 for code, df in data_dict.items()}
                total_rows = sum(company_counts.values())
                print(f"      ✅ {filename}")
                if fx_reader and report_type in ['balance_sheet', 'profit_and_loss', 'trial_balance']:
                    print(f"         Multi-sheet + CONSOLIDATED sheet")
                print(f"         Total rows: {total_rows} across {len([c for c in company_counts.values() if c > 0])} companies")
                for code, count in company_counts.items():
                    if count > 0:
                        print(f"         {code}: {count} rows")
        
        print(f"\n🎉 MULTI-SHEET REPORTS COMPLETED")
        print(f"Generated {len(generated_files)} consolidated files for: {', '.join([c['code'] for c in companies])}")
        
        # Show final file list
        if generated_files:
            print(f"\n📁 Final reports in ../output/xero_data/:")
            for filename in generated_files:
                print(f"   📄 {filename}")
        
    except Exception as e:
        print(f"❌ Error in multi-sheet generation: {e}")

if __name__ == "__main__":
    main()
