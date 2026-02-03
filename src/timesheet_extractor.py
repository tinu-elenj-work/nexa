#!/usr/bin/env python3
"""
Nexa Timesheet Extractor
========================

A utility to extract timesheet data per resource per allocation for a range of months from ElapseIT.

Features:
- Dynamic date range specification
- Excel output with two sheets:
  1. Grouping by Client -> Allocation -> Resource (broken down by month)
  2. Grouping by Resource -> Client -> Allocation (broken down by month)

Usage:
    python timesheet_extractor.py

Author: AI Assistant
Date: 2024
"""

import requests
import pandas as pd
from datetime import datetime, timedelta, date
import calendar
from typing import Dict, List, Optional, Tuple
import os
import glob
import argparse
from urllib.parse import quote
import json

# Import existing API client and config
from elapseit_api_client import ElapseITAPIClient
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from config import ELAPSEIT_CONFIG

# Import centralized color configuration
from config.color_scheme import (
    get_category_color, get_chart_color, get_status_color,
    get_plotly_marker_config, get_plotly_line_config,
    CATEGORY_COLORS, CHART_COLORS
)

# Fix Windows console encoding so emoji/Unicode print without crashing
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass


class ElapseITTimesheetExtractor:
    """Extract timesheet data from ElapseIT API and generate Excel reports."""
    
    def __init__(self):
        """Initialize the extractor with API configuration."""
        # Initialize API client with existing config
        self.client = ElapseITAPIClient(
            domain=ELAPSEIT_CONFIG['domain'],
            username=ELAPSEIT_CONFIG['username'],
            password=ELAPSEIT_CONFIG['password'],
            timezone=ELAPSEIT_CONFIG['timezone']
        )
        
        # Get project root directory (parent of src/)
        script_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.dirname(script_dir)

        # Create directories
        self.data_dir = os.path.join(project_root, "output", "elapseIT_data")
        self.archive_dir = os.path.join(project_root, "output", "elapseIT_data", "archive")
        os.makedirs(self.data_dir, exist_ok=True)
        os.makedirs(self.archive_dir, exist_ok=True)
        
    def authenticate(self) -> bool:
        """
        Authenticate with ElapseIT API using existing client.
        
        Returns:
            bool: True if authentication successful
        """
        try:
            print("🔐 Authenticating with ElapseIT...")
            if self.client.authenticate():
                print("✅ Authentication successful!")
                return True
            else:
                print("❌ Authentication failed!")
                return False
                
        except Exception as e:
            print(f"❌ Authentication error: {str(e)}")
            return False
    
    def create_archive_filename(self, report_type: str, start_date: str, end_date: str, timestamp: str = None) -> str:
        """
        Create archive filename with proper timestamp.
        
        Args:
            report_type: Type of report (e.g., 'timesheets')
            start_date: Start date in YYYY-MM-DD format
            end_date: End date in YYYY-MM-DD format
            timestamp: Optional timestamp, defaults to current time
            
        Returns:
            str: Archive filename
        """
        if not timestamp:
            timestamp = datetime.now().strftime("%H%M%S")
        
        # Format: timesheets_YYYYMMDD_to_YYYYMMDD_HHMMSS.xlsx
        start_str = start_date.replace('-', '')
        end_str = end_date.replace('-', '')
        
        filename = f"{report_type}_{start_str}_to_{end_str}_{timestamp}.xlsx"
        return filename
    
    def move_to_archive(self, filename: str) -> bool:
        """
        Move file to archive directory (automatic archiving).
        
        Args:
            filename: Name of file to archive
            
        Returns:
            bool: True if archived successfully
        """
        source_path = os.path.join(self.data_dir, filename)
        archive_path = os.path.join(self.archive_dir, filename)
        
        try:
            if os.path.exists(source_path):
                os.rename(source_path, archive_path)
                print(f"📁 Auto-archived: {filename}")
                return True
            else:
                print(f"⚠️ File not found for archiving: {filename}")
                return False
        except Exception as e:
            print(f"❌ Error archiving: {e}")
            return False
    
    def archive_existing_files(self) -> int:
        """
        Archive any existing timesheet files in the main directory.
        
        Returns:
            int: Number of files archived
        """
        existing_files = glob.glob(os.path.join(self.data_dir, "*.xlsx"))
        archived_count = 0
        
        for filepath in existing_files:
            filename = os.path.basename(filepath)
            
            # Skip README files
            if filename.lower().startswith('readme'):
                continue
                
            try:
                # Move existing file to archive with archive prefix
                timestamp_str = datetime.now().strftime("%H%M%S")
                archive_name = f"archive_{timestamp_str}_{filename}"
                archive_path = os.path.join(self.archive_dir, archive_name)
                
                os.rename(filepath, archive_path)
                print(f"📁 Archived existing file: {filename} → {archive_name}")
                archived_count += 1
                
            except Exception as e:
                print(f"⚠️ Could not archive {filename}: {e}")
        
        return archived_count
    
    def get_date_range_filter(self, start_date: str, end_date: str) -> str:
        """
        Create OData filter for date range with proper boundary handling.
        
        Args:
            start_date: Start date in YYYY-MM-DD format
            end_date: End date in YYYY-MM-DD format
            
        Returns:
            str: OData filter string
        """
        # Use beginning of day for start and end of day for end to ensure we capture all records
        # This fixes the boundary issue where records with timestamps later in the day were excluded
        filter_str = f"Day ge {start_date}T00:00:00Z and Day le {end_date}T23:59:59Z"
        return filter_str
    
    def fetch_vacation_records(self, start_date: str, end_date: str) -> List[Dict]:
        """
        Fetch vacation records from ElapseIT API for the specified date range.
        
        Args:
            start_date: Start date in YYYY-MM-DD format
            end_date: End date in YYYY-MM-DD format
            
        Returns:
            List[Dict]: List of vacation records
        """
        all_records = []
        skip = 0
        top = 1000  # Batch size
        
        try:
            while True:
                # Build OData parameters with expansion to get Person and VacationType details
                params = {
                    '$top': top,
                    '$skip': skip,
                    '$expand': 'Person($select=FirstName,LastName),VacationType($select=Name)',
                    '$filter': f"StartDate ge {start_date} and StartDate le {end_date}"
                }
                
                print(f"📥 Fetching vacation records (batch {skip//top + 1})...")
                
                # Use the existing API client to make the request
                response = self.client.make_api_request('/public/v1/VacationRecords', method='GET', params=params)
                
                if response is None:
                    print("❌ API request failed")
                    break
                
                records = response.get('value', [])
                
                if not records:
                    break
                
                # On first batch, show structure for debugging
                if skip == 0 and records:
                    print("🔍 Sample vacation record structure:")
                    sample_record = records[0]
                    print(f"   Start Date: {sample_record.get('StartDate')}")
                    print(f"   Business Days: {sample_record.get('BusinessDays')}")
                    print(f"   Hours Per Day: {sample_record.get('HoursPerDay')}")
                    if 'Person' in sample_record:
                        person = sample_record['Person']
                        print(f"   Person: {person.get('FirstName')} {person.get('LastName')}")
                    if 'VacationType' in sample_record:
                        vtype = sample_record['VacationType']
                        print(f"   Vacation Type: {vtype.get('Name')}")
                    print()
                    
                all_records.extend(records)
                print(f"✅ Fetched {len(records)} vacation records (Total: {len(all_records)})")
                
                # Check if there are more records
                if len(records) < top:
                    break
                    
                skip += top
                
            print(f"🎯 Total vacation records fetched: {len(all_records)}")
            return all_records
            
        except Exception as e:
            print(f"❌ Error fetching vacation records: {str(e)}")
            return []

    def fetch_timesheet_records(self, start_date: str, end_date: str) -> List[Dict]:
        """
        Fetch timesheet records from ElapseIT API for the specified date range.
        
        Args:
            start_date: Start date in YYYY-MM-DD format
            end_date: End date in YYYY-MM-DD format
            
        Returns:
            List[Dict]: List of timesheet records
        """
        all_records = []
        skip = 0
        top = 1000  # Batch size
        
        try:
            while True:
                # Build OData parameters with expansion to get Person and Project+Client details
                params = {
                    '$top': top,
                    '$skip': skip,
                    '$expand': 'Person($select=FirstName,LastName),Project($expand=Client($select=Name);$select=Name,Code)',
                    '$filter': f"{self.get_date_range_filter(start_date, end_date)} and Hours gt 0"  # Only get records with actual hours
                }
                
                print(f"📥 Fetching timesheet records (batch {skip//top + 1})...")
                
                # Use the existing API client to make the request
                response = self.client.make_api_request('/public/v1/TimesheetRecords', method='GET', params=params)
                
                if response is None:
                    print("❌ API request failed")
                    break
                
                records = response.get('value', [])
                
                if not records:
                    break
                
                # On first batch, show structure for debugging
                if skip == 0 and records:
                    print("🔍 Sample timesheet record structure:")
                    sample_record = records[0]
                    print(f"   Hours: {sample_record.get('Hours')}")
                    print(f"   Day: {sample_record.get('Day')}")
                    if 'Person' in sample_record:
                        person = sample_record['Person']
                        print(f"   Person: {person.get('FirstName')} {person.get('LastName')}")
                    if 'Project' in sample_record:
                        project = sample_record['Project']
                        print(f"   Project: {project.get('Name')}")
                        if 'Client' in project:
                            client = project['Client']
                            print(f"   Client: {client.get('Name')}")
                    print()
                    
                all_records.extend(records)
                print(f"✅ Fetched {len(records)} records (Total: {len(all_records)})")
                
                # Check if there are more records
                if len(records) < top:
                    break
                    
                skip += top
                
            print(f"🎯 Total timesheet records fetched: {len(all_records)}")
            return all_records
            
        except Exception as e:
            print(f"❌ Error fetching timesheet records: {str(e)}")
            return []
    
    def fetch_allocations(self, start_date: str, end_date: str) -> List[Dict]:
        """
        Fetch allocation records from ElapseIT API for the specified date range.
        
        Args:
            start_date: Start date in YYYY-MM-DD format
            end_date: End date in YYYY-MM-DD format
            
        Returns:
            List[Dict]: List of allocation records
        """
        all_records = []
        skip = 0
        top = 1000  # Batch size
        
        try:
            while True:
                # Build OData parameters with expansion to get Person and Project+Client details
                # Use proper datetime format for OData filtering
                start_datetime = f"{start_date}T00:00:00Z"
                end_datetime = f"{end_date}T23:59:59Z"
                
                params = {
                    '$top': top,
                    '$skip': skip,
                    '$expand': 'Person($select=FirstName,LastName,IsArchived,HasLicense,EndDate),Project($expand=Client($select=Name);$select=Name,Code,IsArchived)',
                    '$filter': f"(StartDate le {end_datetime} and (EndDate ge {start_datetime} or EndDate eq null))"
                }
                
                print(f"📥 Fetching allocation records (batch {skip//top + 1})...")
                
                # Use the existing API client to make the request
                response = self.client.make_api_request('/public/v1/ProjectPersonAllocations', method='GET', params=params)
                
                if response is None:
                    print("❌ API request failed")
                    break
                
                records = response.get('value', [])
                
                if not records:
                    break
                
                # On first batch, show structure for debugging
                if skip == 0 and records:
                    print("🔍 Sample allocation record structure:")
                    sample_record = records[0]
                    print(f"   Start Date: {sample_record.get('StartDate')}")
                    print(f"   End Date: {sample_record.get('EndDate')}")
                    print(f"   Allocation %: {sample_record.get('AllocationPercentage', 0)}")
                    print(f"   Available fields: {list(sample_record.keys())}")
                    if 'Person' in sample_record:
                        person = sample_record['Person']
                        print(f"   Person: {person.get('FirstName')} {person.get('LastName')}")
                    if 'Project' in sample_record:
                        project = sample_record['Project']
                        print(f"   Project: {project.get('Name')}")
                        if 'Client' in project:
                            client = project['Client']
                            print(f"   Client: {client.get('Name')}")
                    
                    # Check for allocation name fields
                    allocation_name_fields = ['AllocationName', 'Name', 'TaskName', 'AllocationType']
                    for field in allocation_name_fields:
                        if field in sample_record:
                            print(f"   {field}: {sample_record.get(field)}")
                    print()
                    
                all_records.extend(records)
                print(f"✅ Fetched {len(records)} allocation records (Total: {len(all_records)})")
                
                # Check if there are more records
                if len(records) < top:
                    break
                    
                skip += top
                
            print(f"🎯 Total allocation records fetched: {len(all_records)}")
            return all_records
            
        except Exception as e:
            print(f"❌ Error fetching allocation records: {str(e)}")
            return []
    
    def convert_vacation_to_timesheet_format(self, vacation_records: List[Dict], filter_start_date: str, filter_end_date: str) -> List[Dict]:
        """
        Convert vacation records to timesheet-like format for unified processing.
        
        Args:
            vacation_records: List of vacation records from API
            filter_start_date: Start date filter in YYYY-MM-DD format
            filter_end_date: End date filter in YYYY-MM-DD format
            
        Returns:
            List[Dict]: Vacation records in timesheet format (only dates within filter range)
        """
        timesheet_format_records = []
        
        # Parse filter dates for comparison
        from datetime import datetime, timedelta
        filter_start = datetime.strptime(filter_start_date, '%Y-%m-%d')
        filter_end = datetime.strptime(filter_end_date, '%Y-%m-%d')
        
        for vacation_record in vacation_records:
            try:
                # Extract Person data
                person_data = vacation_record.get('Person', {})
                resource_name = f"{person_data.get('FirstName', '')} {person_data.get('LastName', '')}".strip()
                if not resource_name:
                    resource_name = 'Unknown Resource'
                
                # Extract vacation type
                vacation_type_data = vacation_record.get('VacationType', {})
                vacation_type_name = vacation_type_data.get('Name', 'Unknown Leave Type')
                
                # Get vacation details
                start_date_str = vacation_record.get('StartDate', '')
                end_date_str = vacation_record.get('EndDate', '')
                business_days = vacation_record.get('BusinessDays', 0)
                hours_per_day = vacation_record.get('HoursPerDay', 8.0)
                
                # Debug: Print vacation record details
                print(f"🔍 Vacation record: {resource_name}")
                print(f"   Start: {start_date_str}, End: {end_date_str}")
                print(f"   Business Days: {business_days}, Hours Per Day: {hours_per_day}")
                
                # Parse start and end dates
                if start_date_str and end_date_str:
                    start_date = datetime.strptime(start_date_str.split('T')[0], '%Y-%m-%d')
                    end_date = datetime.strptime(end_date_str.split('T')[0], '%Y-%m-%d')
                    
                    # Check if the vacation period overlaps with our filter range
                    if start_date <= filter_end and end_date >= filter_start:
                        print(f"   📅 Processing vacation: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')} = {business_days} days")
                        
                        # Create individual day records for each business day in the vacation period
                        # This ensures proper month-by-month allocation
                        current_date = start_date
                        days_created = 0.0  # Use float to handle partial days
                        
                        while current_date <= end_date and days_created < business_days:
                            # Skip weekends (assuming business days only)
                            if current_date.weekday() < 5:  # Monday = 0, Friday = 4
                                # Only create entries for dates within our filter range
                                if filter_start <= current_date <= filter_end:
                                    month_year = current_date.strftime('%Y-%m')
                                    month_name = current_date.strftime('%Y-%m. %B %Y')
                                    
                                    # Calculate days for this day based on HoursPerDay
                                    # Convert hours to days: HoursPerDay / 8.0
                                    hours_per_day = vacation_record.get('HoursPerDay', 8.0)
                                    day_hours_to_days = hours_per_day / 8.0
                                    
                                    # For single-day vacations, convert based on HoursPerDay
                                    # For multi-day vacations, use the hours-based calculation
                                    remaining_days = business_days - days_created
                                    if start_date == end_date:
                                        # Single day vacation - convert based on HoursPerDay
                                        # business_days = 1, but HoursPerDay determines the actual days
                                        day_days = day_hours_to_days
                                    elif remaining_days <= day_hours_to_days:
                                        # Last day of multi-day vacation - use remaining days
                                        day_days = remaining_days
                                    else:
                                        # Regular day of multi-day vacation - use hours-based calculation
                                        day_days = day_hours_to_days
                                    
                                    print(f"   📅 Creating leave day: {current_date.strftime('%Y-%m-%d')} in {month_name} ({day_days} days)")
                                    
                                    timesheet_format_records.append({
                                        'Date': current_date.strftime('%Y-%m-%d'),
                                        'Month_Year': month_year,
                                        'Month_Name': month_name,
                                        'Client_Name': 'LEAVE',  # Special client for leave
                                        'Project_ID': f'LEAVE_{vacation_record.get("VacationTypeID", "UNKNOWN")}',
                                        'Project_Name': vacation_type_name,
                                        'Allocation_ID': f'LEAVE_{vacation_record.get("VacationTypeID", "UNKNOWN")}',
                                        'Allocation_Name': vacation_type_name,
                                        'Resource_ID': vacation_record.get('PersonID', ''),
                                        'Resource_Name': resource_name,
                                        'Hours': day_days,  # Use calculated days (can be partial)
                                        'Status': vacation_record.get('Status', 'Approved'),
                                        'VacationID': vacation_record.get('ID', ''),
                                        'IsVacation': True  # Flag to identify vacation records
                                    })
                                days_created += day_days  # Increment by actual days allocated to this day
                            
                            current_date += timedelta(days=1)
                
            except Exception as e:
                print(f"⚠️ Error converting vacation record: {str(e)}")
                continue
        
        return timesheet_format_records
    
    def convert_allocations_to_timesheet_format(self, allocation_records: List[Dict], start_date: str, end_date: str) -> List[Dict]:
        """
        Convert allocation records to timesheet-like format for resources with zero hours.
        
        Args:
            allocation_records: List of allocation records from API
            start_date: Start date in YYYY-MM-DD format
            end_date: End date in YYYY-MM-DD format
            
        Returns:
            List[Dict]: Allocation records in timesheet format with zero hours
        """
        timesheet_format_records = []
        
        from datetime import datetime, timedelta
        filter_start = datetime.strptime(start_date, '%Y-%m-%d')
        filter_end = datetime.strptime(end_date, '%Y-%m-%d')
        
        # Group allocations by person-project to avoid duplicates
        unique_allocations = {}
        
        for allocation_record in allocation_records:
            try:
                # Extract Person data
                person_data = allocation_record.get('Person', {})
                resource_name = f"{person_data.get('FirstName', '')} {person_data.get('LastName', '')}".strip()
                if not resource_name:
                    resource_name = 'Unknown Resource'
                resource_id = allocation_record.get('PersonID', '')
                
                # Apply employee exclusion rules (same as Nexa)
                # Skip BACKLOG ALLOCATIONS (leave adjustments)
                if resource_name == 'BACKLOG ALLOCATIONS':
                    continue
                
                # Skip if person is archived
                if person_data.get('IsArchived', False):
                    continue
                
                # Skip if person doesn't have license (resigned)
                if not person_data.get('HasLicense', True):
                    continue
                
                # Skip if person has end date before our period
                end_date_str = person_data.get('EndDate', '')
                if end_date_str:
                    from datetime import datetime
                    try:
                        person_end_date = datetime.strptime(end_date_str.split('T')[0], '%Y-%m-%d')
                        if person_end_date < filter_start:
                            continue
                    except:
                        pass  # If can't parse date, include the person
                
                # Extract Project data
                project_data = allocation_record.get('Project', {})
                project_name = project_data.get('Name', 'Unknown Project')
                project_id = allocation_record.get('ProjectID', '')
                
                # Skip archived projects
                if project_data.get('IsArchived', False):
                    continue
                
                # Extract Client data from Project
                client_data = project_data.get('Client', {}) if project_data else {}
                client_name = client_data.get('Name', 'Unknown Client') if client_data else 'Unknown Client'
                
                # Try to get allocation name from various possible fields
                # ElapseIT might store allocation names in different fields
                allocation_name = (
                    allocation_record.get('AllocationName') or
                    allocation_record.get('Name') or
                    allocation_record.get('TaskName') or
                    project_name  # Fallback to project name
                )
                
                # Create unique key for person-project combination
                key = f"{resource_id}_{project_id}"
                
                # Only add if not already processed and passes all exclusion rules
                if key not in unique_allocations:
                    unique_allocations[key] = {
                        'Resource_ID': resource_id,
                        'Resource_Name': resource_name,
                        'Client_Name': client_name,
                        'Project_ID': project_id,
                        'Project_Name': project_name,
                        'Allocation_ID': project_id,
                        'Allocation_Name': allocation_name,  # Use proper allocation name
                        'StartDate': allocation_record.get('StartDate', ''),
                        'EndDate': allocation_record.get('EndDate', '')
                    }
                
            except Exception as e:
                print(f"⚠️ Error processing allocation record: {str(e)}")
                continue
        
        # Create monthly entries for each unique allocation
        for key, allocation in unique_allocations.items():
            try:
                # Parse allocation dates
                alloc_start_str = allocation['StartDate']
                alloc_end_str = allocation['EndDate']
                
                if alloc_start_str:
                    alloc_start = datetime.strptime(alloc_start_str.split('T')[0], '%Y-%m-%d')
                else:
                    alloc_start = filter_start
                
                if alloc_end_str:
                    alloc_end = datetime.strptime(alloc_end_str.split('T')[0], '%Y-%m-%d')
                else:
                    alloc_end = filter_end
                
                # Find the overlap between allocation period and our filter period
                effective_start = max(alloc_start, filter_start)
                effective_end = min(alloc_end, filter_end)
                
                if effective_start <= effective_end:
                    # Calculate total days in the allocation period (including weekends)
                    total_days = (effective_end - effective_start).days + 1
                    
                    # Create entries for each month in the effective period
                    current_month = effective_start.replace(day=1)
                    
                    while current_month <= effective_end:
                        month_year = current_month.strftime('%Y-%m')
                        month_name = current_month.strftime('%Y-%m. %B %Y')
                        
                        # Calculate days for this specific month
                        month_start = max(effective_start, current_month)
                        if current_month.month == 12:
                            next_month = current_month.replace(year=current_month.year + 1, month=1)
                        else:
                            next_month = current_month.replace(month=current_month.month + 1)
                        month_end = min(effective_end, next_month - timedelta(days=1))
                        
                        # Calculate days in this month (including weekends)
                        month_days = (month_end - month_start).days + 1
                        
                        timesheet_format_records.append({
                            'Date': current_month.strftime('%Y-%m-%d'),
                            'Month_Year': month_year,
                            'Month_Name': month_name,
                            'Client_Name': allocation['Client_Name'],
                            'Project_ID': allocation['Project_ID'],
                            'Project_Name': allocation['Project_Name'],
                            'Allocation_ID': allocation['Allocation_ID'],
                            'Allocation_Name': allocation['Allocation_Name'],
                            'Resource_ID': allocation['Resource_ID'],
                            'Resource_Name': allocation['Resource_Name'],
                            'Hours': 0.0,  # Zero hours for allocated resources with no timesheet entries
                            'Status': 'Allocated',
                            'AllocationID': key,
                            'IsAllocation': True  # Flag to identify allocation records
                        })
                        
                        # Move to next month
                        if current_month.month == 12:
                            current_month = current_month.replace(year=current_month.year + 1, month=1)
                        else:
                            current_month = current_month.replace(month=current_month.month + 1)
                
            except Exception as e:
                print(f"⚠️ Error converting allocation record: {str(e)}")
                continue
        
        return timesheet_format_records
    
    def process_timesheet_data(self, timesheet_records: List[Dict], vacation_records: List[Dict], allocation_records: List[Dict], start_date: str, end_date: str) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        """
        Process raw timesheet and vacation data into three DataFrames for Excel sheets.
        
        Args:
            timesheet_records: Raw timesheet records from API (with expanded Person and Project data)
            vacation_records: Raw vacation records from API (with expanded Person and VacationType data)
            allocation_records: Raw allocation records from API
            start_date: Start date in YYYY-MM-DD format
            end_date: End date in YYYY-MM-DD format
            
        Returns:
            Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]: (client_grouped_df, resource_grouped_df, daily_detailed_df)
        """
        print(f"📊 Processing {len(timesheet_records)} timesheet records, {len(vacation_records)} vacation records, and {len(allocation_records)} allocation records...")
        
        # Convert vacation records to timesheet format (only dates within our range)
        vacation_timesheet_records = self.convert_vacation_to_timesheet_format(vacation_records, start_date, end_date)
        print(f"📊 Converted {len(vacation_records)} vacation records to {len(vacation_timesheet_records)} daily vacation entries")
        
        # Convert allocation records to timesheet format (for zero-hour allocated resources)
        allocation_timesheet_records = self.convert_allocations_to_timesheet_format(allocation_records, start_date, end_date)
        print(f"📊 Converted {len(allocation_records)} allocation records to {len(allocation_timesheet_records)} monthly allocation entries")
        
        # Combine timesheet, vacation, and allocation records
        all_records = timesheet_records + vacation_timesheet_records + allocation_timesheet_records
        
        if not all_records:
            print("⚠️ No records to process")
            return pd.DataFrame(), pd.DataFrame()
        
        # Convert to DataFrame
        df_list = []
        
        for record in all_records:
            try:
                # Check if this is a vacation record (already converted) or a regular timesheet record
                if record.get('IsVacation', False):
                    # Vacation record - already converted to days, no need to divide by 8 again
                    df_list.append({
                        'Date': record.get('Date', ''),
                        'Month_Year': record.get('Month_Year', ''),
                        'Month_Name': record.get('Month_Name', ''),
                        'Client_Name': record.get('Client_Name', ''),
                        'Project_ID': record.get('Project_ID', ''),
                        'Project_Name': record.get('Project_Name', ''),
                        'Allocation_ID': record.get('Allocation_ID', ''),
                        'Allocation_Name': record.get('Allocation_Name', ''),
                        'Resource_ID': record.get('Resource_ID', ''),
                        'Resource_Name': record.get('Resource_Name', ''),
                        'Hours': float(record.get('Hours', 0)),  # Already in days, no conversion needed
                        'Status': record.get('Status', ''),
                        'VacationID': record.get('VacationID', '')
                    })
                elif record.get('IsAllocation', False):
                    # Allocation record - already in the right format
                    df_list.append({
                        'Date': record.get('Date', ''),
                        'Month_Year': record.get('Month_Year', ''),
                        'Month_Name': record.get('Month_Name', ''),
                        'Client_Name': record.get('Client_Name', ''),
                        'Project_ID': record.get('Project_ID', ''),
                        'Project_Name': record.get('Project_Name', ''),
                        'Allocation_ID': record.get('Allocation_ID', ''),
                        'Allocation_Name': record.get('Allocation_Name', ''),
                        'Resource_ID': record.get('Resource_ID', ''),
                        'Resource_Name': record.get('Resource_Name', ''),
                        'Hours': float(record.get('Hours', 0)) / 8.0,  # Convert hours to days
                        'Status': record.get('Status', ''),
                        'AllocationID': record.get('AllocationID', '')
                    })
                else:
                    # Regular timesheet record - process as before
                    # Extract Person data
                    person_data = record.get('Person', {})
                    resource_name = f"{person_data.get('FirstName', '')} {person_data.get('LastName', '')}".strip()
                    if not resource_name:
                        resource_name = 'Unknown Resource'
                    resource_id = record.get('PersonID', '')
                    
                    # Extract Project data
                    project_data = record.get('Project', {})
                    project_name = project_data.get('Name', 'Unknown Project')
                    project_id = record.get('ProjectID', '')
                    
                    # Extract Client data from Project
                    client_data = project_data.get('Client', {}) if project_data else {}
                    client_name = client_data.get('Name', 'Unknown Client') if client_data else 'Unknown Client'
                    
                    # For now, use Project name as Allocation name since timesheet records don't have direct allocation references
                    # This matches the logic from the mapper where projects can have multiple allocations
                    allocation_name = project_name
                    allocation_id = project_id
                    
                    # Parse date from Day field
                    day_str = record.get('Day', '')
                    if day_str:
                        date_obj = datetime.strptime(day_str.split('T')[0], '%Y-%m-%d')
                        month_year = date_obj.strftime('%Y-%m')
                        month_name = date_obj.strftime('%Y-%m. %B %Y')
                        date_formatted = day_str.split('T')[0]
                    else:
                        month_year = 'Unknown'
                        month_name = 'Unknown'
                        date_formatted = ''
                    
                    df_list.append({
                        'Date': date_formatted,
                        'Month_Year': month_year,
                        'Month_Name': month_name,
                        'Client_Name': client_name,
                        'Project_ID': project_id,
                        'Project_Name': project_name,
                        'Allocation_ID': allocation_id,
                        'Allocation_Name': allocation_name,
                        'Resource_ID': resource_id,
                        'Resource_Name': resource_name,
                        'Hours': float(record.get('Hours', 0)) / 8.0,  # Convert hours to days
                        'Status': record.get('Status', ''),
                        'TimesheetID': record.get('ID', '')
                    })
                
            except Exception as e:
                print(f"⚠️ Error processing record: {str(e)}")
                print(f"   Record keys: {list(record.keys()) if record else 'None'}")
                continue
        
        if not df_list:
            print("⚠️ No valid records processed")
            return pd.DataFrame(), pd.DataFrame()
        
        df = pd.DataFrame(df_list)
        
        # Simple approach: Only add allocation records for people who don't have any timesheet or vacation records
        print("🔄 Adding missing allocated resources...")
        
        # Get existing resource-project combinations from timesheet and vacation data
        existing_combinations = set()
        for record in df_list:
            if not record.get('IsAllocation', False):  # Only timesheet and vacation records
                key = f"{record.get('Resource_ID', '')}_{record.get('Project_ID', '')}"
                existing_combinations.add(key)
        
        # Filter allocation records to only include those not already covered
        allocation_only_records = []
        for record in df_list:
            if record.get('IsAllocation', False):  # Allocation records
                key = f"{record.get('Resource_ID', '')}_{record.get('Project_ID', '')}"
                if key not in existing_combinations:
                    allocation_only_records.append(record)
        
        print(f"✅ Found {len(allocation_only_records)} allocation-only entries to add (e.g., Tinu Elenjical)")
        
        # Create final dataframe with all records
        df = pd.DataFrame(df_list)
        
        # Generate month columns for the date range
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
        
        month_columns = []
        current_date = start_dt.replace(day=1)
        
        while current_date <= end_dt:
            month_year = current_date.strftime('%Y-%m')
            month_name = current_date.strftime('%Y-%m. %B %Y')  # Add year-month prefix for proper sorting across years
            month_columns.append((month_year, month_name))
            
            # Move to next month
            if current_date.month == 12:
                current_date = current_date.replace(year=current_date.year + 1, month=1)
            else:
                current_date = current_date.replace(month=current_date.month + 1)
        
        # Sheet 1: Client -> Allocation -> Resource breakdown by month
        client_grouped = self._create_client_grouped_sheet(df, month_columns)
        
        # Sheet 2: Resource -> Client -> Allocation breakdown by month
        resource_grouped = self._create_resource_grouped_sheet(df, month_columns)
        
        # Sheet 3: Daily detailed data ordered by Client, Allocation, Employee
        daily_detailed = self._create_daily_detailed_sheet(df)
        
        return client_grouped, resource_grouped, daily_detailed
    
    def _create_client_grouped_sheet(self, df: pd.DataFrame, month_columns: List[Tuple[str, str]]) -> pd.DataFrame:
        """Create Client -> Allocation -> Resource grouped sheet."""
        
        # Group by Client, Allocation, Resource
        grouped = df.groupby(['Client_Name', 'Allocation_Name', 'Resource_Name', 'Month_Year'])['Hours'].sum().reset_index()
        
        # Pivot to get months as columns
        pivot = grouped.pivot_table(
            index=['Client_Name', 'Allocation_Name', 'Resource_Name'],
            columns='Month_Year',
            values='Hours',
            fill_value=0
        )
        
        # Ensure all month columns are present
        for month_year, month_name in month_columns:
            if month_year not in pivot.columns:
                pivot[month_year] = 0
        
        # Sort columns chronologically
        month_order = [month_year for month_year, _ in month_columns]
        pivot = pivot.reindex(columns=month_order)
        
        # Reset index to make grouping columns regular columns (don't add total yet - will use Excel formula)
        result = pivot.reset_index()
        
        # Rename columns to month names for better readability
        column_mapping = {month_year: month_name for month_year, month_name in month_columns}
        result = result.rename(columns=column_mapping)
        
        return result
    
    def _create_daily_detailed_sheet(self, df: pd.DataFrame) -> pd.DataFrame:
        """Create daily detailed data sheet ordered by Client, Allocation, Employee."""
        
        # Create a copy of the dataframe for daily detailed view
        daily_df = df.copy()
        
        # Ensure Date column is properly formatted as datetime
        daily_df['Date'] = pd.to_datetime(daily_df['Date'])
        
        # Add day of week column
        daily_df['Day_Of_Week'] = daily_df['Date'].dt.day_name()
        
        # Add week number column
        daily_df['Week_Number'] = daily_df['Date'].dt.isocalendar().week
        
        # Sort by Date first, then Client_Name, Allocation_Name, Resource_Name for chronological order
        daily_df = daily_df.sort_values(['Date', 'Client_Name', 'Allocation_Name', 'Resource_Name'])
        
        # Include all available columns for comprehensive output
        # Get all columns and reorder them for better readability
        all_columns = list(daily_df.columns)
        
        # Define preferred column order (put most important first)
        preferred_order = [
            'Date', 'Day_Of_Week', 'Week_Number', 'Month_Name', 'Month_Year',
            'Client_Name', 'Allocation_Name', 'Resource_Name',
            'Hours', 'Days', 'Status', 'Project_ID', 'Project_Name',
            'Resource_ID', 'Allocation_ID', 'TimesheetID', 'VacationID', 'AllocationID'
        ]
        
        # Reorder columns: preferred order first, then any remaining columns
        ordered_columns = []
        for col in preferred_order:
            if col in all_columns:
                ordered_columns.append(col)
                all_columns.remove(col)
        
        # Add any remaining columns that weren't in the preferred order
        ordered_columns.extend(all_columns)
        
        # Reorder the dataframe columns
        daily_df = daily_df[ordered_columns]
        
        # Rename Hours to Days for clarity
        if 'Hours' in daily_df.columns:
            daily_df = daily_df.rename(columns={'Hours': 'Days'})
        
        return daily_df
    
    def _create_resource_grouped_sheet(self, df: pd.DataFrame, month_columns: List[Tuple[str, str]]) -> pd.DataFrame:
        """Create Resource -> Client -> Allocation grouped sheet with Month column."""
        
        # Group by Resource, Client, Allocation, Month and sum hours
        grouped = df.groupby(['Resource_Name', 'Client_Name', 'Allocation_Name', 'Month_Year'])['Hours'].sum().reset_index()
        
        # Create month name mapping
        month_mapping = {month_year: month_name for month_year, month_name in month_columns}
        
        # Add Month column with readable month names
        grouped['Month'] = grouped['Month_Year'].map(month_mapping)
        
        # Handle any unmapped months (shouldn't happen but just in case)
        unmapped_mask = grouped['Month'].isna()
        if unmapped_mask.any():
            print(f"⚠️ Warning: {unmapped_mask.sum()} records have unmapped months")
            # For unmapped months, create a readable name from Month_Year
            from datetime import datetime
            for idx in grouped[unmapped_mask].index:
                month_year = grouped.loc[idx, 'Month_Year']
                try:
                    # Parse YYYY-MM format and create readable name
                    date_obj = datetime.strptime(month_year, '%Y-%m')
                    readable_name = date_obj.strftime('%Y-%m. %B %Y')
                    grouped.loc[idx, 'Month'] = readable_name
                except:
                    # Fallback to original value
                    grouped.loc[idx, 'Month'] = str(month_year)
        
        # Select final columns: Resource_Name, Client_Name, Allocation_Name, Month, Days
        result = grouped[['Resource_Name', 'Client_Name', 'Allocation_Name', 'Month', 'Hours']].copy()
        result.rename(columns={'Hours': 'Days'}, inplace=True)
        
        # Sort by Resource, Client, Allocation, and chronologically by month
        month_order = [month_name for _, month_name in month_columns]
        result['Month_Order'] = result['Month'].map({month_name: i for i, month_name in enumerate(month_order)})
        result = result.sort_values(['Resource_Name', 'Client_Name', 'Allocation_Name', 'Month_Order'])
        result = result.drop('Month_Order', axis=1)
        
        return result
    
    def create_histogram_data(self, df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, dict]:
        """
        Create comprehensive histogram data including employee breakdowns, statistical distribution, 
        top/bottom performers relative to mean.
        
        Args:
            df: Resource DataFrame with all timesheet data
            
        Returns:
            Tuple containing:
            - employee_stacked_df: Employee data for stacked bar chart
            - distribution_df: Statistical distribution data
            - top_10_df: Top 10 employees by total hours
            - bottom_10_df: Bottom 10 employees by total hours  
            - stats_dict: Overall statistics
        """
        print("📊 Creating comprehensive histogram data...")
        
        # Group by Resource and Month to get monthly totals
        monthly_data = df.groupby(['Resource_Name', 'Month'])['Days'].sum().reset_index()
        
        # Calculate overall statistics (chronologically ordered)
        print("\n📈 Monthly Statistics:")
        
        # Get chronologically sorted months for statistics
        stats_month_dates = []
        for month_name in monthly_data['Month'].unique():
            try:
                # Parse the new format "YYYY-MM. Month YYYY"
                month_date = datetime.strptime(month_name.split('.')[0], '%Y-%m')
                stats_month_dates.append((month_date, month_name))
            except ValueError:
                try:
                    # Fallback to old format
                    month_date = datetime.strptime(month_name, '%B %Y')
                    stats_month_dates.append((month_date, month_name))
                except ValueError:
                    stats_month_dates.append((datetime.now(), month_name))
        
        stats_months = [month_name for _, month_name in sorted(stats_month_dates)]
        
        for month in stats_months:
            month_data = monthly_data[monthly_data['Month'] == month]['Days']
            mean_days = month_data.mean()
            median_days = month_data.median()
            print(f"   {month}: Mean = {mean_days:.1f} days, Median = {median_days:.1f} days")
        
        # Create separate employee breakdown for each month (chronologically ordered)
        # Convert month names to datetime objects for proper sorting
        month_dates = []
        for month_name in df['Month'].unique():
            try:
                # Parse the new format "YYYY-MM. Month YYYY"
                month_date = datetime.strptime(month_name.split('.')[0], '%Y-%m')
                month_dates.append((month_date, month_name))
            except ValueError:
                try:
                    # Fallback to old format
                    month_date = datetime.strptime(month_name, '%B %Y')
                    month_dates.append((month_date, month_name))
                except ValueError:
                    # Fallback for any parsing issues
                    month_dates.append((datetime.now(), month_name))
        
        # Sort by date and extract month names in chronological order
        months = [month_name for _, month_name in sorted(month_dates)]
        monthly_employee_data = {}
        
        for month in months:
            month_data = df[df['Month'] == month]
            employee_breakdown = []
            
            for resource_name in month_data['Resource_Name'].unique():
                resource_data = month_data[month_data['Resource_Name'] == resource_name]
                
                # Calculate breakdown by category for this month
                leave_hours = resource_data[resource_data['Client_Name'] == 'LEAVE']['Days'].sum()
                internal_hours = resource_data[resource_data['Client_Name'] == 'Elenjical Solutions']['Days'].sum()
                other_hours = resource_data[
                    (resource_data['Client_Name'] != 'LEAVE') & 
                    (resource_data['Client_Name'] != 'Elenjical Solutions')
                ]['Days'].sum()
                
                month_total = leave_hours + internal_hours + other_hours
                
                # Only include employees who worked in this month
                if month_total > 0:
                    employee_breakdown.append({
                        'Resource_Name': resource_name,
                        'LEAVE_Days': leave_hours,
                        'Internal_Days': internal_hours,
                        'Other_Days': other_hours,
                        'Total_Days': month_total
                    })
            
            # Convert to DataFrame and sort by total days (least to most)
            month_df = pd.DataFrame(employee_breakdown)
            if not month_df.empty:
                month_df = month_df.sort_values('Total_Days', ascending=True).reset_index(drop=True)
            monthly_employee_data[month] = month_df
        
        # For compatibility with existing code, create a combined dataframe
        all_monthly_data = []
        for month, month_df in monthly_employee_data.items():
            month_df_copy = month_df.copy()
            month_df_copy['Month'] = month
            all_monthly_data.append(month_df_copy)
        
        employee_stacked_df = pd.concat(all_monthly_data, ignore_index=True) if all_monthly_data else pd.DataFrame()
        
        # Calculate overall statistics based on employee totals across all months
        employee_totals = employee_stacked_df.groupby('Resource_Name')['Total_Days'].sum().reset_index()
        employee_totals.columns = ['Resource_Name', 'Total_Days']
        
        total_days_series = employee_totals['Total_Days']
        mean_total = total_days_series.mean()
        median_total = total_days_series.median()
        std_total = total_days_series.std()
        min_total = total_days_series.min()
        max_total = total_days_series.max()
        
        stats_dict = {
            'mean': mean_total,
            'median': median_total,
            'std': std_total,
            'min': min_total,
            'max': max_total,
            'total_employees': len(employee_totals),
            'total_months': len(months),
            'total_data_points': len(employee_stacked_df)
        }
        
        print(f"\n📊 Overall Statistics:")
        print(f"   Total Employees: {stats_dict['total_employees']}")
        print(f"   Total Months: {stats_dict['total_months']}")
        print(f"   Total Data Points: {stats_dict['total_data_points']}")
        print(f"   Mean Days per Employee: {mean_total:.1f}")
        print(f"   Median Days per Employee: {median_total:.1f}")
        print(f"   Std Deviation: {std_total:.1f}")
        print(f"   Range: {min_total:.1f} to {max_total:.1f}")
        
        # Create statistical distribution (histogram bins) based on employee totals
        bin_count = min(15, len(employee_totals) // 2) if len(employee_totals) > 4 else 5
        bins = pd.cut(total_days_series, bins=bin_count, include_lowest=True)
        distribution_counts = bins.value_counts().sort_index()
        
        # Create distribution DataFrame for charting
        distribution_data = []
        for interval, count in distribution_counts.items():
            bin_center = (interval.left + interval.right) / 2
            distribution_data.append({
                'Days_Range': f"{interval.left:.0f}-{interval.right:.0f}",
                'Bin_Center': bin_center,
                'Employee_Count': count
            })
        
        distribution_df = pd.DataFrame(distribution_data)
        
        # Identify top 10 and bottom 10 employees by total days
        sorted_employees = employee_totals.sort_values('Total_Days', ascending=False)
        top_10_df = sorted_employees.head(10).copy()
        bottom_10_df = sorted_employees.tail(10).copy()
        
        # Add deviation from mean
        top_10_df['Days_Above_Mean'] = top_10_df['Total_Days'] - mean_total
        bottom_10_df['Days_Below_Mean'] = mean_total - bottom_10_df['Total_Days']
        
        print(f"✅ Comprehensive histogram data created:")
        print(f"   - {len(employee_stacked_df)} employee-month data points for stacked bar chart")
        print(f"   - {len(distribution_df)} bins for statistical distribution")
        print(f"   - Top 10 performers: {top_10_df['Total_Days'].min():.1f} to {top_10_df['Total_Days'].max():.1f} days")
        print(f"   - Bottom 10 performers: {bottom_10_df['Total_Days'].min():.1f} to {bottom_10_df['Total_Days'].max():.1f} days")
        
        return employee_stacked_df, distribution_df, top_10_df, bottom_10_df, stats_dict, monthly_employee_data
    
    def create_interactive_dashboard(self, df: pd.DataFrame, monthly_employee_data: dict, stats_dict: dict, filename: str):
        """
        Create interactive Plotly pop-up charts with monthly statistics and drill-down capabilities.
        
        Args:
            df: Main timesheet DataFrame
            monthly_employee_data: Dictionary of monthly employee DataFrames
            stats_dict: Overall statistics dictionary
            filename: Base filename for outputs
        """
        try:
            import plotly.graph_objects as go
            import plotly.express as px
            from plotly.subplots import make_subplots
            import plotly.offline as pyo
            import os
            from datetime import datetime
            
            print("📊 Creating interactive Plotly charts...")
            print("💡 Charts will open in your default browser as interactive pop-ups")
            
            # 1. Monthly Overview Dashboard (chronologically ordered)
            # Sort monthly_employee_data by chronological order
            monthly_dates = []
            for month_name in monthly_employee_data.keys():
                try:
                    # Parse the new format "YYYY-MM. Month YYYY"
                    month_date = datetime.strptime(month_name.split('.')[0], '%Y-%m')
                    monthly_dates.append((month_date, month_name))
                except ValueError:
                    try:
                        # Fallback to old format
                        month_date = datetime.strptime(month_name, '%B %Y')
                        monthly_dates.append((month_date, month_name))
                    except ValueError:
                        monthly_dates.append((datetime.now(), month_name))
            
            sorted_months = [month_name for _, month_name in sorted(monthly_dates)]
            
            monthly_totals = []
            for month in sorted_months:
                month_df = monthly_employee_data[month]
                if not month_df.empty:
                    total_days = month_df['Total_Days'].sum()
                    avg_days = month_df['Total_Days'].mean()
                    employee_count = len(month_df)
                    
                    monthly_totals.append({
                        'Month': month,
                        'Total_Days': total_days,
                        'Average_Days': avg_days,
                        'Employee_Count': employee_count,
                        'Leave_Days': month_df['LEAVE_Days'].sum(),
                        'Internal_Days': month_df['Internal_Days'].sum(),
                        'Other_Days': month_df['Other_Days'].sum()
                    })
            
            monthly_df = pd.DataFrame(monthly_totals)
            
            # Debug: Print monthly data for verification
            print(f"📊 Creating overview dashboard with {len(monthly_df)} months:")
            for _, row in monthly_df.iterrows():
                print(f"   📅 {row['Month']}: {row['Total_Days']:.1f} total days, {row['Average_Days']:.1f} avg, {row['Employee_Count']} employees")
            
            # Create main dashboard with subplots
            fig = make_subplots(
                rows=2, cols=2,
                subplot_titles=(f'Monthly Total Days ({len(sorted_months)} months)', 
                              f'Average Days per Employee ({len(sorted_months)} months)', 
                              f'Employee Count by Month ({len(sorted_months)} months)', 
                              f'Days Breakdown by Category ({len(sorted_months)} months)'),
                specs=[[{"secondary_y": False}, {"secondary_y": False}],
                       [{"secondary_y": False}, {"type": "bar"}]]
            )
            
            # Monthly total days (clickable) - use appropriate mode based on data points
            chart_mode = 'lines+markers' if len(monthly_df) > 1 else 'markers'
            fig.add_trace(
                go.Scatter(
                    x=monthly_df['Month'],
                    y=monthly_df['Total_Days'],
                    mode=chart_mode,
                    name='Total Days',
                    hovertemplate='<b>%{x}</b><br>Total Days: %{y}<br><extra></extra>',
                    customdata=monthly_df['Month'],
                    marker=get_plotly_marker_config('TOTAL', 8),
                    line=get_plotly_line_config('TOTAL', 3) if len(monthly_df) > 1 else None
                ),
                row=1, col=1
            )
            
            # Average days per employee with bounds
            overall_mean = monthly_df['Average_Days'].mean()
            overall_std = monthly_df['Average_Days'].std()
            # Handle case where std is NaN (single data point)
            if pd.isna(overall_std) or overall_std == 0:
                overall_std = overall_mean * 0.1  # Use 10% of mean as fallback
            overall_upper_bound = overall_mean + 1.5 * overall_std
            overall_lower_bound = overall_mean - 1.5 * overall_std
            
            fig.add_trace(
                go.Scatter(
                    x=monthly_df['Month'],
                    y=monthly_df['Average_Days'],
                    mode=chart_mode,
                    name='Avg Days/Employee',
                    hovertemplate='<b>%{x}</b><br>Avg Days: %{y:.1f}<br><extra></extra>',
                    line=get_plotly_line_config('TREND_LINE', 3) if len(monthly_df) > 1 else None,
                    marker=get_plotly_marker_config('TREND_LINE', 8)
                ),
                row=1, col=2
            )
            
            # Add bounds to the average days chart
            fig.add_hline(
                y=overall_upper_bound,
                line_dash="dash",
                line_color=get_chart_color('UPPER_BOUND'),
                annotation_text=f"Upper: {overall_upper_bound:.1f}",
                annotation_position="top right",
                row=1, col=2
            )
            
            fig.add_hline(
                y=overall_lower_bound,
                line_dash="dash",
                line_color=get_chart_color('LOWER_BOUND'), 
                annotation_text=f"Lower: {overall_lower_bound:.1f}",
                annotation_position="bottom right",
                row=1, col=2
            )
            
            # Employee count
            fig.add_trace(
                go.Bar(
                    x=monthly_df['Month'],
                    y=monthly_df['Employee_Count'],
                    name='Employee Count',
                    hovertemplate='<b>%{x}</b><br>Employees: %{y}<br><extra></extra>',
                    marker_color=get_category_color('EMPLOYEE_COUNT')
                ),
                row=2, col=1
            )
            
            # Stacked bar chart for categories
            fig.add_trace(
                go.Bar(
                    x=monthly_df['Month'],
                    y=monthly_df['Leave_Days'],
                    name='LEAVE Days',
                    hovertemplate='<b>%{x}</b><br>LEAVE: %{y} days<br><extra></extra>',
                    marker_color=get_category_color('LEAVE')
                ),
                row=2, col=2
            )
            
            fig.add_trace(
                go.Bar(
                    x=monthly_df['Month'],
                    y=monthly_df['Internal_Days'],
                    name='Internal Days',
                    hovertemplate='<b>%{x}</b><br>Internal: %{y} days<br><extra></extra>',
                    marker_color=get_category_color('INTERNAL')
                ),
                row=2, col=2
            )
            
            fig.add_trace(
                go.Bar(
                    x=monthly_df['Month'],
                    y=monthly_df['Other_Days'],
                    name='Other Days',
                    hovertemplate='<b>%{x}</b><br>Other: %{y} days<br><extra></extra>',
                    marker_color=get_category_color('OTHER')
                ),
                row=2, col=2
            )
            
            # Update layout
            fig.update_layout(
                height=900,
                title_text=f"📊 Interactive Monthly Timesheet Dashboard ({len(sorted_months)} months) - Click on data points to drill down!",
                title_x=0.5,
                showlegend=True,
                barmode='stack',
                margin=dict(l=80, r=80, t=120, b=80)
            )
            
            # Update axes labels
            fig.update_xaxes(title_text="Month", row=2, col=1)
            fig.update_xaxes(title_text="Month", row=2, col=2)
            fig.update_yaxes(title_text="Total Days", row=1, col=1)
            fig.update_yaxes(title_text="Average Days", row=1, col=2)
            fig.update_yaxes(title_text="Employee Count", row=2, col=1)
            fig.update_yaxes(title_text="Days", row=2, col=2)
            
            # Show main dashboard in browser
            print("📊 Opening Monthly Overview Dashboard...")
            print(f"📈 Dashboard will show trends across {len(monthly_df)} months: {', '.join(monthly_df['Month'].tolist())}")
            print(f"📊 Dashboard data summary:")
            print(f"   - Total Days range: {monthly_df['Total_Days'].min():.1f} to {monthly_df['Total_Days'].max():.1f}")
            print(f"   - Average Days range: {monthly_df['Average_Days'].min():.1f} to {monthly_df['Average_Days'].max():.1f}")
            print(f"   - Employee Count range: {monthly_df['Employee_Count'].min()} to {monthly_df['Employee_Count'].max()}")
            
            try:
                # Configure plotly to use browser
                import plotly.io as pio
                pio.renderers.default = "browser"
                
                # Show the dashboard with explicit browser opening
                fig.show()
                print("✅ Overview dashboard opened successfully!")
                print("🌐 If the overview dashboard didn't open, check your default browser settings")
                
            except Exception as e:
                print(f"❌ Error opening overview dashboard: {str(e)}")
                # Try alternative display method
                try:
                    import plotly.offline as pyo
                    pyo.plot(fig, auto_open=True, filename='temp_overview.html')
                    print("✅ Overview dashboard opened using alternative method!")
                except Exception as e2:
                    print(f"❌ Alternative method also failed: {str(e2)}")
            
            # Small delay to ensure dashboard opens properly
            import time
            time.sleep(2)
            
            # 2. Create individual monthly drill-down charts with adjacent distributions (chronologically ordered)
            print("📊 Opening monthly detail charts with distributions...")
            for month in sorted_months:
                month_df = monthly_employee_data[month]
                if not month_df.empty:
                    # Sort by total days to identify outliers
                    month_df_sorted = month_df.sort_values('Total_Days', ascending=True)
                    
                    # Calculate statistics
                    mean_days = month_df_sorted['Total_Days'].mean()
                    median_days = month_df_sorted['Total_Days'].median()
                    std_days = month_df_sorted['Total_Days'].std()
                    upper_bound = mean_days + 1.5 * std_days
                    lower_bound = mean_days - 1.5 * std_days
                    
                    # Create subplot with 2 rows: bar chart above, distribution below
                    fig_month = make_subplots(
                        rows=2, cols=1,
                        subplot_titles=(f'{month} - Employee Days Breakdown', f'{month} - Distribution & Stats'),
                        specs=[[{"secondary_y": False}], [{"secondary_y": False}]],
                        row_heights=[0.65, 0.35],
                        vertical_spacing=0.18
                    )
                    
                    # TOP: Stacked bar chart
                    fig_month.add_trace(
                        go.Bar(
                            name='LEAVE Days',
                            x=month_df_sorted['Resource_Name'],
                            y=month_df_sorted['LEAVE_Days'],
                            hovertemplate='<b>%{x}</b><br>LEAVE: %{y} days<br><extra></extra>',
                            marker_color=get_category_color('LEAVE'),
                            showlegend=True
                        ),
                        row=1, col=1
                    )
                    
                    fig_month.add_trace(
                        go.Bar(
                            name='Internal Days',
                            x=month_df_sorted['Resource_Name'],
                            y=month_df_sorted['Internal_Days'],
                            hovertemplate='<b>%{x}</b><br>Internal: %{y} days<br><extra></extra>',
                            marker_color=get_category_color('INTERNAL'),
                            showlegend=True
                        ),
                        row=1, col=1
                    )
                    
                    fig_month.add_trace(
                        go.Bar(
                            name='Other Days',
                            x=month_df_sorted['Resource_Name'],
                            y=month_df_sorted['Other_Days'],
                            hovertemplate='<b>%{x}</b><br>Other: %{y} days<br><extra></extra>',
                            marker_color=get_category_color('OTHER'),
                            showlegend=True
                        ),
                        row=1, col=1
                    )
                    
                    # Add horizontal lines for outlier bounds to bar chart
                    fig_month.add_hline(
                        y=upper_bound,
                        line_dash="dash",
                        line_color=get_chart_color('UPPER_BOUND'),
                        annotation_text=f"Upper: {upper_bound:.1f}",
                        annotation_position="top right",
                        row=1, col=1
                    )
                    
                    fig_month.add_hline(
                        y=lower_bound,
                        line_dash="dash", 
                        line_color=get_chart_color('LOWER_BOUND'),
                        annotation_text=f"Lower: {lower_bound:.1f}",
                        annotation_position="bottom right",
                        row=1, col=1
                    )
                    
                    # Identify outliers (more than 1.5 std deviations from mean)
                    outliers = month_df_sorted[
                        (month_df_sorted['Total_Days'] > upper_bound) |
                        (month_df_sorted['Total_Days'] < lower_bound)
                    ]
                    
                    # Add outlier markers to bar chart
                    if not outliers.empty:
                        fig_month.add_trace(
                            go.Scatter(
                                x=outliers['Resource_Name'],
                                y=outliers['Total_Days'],
                                mode='markers',
                                name='Outliers',
                                marker=dict(
                                    size=15,
                                    color=get_chart_color('OUTLIER_MARKER'),
                                    symbol='star',
                                    line=dict(width=2, color=get_chart_color('OUTLIER_BORDER'))
                                ),
                                hovertemplate='<b>OUTLIER: %{x}</b><br>Total: %{y} days<br>' +
                                            f'Mean: {mean_days:.1f} days<br>' +
                                            f'Upper Bound: {upper_bound:.1f} days<br>' +
                                            f'Lower Bound: {lower_bound:.1f} days<br>' +
                                            'Deviation: %{customdata:.1f} days<br><extra></extra>',
                                customdata=outliers['Total_Days'] - mean_days,
                                showlegend=True
                            ),
                            row=1, col=1
                        )
                    
                    # BOTTOM: Distribution histogram
                    # Calculate appropriate number of bins based on data size
                    n_employees = len(month_df_sorted)
                    n_bins = min(max(int(n_employees / 5), 5), 15)  # Between 5-15 bins
                    
                    fig_month.add_trace(
                        go.Histogram(
                            x=month_df_sorted['Total_Days'],
                            nbinsx=n_bins,
                            name='Distribution',
                            marker_color=get_chart_color('DISTRIBUTION'),
                            opacity=0.7,
                            hovertemplate='Days Range: %{x}<br>Count: %{y}<br><extra></extra>',
                            showlegend=False
                        ),
                        row=2, col=1
                    )
                    
                    # Add statistical lines to distribution with better annotation positioning
                    fig_month.add_vline(
                        x=mean_days,
                        line_dash="solid",
                        line_color=get_chart_color('MEAN_LINE'),
                        annotation_text=f"Mean: {mean_days:.1f}",
                        annotation_position="top left",
                        row=2, col=1
                    )
                    
                    fig_month.add_vline(
                        x=median_days,
                        line_dash="dot",
                        line_color=get_chart_color('MEDIAN_LINE'),
                        annotation_text=f"Median: {median_days:.1f}",
                        annotation_position="top right",
                        row=2, col=1
                    )
                    
                    fig_month.add_vline(
                        x=upper_bound,
                        line_dash="dash",
                        line_color=get_chart_color('UPPER_BOUND'),
                        annotation_text=f"Upper: {upper_bound:.1f}",
                        annotation_position="bottom left",
                        row=2, col=1
                    )
                    
                    fig_month.add_vline(
                        x=lower_bound,
                        line_dash="dash",
                        line_color=get_chart_color('LOWER_BOUND'),
                        annotation_text=f"Lower: {lower_bound:.1f}",
                        annotation_position="bottom right",
                        row=2, col=1
                    )
                    
                    # Update layout
                    fig_month.update_layout(
                        barmode='stack',
                        title=f'📊 {month} - Employee Analysis & Distribution<br>' +
                              f'<sub>Employees: {len(month_df_sorted)} | Mean: {mean_days:.1f} days | ' +
                              f'Median: {median_days:.1f} days | Std: {std_days:.1f} | Outliers: {len(outliers)}</sub>',
                        title_x=0.5,
                        height=1300,
                        hovermode='closest',
                        margin=dict(l=80, r=80, t=140, b=120)
                    )
                    
                    # Adjust subplot title positions to prevent overlap
                    for i, annotation in enumerate(fig_month['layout']['annotations']):
                        if i == 1:  # Second subplot title (Distribution & Stats)
                            annotation['y'] = annotation['y'] + 0.02  # Move title up slightly
                    
                    # Update axes
                    fig_month.update_xaxes(
                        title_text="Employees (Ranked Least to Most Days)", 
                        tickangle=-45, 
                        tickfont=dict(size=10),
                        row=1, col=1
                    )
                    fig_month.update_xaxes(
                        title_text="Total Days", 
                        tickfont=dict(size=12),
                        row=2, col=1
                    )
                    fig_month.update_yaxes(
                        title_text="Days", 
                        tickfont=dict(size=12),
                        row=1, col=1
                    )
                    fig_month.update_yaxes(
                        title_text="Employee Count", 
                        tickfont=dict(size=12),
                        row=2, col=1
                    )
                    
                    # Show monthly detail chart with distribution
                    print(f"📊 Opening {month} detail chart with distribution...")
                    fig_month.show()
            
            print(f"\n✅ Interactive charts opened successfully!")
            print(f"📊 {len(sorted_months)} monthly detail charts displayed")
            print(f"💡 All charts are now open in your browser for interactive exploration!")
            
            return len(sorted_months), 0
            
        except ImportError:
            print("⚠️ Plotly not available - install with: pip install plotly kaleido")
            return None, None
        except Exception as e:
            print(f"❌ Error creating interactive dashboard: {str(e)}")
            return None, None
    
    def _create_excel_safe_sheet_name(self, base_name: str, max_length: int = 31) -> str:
        """
        Create an Excel-safe sheet name that doesn't exceed the character limit.
        
        Args:
            base_name: The desired sheet name
            max_length: Maximum allowed length (default 31 for Excel)
            
        Returns:
            str: Excel-safe sheet name
        """
        # Remove invalid characters for Excel sheet names
        invalid_chars = ['[', ']', '*', '?', ':', '\\', '/']
        safe_name = base_name
        for char in invalid_chars:
            safe_name = safe_name.replace(char, '_')
        
        # Truncate if too long, but preserve meaningful part
        if len(safe_name) > max_length:
            # Try to keep the most important part (usually the month/year)
            if '_' in safe_name:
                parts = safe_name.split('_')
                # Keep the last part (usually the month) and truncate the first part
                if len(parts) > 1:
                    last_part = parts[-1]
                    remaining_length = max_length - len(last_part) - 1  # -1 for underscore
                    if remaining_length > 0:
                        first_part = '_'.join(parts[:-1])[:remaining_length]
                        safe_name = f"{first_part}_{last_part}"
                    else:
                        safe_name = last_part[:max_length]
                else:
                    safe_name = safe_name[:max_length]
            else:
                safe_name = safe_name[:max_length]
        
        return safe_name

    def save_to_excel(self, client_df: pd.DataFrame, resource_df: pd.DataFrame, daily_detailed_df: pd.DataFrame, 
                      employee_stacked_df: pd.DataFrame, distribution_df: pd.DataFrame, top_10_df: pd.DataFrame, 
                      bottom_10_df: pd.DataFrame, stats_dict: dict, monthly_employee_data: dict, filename: str):
        """
        Save the processed data to Excel with core sheets and monthly breakdowns.
        
        Args:
            client_df: Client grouped DataFrame (ignored, can be None)
            resource_df: Resource grouped DataFrame
            daily_detailed_df: Daily detailed DataFrame ordered by Client, Allocation, Employee
            employee_stacked_df: Employee data (not used for sheets, kept for compatibility)
            distribution_df: Statistical distribution data (not used for sheets, kept for compatibility)
            top_10_df: Top 10 employees (not used for sheets, kept for compatibility)
            bottom_10_df: Bottom 10 employees (not used for sheets, kept for compatibility)
            stats_dict: Overall statistics (not used for sheets, kept for compatibility)
            monthly_employee_data: Dictionary of monthly employee DataFrames
            filename: Output Excel filename
        """
        import os
        
        # Update filename to include directory path
        filepath = os.path.join(self.data_dir, filename)
        
        try:
            # First, save the data using pandas/openpyxl
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            wb = Workbook()
            ws_data = wb.active
            ws_data.title = 'Resource_Client_Allocation'
            
            # Write data to the sheet
            for r in dataframe_to_rows(resource_df, index=False, header=True):
                ws_data.append(r)
            
            # Add daily detailed sheets for each month
            daily_detailed_sheets = {}
            for month, month_df in monthly_employee_data.items():
                if not month_df.empty:
                    # Filter daily_detailed_df for this specific month
                    month_daily_df = daily_detailed_df[daily_detailed_df['Month_Name'] == month].copy()
                    
                    if not month_daily_df.empty:
                        # Create Excel-safe sheet name (max 31 characters)
                        base_sheet_name = f'Daily_Detailed_{month.replace(" ", "_")}'
                        sheet_name = self._create_excel_safe_sheet_name(base_sheet_name)
                        ws_daily = wb.create_sheet(sheet_name)
                        
                        # Write daily detailed data for this month
                        for r in dataframe_to_rows(month_daily_df, index=False, header=True):
                            ws_daily.append(r)
                        
                        # Format Date column as proper date format in Excel
                        if 'Date' in month_daily_df.columns:
                            date_col_idx = list(month_daily_df.columns).index('Date') + 1  # +1 for 1-based indexing
                            date_col_letter = ws_daily.cell(row=1, column=date_col_idx).column_letter
                            
                            # Apply date formatting to all data rows (skip header row)
                            for row in range(2, len(month_daily_df) + 2):
                                cell = ws_daily[f'{date_col_letter}{row}']
                                cell.number_format = 'yyyy-mm-dd'
                        
                        # Auto-adjust column widths for better readability
                        for column in ws_daily.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                            ws_daily.column_dimensions[column_letter].width = adjusted_width
                        
                        daily_detailed_sheets[month] = sheet_name
            
            # Note: Employee sheets removed as requested
            
            # Note: Removed All_Months_Combined, Distribution, Top_10_Performers, 
            # Bottom_10_Performers, and Statistics_Summary sheets per user request
            
            wb.save(filepath)
            print(f"✅ Excel file saved: {filepath}")
            
            # Now create native Excel pivot table and histogram chart using pywin32
            excel = None
            workbook = None
            try:
                import win32com.client as win32
                
                print("📊 Creating native Excel pivot table and histogram chart...")
                
                # Open Excel application
                excel = win32.Dispatch('Excel.Application')
                excel.Visible = False  # Keep Excel hidden
                excel.DisplayAlerts = False  # Disable alerts to prevent interruption
                
                # Open the workbook we just created
                workbook = excel.Workbooks.Open(os.path.abspath(filepath))
                
                # Get the data sheet
                data_sheet = workbook.Sheets('Resource_Client_Allocation')
                
                # Define data range (A1 to last used cell)
                last_row = len(resource_df) + 1  # +1 for header
                data_range = data_sheet.Range(f'A1:E{last_row}')
                
                # Create pivot cache
                pivot_cache = workbook.PivotCaches().Create(
                    SourceType=1,  # xlDatabase
                    SourceData=data_range
                )
                
                # Add new sheet for pivot table
                pivot_sheet = workbook.Sheets.Add()
                pivot_sheet.Name = 'Resource_Summary'
                
                # Create pivot table
                pivot_table = pivot_cache.CreatePivotTable(
                    TableDestination=pivot_sheet.Range('A1'),
                    TableName='ResourcePivotTable'
                )
                
                # Configure pivot table fields
                # Rows: Resource_Name, Client_Name, Allocation_Name (in order)
                pivot_table.PivotFields('Resource_Name').Orientation = 1  # xlRowField
                pivot_table.PivotFields('Resource_Name').Position = 1
                
                pivot_table.PivotFields('Client_Name').Orientation = 1  # xlRowField  
                pivot_table.PivotFields('Client_Name').Position = 2
                
                pivot_table.PivotFields('Allocation_Name').Orientation = 1  # xlRowField
                pivot_table.PivotFields('Allocation_Name').Position = 3
                
                # Columns: Month
                pivot_table.PivotFields('Month').Orientation = 2  # xlColumnField
                
                # Values: Sum of Days
                data_field = pivot_table.PivotFields('Days')
                data_field.Orientation = 4  # xlDataField
                data_field.Function = -4157  # xlSum
                data_field.Name = "Sum of Days"
                
                # Set Resource_Name to be collapsed by default (grouped)
                pivot_table.PivotFields('Resource_Name').ShowDetail = False
                
                # Note: Employee charts removed as requested
                
                # Save and close
                workbook.Save()
                workbook.Close()
                excel.Quit()
                
                print(f"✅ Excel analysis created successfully!")
                print(f"📊 Resource_Summary sheet with interactive pivot table")
                print(f"📊 {len(daily_detailed_sheets)} daily detailed sheets with comprehensive daily data:")
                for month, sheet_name in daily_detailed_sheets.items():
                    print(f"   - {sheet_name} ({month})")
                
            except ImportError:
                print("⚠️ pywin32 not available - pivot table and chart will be manual")
            except Exception as e_pivot:
                print(f"⚠️ Error creating native pivot table/chart: {str(e_pivot)}")
                print("📊 File saved with data sheets only")
                
                # Clean up Excel resources if they were created
                try:
                    if workbook:
                        workbook.Close(SaveChanges=False)  # Don't save if there was an error
                    if excel:
                        excel.Quit()
                except Exception as cleanup_error:
                    print(f"⚠️ Error during cleanup: {cleanup_error}")
            finally:
                # Ensure Excel is properly closed even if an exception occurs
                try:
                    if workbook and not workbook.Saved:
                        workbook.Close(SaveChanges=False)
                    if excel:
                        excel.Quit()
                except:
                    pass  # Ignore cleanup errors
                
        except Exception as e:
            print(f"❌ Error saving Excel file: {str(e)}")
            return False
    
    def extract_timesheets(self, start_date: str, end_date: str, output_filename: Optional[str] = None) -> bool:
        """
        Main method to extract timesheets for a date range.
        
        Args:
            start_date: Start date in YYYY-MM-DD format
            end_date: End date in YYYY-MM-DD format
            output_filename: Optional custom filename
            
        Returns:
            bool: True if extraction successful
        """
        try:
            print(f"🚀 Starting timesheet extraction from {start_date} to {end_date}")
            
            # Archive existing files first
            print("\n🗂️ CHECKING FOR EXISTING FILES TO ARCHIVE")
            print("-" * 40)
            archived_count = self.archive_existing_files()
            
            if archived_count == 0:
                print("✅ No existing files found to archive")
            else:
                print(f"✅ Archived {archived_count} existing files")
            
            # Fetch timesheet data from API
            timesheet_records = self.fetch_timesheet_records(start_date, end_date)
            
            # Fetch vacation data from API
            print("\n🏖️ Fetching vacation records...")
            vacation_records = self.fetch_vacation_records(start_date, end_date)
            
            # Fetch allocation data from API
            print("\n📋 Fetching allocation records...")
            allocation_records = self.fetch_allocations(start_date, end_date)
            
            if not timesheet_records and not vacation_records and not allocation_records:
                print("❌ No timesheet, vacation, or allocation records found for the specified date range")
                return False
            
            # Process data
            print("\n📊 Processing combined timesheet, vacation, and allocation data...")
            _, resource_df, daily_detailed_df = self.process_timesheet_data(timesheet_records, vacation_records, allocation_records, start_date, end_date)
            
            if resource_df.empty:
                print("❌ No data to export after processing")
                return False
            
            # Create histogram data
            print("\n📊 Creating histogram analysis...")
            employee_stacked_df, distribution_df, top_10_df, bottom_10_df, stats_dict, monthly_employee_data = self.create_histogram_data(resource_df)
            
            # Generate filename if not provided
            if not output_filename:
                timestamp = datetime.now().strftime('%H%M%S')
                output_filename = self.create_archive_filename("timesheets", start_date, end_date, timestamp)
            
            # Save to Excel
            print("💾 Saving to Excel...")
            self.save_to_excel(None, resource_df, daily_detailed_df, employee_stacked_df, distribution_df, top_10_df, bottom_10_df, stats_dict, monthly_employee_data, output_filename)
            
            # Create interactive dashboard
            print("🚀 Creating interactive charts...")
            monthly_charts, _ = self.create_interactive_dashboard(resource_df, monthly_employee_data, stats_dict, output_filename)
            
            print(f"🎉 Timesheet extraction completed successfully!")
            print(f"📁 Excel output: {os.path.join(self.data_dir, output_filename)}")
            print(f"📈 Resource-grouped sheet: {len(resource_df)} rows")
            if monthly_charts:
                print(f"📊 Interactive charts: {monthly_charts} monthly charts + overview dashboard")
                print(f"💡 Charts are now open in your browser for interactive exploration!")
            
            return True
            
        except Exception as e:
            print(f"❌ Error during timesheet extraction: {str(e)}")
            return False


def get_user_input_dates() -> Tuple[str, str]:
    """Get user input for date range only (authentication uses config)."""
    
    print("=" * 60)
    print("🕒 ElapseIT Timesheet Extractor")
    print("=" * 60)
    print(f"🔧 Using config: {ELAPSEIT_CONFIG['domain']}")
    
    print("\n📅 Date Range Selection")
    print("-" * 30)
    
    # Get date range
    while True:
        try:
            start_date = input("Enter start date (YYYY-MM-DD): ").strip()
            datetime.strptime(start_date, '%Y-%m-%d')
            break
        except ValueError:
            print("❌ Invalid date format. Please use YYYY-MM-DD")
    
    while True:
        try:
            end_date = input("Enter end date (YYYY-MM-DD): ").strip()
            end_dt = datetime.strptime(end_date, '%Y-%m-%d')
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            
            if end_dt < start_dt:
                print("❌ End date must be after start date")
                continue
            break
        except ValueError:
            print("❌ Invalid date format. Please use YYYY-MM-DD")
    
    return start_date, end_date


def validate_date(date_string: str) -> str:
    """Validate date format and return the date string."""
    try:
        datetime.strptime(date_string, '%Y-%m-%d')
        return date_string
    except ValueError:
        raise argparse.ArgumentTypeError(f"Invalid date format: {date_string}. Use YYYY-MM-DD")


def main():
    """Main function to run the timesheet extractor."""
    
    parser = argparse.ArgumentParser(
        description="ElapseIT Timesheet Extractor - Extract timesheet data for a date range",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Interactive mode (prompts for all inputs)
  python timesheet_extractor.py
  
  # Command-line mode with parameters (uses config for authentication)
  python timesheet_extractor.py --start-date 2024-01-01 --end-date 2024-03-31
  
  # Custom output filename
  python timesheet_extractor.py --start-date 2024-01-01 --end-date 2024-03-31 --output custom_report.xlsx
  
  # Short form parameters
  python timesheet_extractor.py -s 2024-01-01 -e 2024-03-31 -o Q1_report.xlsx
        """
    )
    
    parser.add_argument('--start-date', '-s', type=validate_date,
                        help='Start date for extraction (YYYY-MM-DD)')
    parser.add_argument('--end-date', '-e', type=validate_date,
                        help='End date for extraction (YYYY-MM-DD)')
    parser.add_argument('--output', '-o', type=str,
                        help='Custom output filename (optional)')
    parser.add_argument('--interactive', '-i', action='store_true',
                        help='Force interactive mode (ignore other parameters)')
    
    args = parser.parse_args()
    
    try:
        # Determine if we should use interactive mode
        use_interactive = args.interactive or not (args.start_date and args.end_date)
        
        if use_interactive:
            # Interactive mode
            start_date, end_date = get_user_input_dates()
            output_filename = args.output
        else:
            # Command-line mode
            start_date = args.start_date
            end_date = args.end_date
            output_filename = args.output
            
            # Validate date range
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            end_dt = datetime.strptime(end_date, '%Y-%m-%d')
            
            if end_dt < start_dt:
                print("❌ Error: End date must be after start date")
                return
            
            print("=" * 60)
            print("🕒 ElapseIT Timesheet Extractor")
            print("=" * 60)
            print(f"📅 Date Range: {start_date} to {end_date}")
            print(f"🔧 Using config: {ELAPSEIT_CONFIG['domain']}")
            if output_filename:
                print(f"📄 Output File: {output_filename}")
        
        # Create extractor instance
        extractor = ElapseITTimesheetExtractor()
        
        # Authenticate
        if not extractor.authenticate():
            print("❌ Authentication failed. Please check your credentials.")
            return
        
        # Extract timesheets
        print(f"\n📊 Extracting timesheets from {start_date} to {end_date}...")
        success = extractor.extract_timesheets(start_date, end_date, output_filename)
        
        if success:
            print("\n✅ Extraction completed successfully!")
        else:
            print("\n❌ Extraction failed!")
            
    except KeyboardInterrupt:
        print("\n\n⚠️ Operation cancelled by user")
    except Exception as e:
        print(f"\n❌ Unexpected error: {str(e)}")


if __name__ == "__main__":
    main()
